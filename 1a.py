# 1a.py
# -*- coding: utf-8 -*-

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import os
import json
from persiantools.jdatetime import JalaliDate
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from PIL import Image, ImageTk
import traceback
import warnings
import logging
import re
import numpy as np
from datetime import datetime

# تنظیمات لاگینگ
logging.basicConfig(
    filename='app_errors.log',
    level=logging.ERROR,
    format='%(asctime)s - %(levelname)s - %(message)s',
    encoding='utf-8'
)

warnings.simplefilter("ignore", UserWarning)

# بررسی وجود کتابخانه‌های گرافیکی
try:
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    from matplotlib.figure import Figure
    import seaborn as sns
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False
    print("⚠️ کتابخانه‌های گرافیکی نصب نیستند. نمودارها غیرفعال خواهند بود.")

# -----------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SETTINGS_PATH = os.path.join(BASE_DIR, "settings.json")


# -----------------------------
def register_persian_fonts():
    """ثبت فونت‌های فارسی برای استفاده در PDF"""
    try:
        font_paths = [
            "C:/Windows/Fonts/arial.ttf",
            "C:/Windows/Fonts/tahoma.ttf",
            "C:/Windows/Fonts/times.ttf",
            "/usr/share/fonts/truetype/freefont/FreeSans.ttf",
            "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
            "/Library/Fonts/Arial.ttf",
            "/System/Library/Fonts/Tahoma.ttf"
        ]

        for font_path in font_paths:
            if os.path.exists(font_path):
                try:
                    pdfmetrics.registerFont(TTFont('PersianFont', font_path))
                    return 'PersianFont'
                except Exception as e:
                    logging.error(f"Error registering font from {font_path}: {e}")
                    continue

        return 'Helvetica'
    except Exception as e:
        logging.error(f"Error registering Persian fonts: {e}")
        return 'Helvetica'


# -----------------------------
def load_settings():
    """لود کردن تنظیمات از فایل JSON"""
    default = {
        "logo_path": "",
        "last_excel_path": "",
        "last_sheet": "",
        "window_size": "1200x800",
        "filters": {
            "start_date": "",
            "end_date": "",
            "repair_type": "",
            "part_type": ""
        },
        "colors": {
            "bg_main": "#FFA500",
            "frame_bg": "#FFE5B4",
            "button_bg": "#FF8C00",
            "button_fg": "#FFFFFF",
            "tree_bg": "#FFFFFF",
            "tree_alt_bg": "#FFF5E0",
            "tree_font_color": "#000000",
            "tree_total_bg": "#0000FF",
            "tree_total_fg": "#FFFFFF"
        }
    }
    try:
        if not os.path.exists(SETTINGS_PATH):
            with open(SETTINGS_PATH, "w", encoding="utf-8") as f:
                json.dump(default, f, ensure_ascii=False, indent=4)
            return default
        with open(SETTINGS_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
        # تکمیل کلیدهای گم‌شده
        for k, v in default.items():
            if k not in data:
                data[k] = v
        if "filters" not in data or not isinstance(data["filters"], dict):
            data["filters"] = default["filters"]
        if "colors" not in data or not isinstance(data["colors"], dict):
            data["colors"] = default["colors"]
        return data
    except Exception as e:
        logging.error(f"Error loading settings: {e}")
        traceback.print_exc()
        return default


def save_settings(data):
    """ذخیره تنظیمات در فایل کنار اسکریپت"""
    try:
        with open(SETTINGS_PATH, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
    except Exception as e:
        logging.error(f"Error saving settings: {e}")
        traceback.print_exc()


# -----------------------------
def find_column(columns, possible_names):
    """
    پیدا کردن ستون در لیست ستون‌ها بر اساس نام‌های احتمالی
    """
    for name in possible_names:
        for col in columns:
            if name.strip().lower() in str(col).strip().lower():
                return col
    return None


# -----------------------------
def normalize_repair_type(repair_type):
    """نرمالایز کردن نوع تعمیر برای تطبیق بهتر"""
    if not isinstance(repair_type, str):
        return str(repair_type)

    repair_type = repair_type.strip()
    repair_type = re.sub(r'[:]', '', repair_type)
    repair_type = re.sub(r'\s+', ' ', repair_type)

    if 'قالب' in repair_type and 'تعمیر' in repair_type:
        return 'قالب تعمیری'
    elif 'قطعه' in repair_type and 'تعمیر' in repair_type:
        return 'قطعه تعمیری'
    elif 'دستگاه' in repair_type and 'تعمیر' in repair_type:
        return 'دستگاه تعمیری'
    elif 'قالب' in repair_type:
        return 'قالب'
    elif 'قطعه' in repair_type:
        return 'قطعه'
    elif 'دستگاه' in repair_type:
        return 'دستگاه'
    elif 'تعمیر' in repair_type:
        return 'تعمیری'
    else:
        return repair_type


# -----------------------------
class PowerBIDashboard:
    """
    داشبورد نمایشی شبیه Power BI
    - از self.main_app.df به عنوان منبع داده استفاده می‌کند.
    - اگر فیلترهای داخلی داشبورد اعمال شوند، روی self.filtered_df بازتاب داده می‌شود.
    - ویژوال‌های پیش‌فرض (bar, pie, line, summary) با داده‌ی فیلترشده به‌روزرسانی می‌شوند
      بدون این‌که ویژوال‌های سفارشی کاربر حذف شوند.
    """

    def __init__(self, parent, main_app):
        self.parent = parent
        self.main_app = main_app

        # DataFrame فیلترشده داخلی داشبورد
        self.filtered_df = None
        self.current_filters = {}
        self.visuals = []

        self.setup_ui()

    # ========================= UI Setup =========================
    def setup_ui(self):
        self.parent.title("Power BI Dashboard - گزارش‌گیری قالب‌سازی")
        self.parent.geometry("1400x900")
        self.create_main_layout()
        self.populate_fields_panel()
        self.create_default_visuals()

    def create_main_layout(self):
        self.create_toolbar()

        main_container = ttk.PanedWindow(self.parent, orient=tk.HORIZONTAL)
        main_container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        self.left_panel = ttk.Frame(main_container, width=300)
        main_container.add(self.left_panel, weight=0)

        self.center_panel = ttk.Frame(main_container)
        main_container.add(self.center_panel, weight=1)

        self.right_panel = ttk.Frame(main_container, width=250)
        main_container.add(self.right_panel, weight=0)

        self.create_left_panel_content()
        self.create_center_panel_content()
        self.create_right_panel_content()

    def create_toolbar(self):
        toolbar = ttk.Frame(self.parent, height=40)
        toolbar.pack(fill=tk.X, padx=5, pady=2)

        buttons = [
            ("📊 نمودار جدید", self.add_chart),
            ("📋 جدول جدید", self.add_table),
            ("🔍 فیلتر جدید", self.add_filter),
            ("💾 ذخیره گزارش", self.save_report),
            ("📂 بارگذاری گزارش", self.load_report),
            ("📤 خروجی PDF", self.export_pdf),
            ("🔄 بروزرسانی داده", self.refresh_data)
        ]

        for text, command in buttons:
            ttk.Button(toolbar, text=text, command=command).pack(side=tk.LEFT, padx=2)

        self.status_label = ttk.Label(toolbar, text="حالت طراحی فعال")
        self.status_label.pack(side=tk.RIGHT, padx=10)

        if not MATPLOTLIB_AVAILABLE:
            warning_label = ttk.Label(
                toolbar,
                text="⚠️ نمودارها غیرفعال - کتابخانه‌ها نصب نیستند",
                foreground="red"
            )
            warning_label.pack(side=tk.RIGHT, padx=10)

    def create_left_panel_content(self):
        notebook = ttk.Notebook(self.left_panel)
        notebook.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        fields_frame = ttk.Frame(notebook)
        notebook.add(fields_frame, text="فیلدها")

        fields_tree_frame = ttk.Frame(fields_frame)
        fields_tree_frame.pack(fill=tk.BOTH, expand=True, pady=5)

        ttk.Label(fields_tree_frame, text="فیلدهای موجود:", font=('Arial', 10, 'bold')).pack(anchor='w')

        self.fields_tree = ttk.Treeview(fields_tree_frame, show="tree", height=15)
        self.fields_tree.pack(fill=tk.BOTH, expand=True, pady=5)

        filters_frame = ttk.Frame(notebook)
        notebook.add(filters_frame, text="فیلترها")

        self.filters_container = ttk.Frame(filters_frame)
        self.filters_container.pack(fill=tk.BOTH, expand=True)

        ttk.Button(filters_frame, text="+ افزودن فیلتر", command=self.add_filter_dialog).pack(pady=10)

    def create_center_panel_content(self):
        vis_toolbar = ttk.Frame(self.center_panel)
        vis_toolbar.pack(fill=tk.X, pady=2)
        ttk.Label(vis_toolbar, text="ویژوال‌های گزارش", font=('Arial', 12, 'bold')).pack(side=tk.LEFT)

        self.canvas_frame = ttk.Frame(self.center_panel)
        self.canvas_frame.pack(fill=tk.BOTH, expand=True)

        self.grid_frame = ttk.Frame(self.canvas_frame)
        self.grid_frame.pack(fill=tk.BOTH, expand=True)

    def create_right_panel_content(self):
        notebook = ttk.Notebook(self.right_panel)
        notebook.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        visual_settings_frame = ttk.Frame(notebook)
        notebook.add(visual_settings_frame, text="تنظیمات بصری")

        self.setup_visual_settings(visual_settings_frame)

    # ========================= Visuals =========================
    def get_base_df(self):
        """دیتافریم پایه برای نمودارها: اگر فیلتر داشبورد هست از آن، وگرنه df برنامه اصلی"""
        if self.filtered_df is not None:
            return self.filtered_df
        return self.main_app.df

    def create_default_visuals(self):
        """ایجاد ویژوال‌های پیش‌فرض فقط یک بار"""
        if self.main_app.df is None:
            ttk.Label(
                self.grid_frame,
                text="لطفاً ابتدا داده‌ها را از برنامه اصلی بارگذاری کنید",
                font=('Arial', 14),
                foreground='red'
            ).pack(expand=True)
            return

        # ما فقط یک‌بار ویژوال‌های پیش‌فرض را می‌سازیم. بعداً فقط داده‌شان را به‌روز می‌کنیم.
        # برای سادگی، ویژوال‌ها را در لیست نگه می‌داریم تا در refresh از آن‌ها استفاده کنیم.

        # نمودار میله‌ای انواع تعمیر
        if MATPLOTLIB_AVAILABLE:
            self.bar_chart_frame = ttk.LabelFrame(self.grid_frame, text="توزیع انواع تعمیر", width=400, height=300)
            self.bar_chart_frame.grid(row=0, column=0, padx=5, pady=5, sticky='nsew')
            self.bar_chart_frame.grid_propagate(False)
            self.create_bar_chart()
        else:
            self.create_text_visual("نمودار میله‌ای", "برای نمایش نمودارها، کتابخانه matplotlib را نصب کنید", 0, 0)

        # نمودار دایره‌ای ساعات
        if MATPLOTLIB_AVAILABLE:
            self.pie_chart_frame = ttk.LabelFrame(self.grid_frame, text="توزیع ساعت کاری", width=400, height=300)
            self.pie_chart_frame.grid(row=0, column=1, padx=5, pady=5, sticky='nsew')
            self.pie_chart_frame.grid_propagate(False)
            self.create_pie_chart()
        else:
            self.create_text_visual("نمودار دایره‌ای", "برای نمایش نمودارها، کتابخانه matplotlib را نصب کنید", 0, 1)

        # جدول خلاصه
        self.summary_frame = ttk.LabelFrame(self.grid_frame, text="خلاصه آماری", width=400, height=300)
        self.summary_frame.grid(row=1, column=0, padx=5, pady=5, sticky='nsew')
        self.summary_frame.grid_propagate(False)
        self.create_summary_table()

        # نمودار خطی
        if MATPLOTLIB_AVAILABLE:
            self.line_chart_frame = ttk.LabelFrame(self.grid_frame, text="روند ساعت کاری", width=400, height=300)
            self.line_chart_frame.grid(row=1, column=1, padx=5, pady=5, sticky='nsew')
            self.line_chart_frame.grid_propagate(False)
            self.create_line_chart()
        else:
            self.create_text_visual("نمودار خطی", "برای نمایش نمودارها، کتابخانه matplotlib را نصب کنید", 1, 1)

    def refresh_default_visuals(self):
        """به‌روزرسانی داده‌ی ویژوال‌های پیش‌فرض بدون تخریب فریم‌ها"""
        if self.main_app.df is None:
            return

        if MATPLOTLIB_AVAILABLE:
            # پاک‌کردن محتوای قدیمی فریم‌ها
            for frame in [getattr(self, 'bar_chart_frame', None),
                          getattr(self, 'pie_chart_frame', None),
                          getattr(self, 'line_chart_frame', None),
                          getattr(self, 'summary_frame', None)]:
                if frame is not None:
                    for child in frame.winfo_children():
                        child.destroy()

            self.create_bar_chart()
            self.create_pie_chart()
            self.create_line_chart()
        # خلاصه آماری
        for child in self.summary_frame.winfo_children():
            child.destroy()
        self.create_summary_table()

    def create_text_visual(self, title, message, row, col):
        frame = ttk.LabelFrame(self.grid_frame, text=title, width=400, height=300)
        frame.grid(row=row, column=col, padx=5, pady=5, sticky='nsew')
        frame.grid_propagate(False)

        ttk.Label(
            frame,
            text=message,
            font=('Arial', 10),
            foreground='red',
            wraplength=350
        ).pack(expand=True)

        ttk.Button(frame, text="نصب کتابخانه‌ها", command=self.show_install_instructions).pack(pady=10)

        self.visuals.append(('text_visual', frame))

    def show_install_instructions(self):
        instructions = """
برای استفاده از قابلیت‌های نموداری Power BI، لطفاً کتابخانه‌های زیر را نصب کنید:

pip install matplotlib seaborn numpy

پس از نصب، برنامه را مجدداً راه‌اندازی کنید.
        """
        messagebox.showinfo("راهنمای نصب", instructions)

    def create_bar_chart(self):
        if not MATPLOTLIB_AVAILABLE or self.main_app.df is None:
            return

        frame = self.bar_chart_frame
        try:
            df = self.get_base_df()
            if df is None or df.empty:
                ttk.Label(frame, text="هیچ داده‌ای موجود نیست").pack(expand=True)
                return

            if self.main_app.repair_col in df.columns:
                repair_counts = df[self.main_app.repair_col].value_counts().head(10)

                fig = Figure(figsize=(4, 3), dpi=100)
                ax = fig.add_subplot(111)

                bars = ax.bar(range(len(repair_counts)), repair_counts.values, color='skyblue')
                ax.set_title('توزیع انواع تعمیر', fontsize=12)
                ax.set_xticks(range(len(repair_counts)))
                ax.set_xticklabels(repair_counts.index, rotation=45, ha='right')

                for bar in bars:
                    height = bar.get_height()
                    ax.text(bar.get_x() + bar.get_width() / 2., height,
                            f'{int(height)}', ha='center', va='bottom')

                canvas = FigureCanvasTkAgg(fig, frame)
                canvas.draw()
                canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

                self.visuals.append(('bar_chart', frame))
            else:
                ttk.Label(frame, text="ستون نوع تعمیر یافت نشد").pack(expand=True)

        except Exception as e:
            logging.error(f"Error creating bar chart: {e}")
            ttk.Label(frame, text=f"خطا در ایجاد نمودار: {e}").pack()

    def create_pie_chart(self):
        if not MATPLOTLIB_AVAILABLE or self.main_app.df is None:
            return

        frame = self.pie_chart_frame
        try:
            df = self.get_base_df()
            if df is None or df.empty:
                ttk.Label(frame, text="هیچ داده‌ای موجود نیست").pack(expand=True)
                return

            if (self.main_app.repair_col in df.columns and
                    self.main_app.perf_col in df.columns):

                df = df.copy()
                df[self.main_app.perf_col] = pd.to_numeric(df[self.main_app.perf_col], errors='coerce')

                grouped = df.groupby(self.main_app.repair_col)[self.main_app.perf_col].sum()
                grouped = grouped[grouped > 0].head(6)

                if grouped.empty:
                    ttk.Label(frame, text="داده‌ی معتبری برای ساعت کاری یافت نشد").pack(expand=True)
                    return

                fig = Figure(figsize=(4, 3), dpi=100)
                ax = fig.add_subplot(111)

                colors = plt.cm.Set3(np.linspace(0, 1, len(grouped)))
                ax.pie(grouped.values,
                       labels=grouped.index,
                       autopct='%1.1f%%',
                       colors=colors,
                       startangle=90)

                ax.set_title('توزیع ساعت کاری بر اساس نوع تعمیر', fontsize=10)

                canvas = FigureCanvasTkAgg(fig, frame)
                canvas.draw()
                canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

                self.visuals.append(('pie_chart', frame))
            else:
                ttk.Label(frame, text="ستون‌های لازم برای نمودار دایره‌ای یافت نشد").pack(expand=True)

        except Exception as e:
            logging.error(f"Error creating pie chart: {e}")
            ttk.Label(frame, text=f"خطا در ایجاد نمودار: {e}").pack()

    def create_summary_table(self):
        frame = self.summary_frame
        try:
            df = self.get_base_df()
            if df is None or df.empty:
                ttk.Label(frame, text="هیچ داده‌ای موجود نیست").pack(expand=True)
                return

            columns = ("معیار", "مقدار")
            tree = ttk.Treeview(frame, columns=columns, show="headings", height=8)

            for col in columns:
                tree.heading(col, text=col)
                tree.column(col, width=150)

            stats = [("تعداد رکوردها", len(df))]

            if self.main_app.perf_col in df.columns:
                df[self.main_app.perf_col] = pd.to_numeric(df[self.main_app.perf_col], errors='coerce')
                stats.extend([
                    ("مجموع ساعت کاری", f"{df[self.main_app.perf_col].sum():.2f}"),
                    ("میانگین ساعت کاری", f"{df[self.main_app.perf_col].mean():.2f}"),
                    ("بیشترین ساعت کاری", f"{df[self.main_app.perf_col].max():.2f}"),
                    ("کمترین ساعت کاری", f"{df[self.main_app.perf_col].min():.2f}")
                ])

            if self.main_app.repair_col in df.columns:
                unique_repairs = df[self.main_app.repair_col].nunique()
                stats.append(("انواع تعمیر منحصر بفرد", unique_repairs))

            for stat in stats:
                tree.insert("", "end", values=stat)

            scrollbar = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
            tree.configure(yscrollcommand=scrollbar.set)

            tree.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")

            self.visuals.append(('summary_table', frame))

        except Exception as e:
            logging.error(f"Error creating summary table: {e}")
            ttk.Label(frame, text=f"خطا در ایجاد جدول: {e}").pack()

    def create_line_chart(self):
        if not MATPLOTLIB_AVAILABLE or self.main_app.df is None:
            return

        frame = self.line_chart_frame
        try:
            df = self.get_base_df()
            if df is None or df.empty:
                ttk.Label(frame, text="هیچ داده‌ای موجود نیست").pack(expand=True)
                return

            if (self.main_app.date_col in df.columns and
                    self.main_app.perf_col in df.columns):

                df = df.copy()
                df[self.main_app.perf_col] = pd.to_numeric(df[self.main_app.perf_col], errors='coerce')
                df[self.main_app.date_col] = pd.to_datetime(df[self.main_app.date_col], errors='coerce')

                daily_hours = df.groupby(df[self.main_app.date_col].dt.date)[self.main_app.perf_col].sum()
                daily_hours = daily_hours.sort_index().tail(30)

                if daily_hours.empty:
                    ttk.Label(frame, text="داده‌ی معتبری برای نمودار زمانی یافت نشد").pack(expand=True)
                    return

                fig = Figure(figsize=(4, 3), dpi=100)
                ax = fig.add_subplot(111)

                ax.plot(range(len(daily_hours)), daily_hours.values, marker='o', linewidth=2, color='green')
                ax.set_title('روند ساعت کاری روزانه', fontsize=12)
                ax.set_xticks(range(len(daily_hours)))
                ax.set_xticklabels([d.strftime('%m/%d') for d in daily_hours.index], rotation=45)
                ax.grid(True, alpha=0.3)

                canvas = FigureCanvasTkAgg(fig, frame)
                canvas.draw()
                canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

                self.visuals.append(('line_chart', frame))
            else:
                ttk.Label(frame, text="ستون‌های لازم برای نمودار خطی یافت نشد").pack(expand=True)

        except Exception as e:
            logging.error(f"Error creating line chart: {e}")
            ttk.Label(frame, text=f"خطا در ایجاد نمودار: {e}").pack()

    # ========================= Right Panel Settings =========================
    def setup_visual_settings(self, parent):
        ttk.Label(parent, text="نوع نمودار:", font=('Arial', 9, 'bold')).pack(anchor='w', pady=(10, 2))

        self.chart_type = ttk.Combobox(
            parent,
            values=["نمودار میله‌ای", "نمودار خطی", "نمودار دایره‌ای"],
            state="readonly"
        )
        self.chart_type.pack(fill=tk.X, pady=2)
        self.chart_type.set("نمودار میله‌ای")

        if not MATPLOTLIB_AVAILABLE:
            self.chart_type.config(state="disabled")
            ttk.Label(
                parent,
                text="⚠️ برای ایجاد نمودار، کتابخانه‌ها را نصب کنید",
                foreground="red"
            ).pack(pady=5)

        ttk.Button(parent, text="ایجاد نمودار", command=self.create_custom_chart).pack(pady=10)

    def populate_fields_panel(self):
        if self.main_app.df is None:
            return

        for item in self.fields_tree.get_children():
            self.fields_tree.delete(item)

        df = self.main_app.df

        if MATPLOTLIB_AVAILABLE:
            numeric_fields = df.select_dtypes(include=[np.number]).columns.tolist()
        else:
            numeric_fields = []

        if numeric_fields:
            numeric_node = self.fields_tree.insert("", "end", text="فیلدهای عددی", values=("numeric",))
            for field in numeric_fields:
                self.fields_tree.insert(numeric_node, "end", text=field, values=("field", field, "numeric"))

        text_fields = df.select_dtypes(include=['object']).columns.tolist()
        if text_fields:
            text_node = self.fields_tree.insert("", "end", text="فیلدهای متنی", values=("text",))
            for field in text_fields:
                self.fields_tree.insert(text_node, "end", text=field, values=("field", field, "text"))

    # ========================= Dashboard Actions =========================
    def add_chart(self):
        if not MATPLOTLIB_AVAILABLE:
            messagebox.showwarning(
                "هشدار",
                "برای ایجاد نمودار، لطفاً کتابخانه‌های زیر را نصب کنید:\n\npip install matplotlib seaborn numpy"
            )
            return
        messagebox.showinfo("افزودن نمودار", "از پنل سمت راست برای ایجاد نمودارهای جدید استفاده کنید")

    def add_table(self):
        self.create_data_table()

    def add_filter(self):
        self.add_filter_dialog()

    def add_filter_dialog(self):
        dialog = tk.Toplevel(self.parent)
        dialog.title("افزودن فیلتر جدید")
        dialog.geometry("400x300")
        dialog.transient(self.parent)
        dialog.grab_set()

        ttk.Label(dialog, text="انتخاب فیلد برای فیلتر:", font=('Arial', 10, 'bold')).pack(pady=10)

        field_var = tk.StringVar()
        fields_combobox = ttk.Combobox(dialog, textvariable=field_var, state="readonly")

        if self.main_app.df is not None:
            fields_combobox['values'] = self.main_app.df.columns.tolist()
        fields_combobox.pack(fill=tk.X, padx=20, pady=5)

        ttk.Label(dialog, text="مقادیر فیلتر:", font=('Arial', 10, 'bold')).pack(pady=10)

        values_listbox = tk.Listbox(dialog, selectmode=tk.MULTIPLE, height=8)
        values_listbox.pack(fill=tk.BOTH, expand=True, padx=20, pady=5)

        def on_field_selected(event):
            field = field_var.get()
            if field and self.main_app.df is not None:
                values_listbox.delete(0, tk.END)
                unique_values = self.main_app.df[field].dropna().unique()[:50]
                for value in unique_values:
                    values_listbox.insert(tk.END, str(value))

        fields_combobox.bind('<<ComboboxSelected>>', on_field_selected)

        def apply_filter():
            field = field_var.get()
            selected_indices = values_listbox.curselection()
            selected_values = [values_listbox.get(i) for i in selected_indices]

            if field and selected_values:
                filter_key = f"{field}_filter"
                self.current_filters[filter_key] = {
                    'field': field,
                    'values': selected_values,
                    'type': 'multi_select'
                }
                self.apply_filters_to_visuals()
                dialog.destroy()
            else:
                messagebox.showwarning("هشدار", "لطفاً فیلد و مقادیر را انتخاب کنید")

        ttk.Button(dialog, text="اعمال فیلتر", command=apply_filter).pack(pady=10)

    def apply_filters_to_visuals(self):
        """اعمال فیلترهای داشبورد روی df اصلی و به‌روزرسانی ویژوال‌ها"""
        if self.main_app.df is None:
            return

        df = self.main_app.df.copy()

        for f in self.current_filters.values():
            if f['type'] == 'multi_select':
                field = f['field']
                values = f['values']
                if field in df.columns:
                    df = df[df[field].astype(str).isin(values)]

        self.filtered_df = df
        self.refresh_default_visuals()
        self.status_label.config(text="فیلترها اعمال شدند")

    def create_custom_chart(self):
        if not MATPLOTLIB_AVAILABLE:
            messagebox.showwarning(
                "هشدار",
                "برای ایجاد نمودار، لطفاً کتابخانه‌های زیر را نصب کنید:\n\npip install matplotlib seaborn numpy"
            )
            return

        if self.main_app.df is None:
            messagebox.showwarning("هشدار", "ابتدا داده‌ها را بارگذاری کنید.")
            return

        chart_type = self.chart_type.get()
        title = f"نمودار {chart_type}"

        frame = ttk.LabelFrame(self.grid_frame, text=title, width=400, height=300)
        row, col = self.find_empty_grid_position()
        frame.grid(row=row, column=col, padx=5, pady=5, sticky='nsew')
        frame.grid_propagate(False)

        try:
            df = self.get_base_df().copy()
            if df is None or df.empty:
                ttk.Label(frame, text="هیچ داده‌ای وجود ندارد").pack(expand=True)
                return

            if chart_type == "نمودار میله‌ای":
                self.create_custom_bar_chart(frame, df, title)
            elif chart_type == "نمودار خطی":
                self.create_custom_line_chart(frame, df, title)
            elif chart_type == "نمودار دایره‌ای":
                self.create_custom_pie_chart(frame, df, title)

            self.visuals.append(('custom_chart', frame))

        except Exception as e:
            logging.error(f"Error creating custom chart: {e}")
            ttk.Label(frame, text=f"خطا در ایجاد نمودار: {e}").pack()

    def find_empty_grid_position(self):
        for row in range(3):
            for col in range(3):
                exists = False
                for child in self.grid_frame.grid_slaves():
                    info = child.grid_info()
                    if info.get('row') == row and info.get('column') == col:
                        exists = True
                        break
                if not exists:
                    return row, col
        return 2, 2

    def create_custom_bar_chart(self, frame, df, title):
        fig = Figure(figsize=(4, 3), dpi=100)
        ax = fig.add_subplot(111)

        if self.main_app.repair_col in df.columns:
            grouped = df[self.main_app.repair_col].value_counts().head(10)
        else:
            grouped = df.iloc[:, 0].value_counts().head(10)

        bars = ax.bar(range(len(grouped)), grouped.values, color='lightblue')
        ax.set_title(title, fontsize=10)
        ax.set_xticks(range(len(grouped)))
        ax.set_xticklabels(grouped.index, rotation=45, ha='right')

        for bar in bars:
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width() / 2., height,
                    f'{int(height)}', ha='center', va='bottom')

        canvas = FigureCanvasTkAgg(fig, frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

    def create_custom_line_chart(self, frame, df, title):
        fig = Figure(figsize=(4, 3), dpi=100)
        ax = fig.add_subplot(111)

        if self.main_app.perf_col in df.columns:
            df[self.main_app.perf_col] = pd.to_numeric(df[self.main_app.perf_col], errors='coerce')
            values = df[self.main_app.perf_col].head(15).dropna().values
            if len(values) == 0:
                ttk.Label(frame, text="داده‌ی عددی برای نمودار خطی یافت نشد").pack(expand=True)
                return
            ax.plot(range(len(values)), values, marker='o', linewidth=2)
        else:
            ttk.Label(frame, text="ستون عددی برای نمودار خطی یافت نشد").pack(expand=True)
            return

        ax.set_title(title, fontsize=10)
        ax.grid(True, alpha=0.3)

        canvas = FigureCanvasTkAgg(fig, frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

    def create_custom_pie_chart(self, frame, df, title):
        fig = Figure(figsize=(4, 3), dpi=100)
        ax = fig.add_subplot(111)

        if self.main_app.repair_col in df.columns:
            grouped = df[self.main_app.repair_col].value_counts().head(6)
        else:
            grouped = df.iloc[:, 0].value_counts().head(6)

        if grouped.empty:
            ttk.Label(frame, text="داده‌ای برای نمودار دایره‌ای یافت نشد").pack(expand=True)
            return

        colors = plt.cm.Pastel1(np.linspace(0, 1, len(grouped)))
        ax.pie(grouped.values, labels=grouped.index, autopct='%1.1f%%', colors=colors)

        ax.set_title(title, fontsize=10)

        canvas = FigureCanvasTkAgg(fig, frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

    def create_data_table(self):
        if self.main_app.df is None:
            messagebox.showwarning("هشدار", "هیچ داده‌ای برای نمایش وجود ندارد")
            return

        frame = ttk.LabelFrame(self.grid_frame, text="جدول داده‌ها", width=600, height=400)
        row, col = self.find_empty_grid_position()
        frame.grid(row=row, column=col, padx=5, pady=5, sticky='nsew', columnspan=2)
        frame.grid_propagate(False)

        df = self.get_base_df()
        if df is None or df.empty:
            ttk.Label(frame, text="هیچ داده‌ای موجود نیست").pack(expand=True)
            return

        columns = df.columns.tolist()[:6]
        tree = ttk.Treeview(frame, columns=columns, show="headings", height=15)

        for col_name in columns:
            tree.heading(col_name, text=col_name)
            tree.column(col_name, width=100)

        for _, row_data in df.head(50).iterrows():
            tree.insert("", "end", values=row_data.tolist()[:6])

        v_scrollbar = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        h_scrollbar = ttk.Scrollbar(frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)

        tree.pack(side="left", fill="both", expand=True)
        v_scrollbar.pack(side="right", fill="y")
        h_scrollbar.pack(side="bottom", fill="x")

        self.visuals.append(('data_table', frame))

    def refresh_data(self):
        if hasattr(self.main_app, 'df') and self.main_app.df is not None:
            # فیلترهای داشبورد را دوباره روی df جدید اعمال می‌کنیم
            self.apply_filters_to_visuals()
            self.populate_fields_panel()
            self.status_label.config(text="داده‌ها بروزرسانی شدند")
        else:
            messagebox.showwarning("هشدار", "هیچ داده‌ای برای بروزرسانی وجود ندارد")

    def save_report(self):
        try:
            filename = filedialog.asksaveasfilename(
                defaultextension=".json",
                filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
            )
            if filename:
                report_data = {
                    'filters': self.current_filters,
                    'visuals_count': len(self.visuals),
                    'saved_at': datetime.now().isoformat()
                }
                with open(filename, 'w', encoding='utf-8') as f:
                    json.dump(report_data, f, ensure_ascii=False, indent=2)
                messagebox.showinfo("موفق", "گزارش با موفقیت ذخیره شد")
        except Exception as e:
            logging.error(f"Error saving dashboard report: {e}")
            messagebox.showerror("خطا", f"خطا در ذخیره گزارش: {e}")

    def load_report(self):
        try:
            filename = filedialog.askopenfilename(
                filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
            )
            if filename and os.path.exists(filename):
                with open(filename, 'r', encoding='utf-8') as f:
                    report_data = json.load(f)
                # در این نسخه، فقط نمایش پیام؛ می‌توانید در آینده بازسازی فیلترها و ویژوال‌ها را اضافه کنید
                messagebox.showinfo("موفق", "گزارش با موفقیت بارگذاری شد")
        except Exception as e:
            logging.error(f"Error loading dashboard report: {e}")
            messagebox.showerror("خطا", f"خطا در بارگذاری گزارش: {e}")

    def export_pdf(self):
        messagebox.showinfo("خروجی PDF", "این قابلیت در نسخه بعدی اضافه خواهد شد")


# -----------------------------
class ExcelReportApp:
    def __init__(self, root):
        self.root = root
        self.settings = load_settings()
        self.colors = self.settings.get("colors", {})

        self.tk_img = None
        self.logo_label = None
        self.df = None
        self.df_filtered = None
        self.df_normalized = None
        self.df_grouped = None

        self.repair_col = None
        self.part_col = None
        self.date_col = None
        self.perf_col = None
        self.req_col = None
        self.code_col = None

        self.persian_font = register_persian_fonts()
        self.has_persian_support = self.check_persian_support()

        self.status_var = tk.StringVar()
        self.status_var.set("آماده")

        self.root.title("گزارش قالبسازی - برنامه‌ریز تعمیر و ساخت")
        geom = self.settings.get("window_size", "1200x800")
        try:
            self.root.geometry(geom)
        except Exception:
            self.root.geometry("1200x800")

        self.root.configure(bg=self.colors.get("bg_main", "#FFA500"))
        self.logo_path = self.fix_logo_path(self.settings.get("logo_path", ""))

        self.create_menu()
        self.setup_ui()
        self.load_saved_fields()

        self.root.after(1000, self.debug_logo_info)
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

    # -------------------- Persian Support --------------------
    def fix_logo_path(self, path):
        """اصلاح مسیر لوگو و بررسی وجود فایل"""
        if not path:
            return ""

        path = path.replace("\\", "/").strip()

        if os.path.exists(path):
            return path
        else:
            base_dir = os.path.dirname(os.path.abspath(__file__))
            possible_paths = [
                path,
                os.path.join(base_dir, path),
                os.path.join(base_dir, "logo.png"),
                os.path.join(base_dir, "logo.jpg"),
                os.path.join(base_dir, "images", "logo.png"),
            ]

            for possible_path in possible_paths:
                if os.path.exists(possible_path):
                    return possible_path

            return ""

    def check_persian_support(self):
        try:
            import arabic_reshaper  # noqa
            from bidi.algorithm import get_display  # noqa
            return True
        except ImportError:
            print("کتابخانه‌های arabic-reshaper و python-bidi نصب نیستند.")
            print("برای نصب از دستور زیر استفاده کنید:")
            print("pip install arabic-reshaper python-bidi")
            return False

    def reshape_persian_text(self, text):
        if not self.has_persian_support:
            return str(text)

        try:
            import arabic_reshaper
            from bidi.algorithm import get_display

            text_str = str(text)
            if all(ord(c) < 128 for c in text_str):
                return text_str

            reshaped_text = arabic_reshaper.reshape(text_str)
            bidi_text = get_display(reshaped_text)
            return bidi_text
        except Exception as e:
            logging.error(f"Error reshaping Persian text: {e}")
            return str(text)

    # -------------------- Menu & Top-level --------------------
    def create_menu(self):
        self.menubar = tk.Menu(self.root)

        file_menu = tk.Menu(self.menubar, tearoff=0)
        file_menu.add_command(label="انتخاب لوگو", command=self.select_logo)
        file_menu.add_command(label="حذف لوگو", command=self.remove_logo)
        file_menu.add_command(label="اطلاعات دیباگ لوگو", command=self.debug_logo_info)
        file_menu.add_command(label="اطلاعات دیباگ ستون‌ها", command=self.debug_columns_info)
        file_menu.add_command(label="اطلاعات دیباگ فیلتر هوشمند", command=self.debug_smart_filter)
        file_menu.add_command(label="ذخیره تنظیمات", command=lambda: save_settings(self.settings))
        file_menu.add_command(label="بارگذاری دستی settings.json", command=self.debug_show_settings)
        file_menu.add_separator()
        file_menu.add_command(label="❌ خروج", command=self.root.quit)
        self.menubar.add_cascade(label="فایل", menu=file_menu)

        powerbi_menu = tk.Menu(self.menubar, tearoff=0)
        powerbi_menu.add_command(label="🚀 بازکردن داشبورد Power BI", command=self.open_power_bi_dashboard)
        self.menubar.add_cascade(label="Power BI", menu=powerbi_menu)

        help_menu = tk.Menu(self.menubar, tearoff=0)
        help_menu.add_command(label="راهنمای برنامه", command=self.show_help)
        self.menubar.add_cascade(label="راهنما", menu=help_menu)

        self.root.config(menu=self.menubar)

    def open_power_bi_dashboard(self):
        try:
            if self.df is None:
                messagebox.showwarning("هشدار", "لطفاً ابتدا داده‌ها را بارگذاری کنید")
                return

            dashboard_window = tk.Toplevel(self.root)
            dashboard_window.title("Power BI Dashboard - گزارش‌گیری قالب‌سازی")
            dashboard_window.geometry("1400x900")

            PowerBIDashboard(dashboard_window, self)

        except Exception as e:
            logging.error(f"Error opening Power BI dashboard: {e}")
            messagebox.showerror("خطا", f"خطا در باز کردن داشبورد: {e}")

    # -------------------- Debug Info --------------------
    def debug_columns_info(self):
        if self.df is not None and self.repair_col is not None:
            raw_unique = self.df[self.repair_col].dropna().astype(str).unique()
            if self.df_normalized is not None:
                norm_unique = self.df_normalized[self.repair_col].dropna().astype(str).unique()
            else:
                norm_unique = []

            info_msg = f"""اطلاعات دیباگ ستون‌ها:

ستون نوع تعمیر: '{self.repair_col}'
ستون قالب/قطعه/دستگاه: '{self.part_col}'

مقادیر نوع تعمیر (خام): {list(raw_unique)}
مقادیر نوع تعمیر (نرمالایز شده): {list(norm_unique)}
            """
            messagebox.showinfo("اطلاعات دیباگ ستون‌ها", info_msg)
        else:
            messagebox.showwarning("هشدار", "ابتدا داده‌ها را بارگذاری کنید.")

    def debug_smart_filter(self):
        if self.df is None or self.df_normalized is None or self.repair_col is None or self.part_col is None:
            messagebox.showwarning("هشدار", "ابتدا داده‌ها را بارگذاری کنید.")
            return

        selected_repair = self.repair_cb.get()
        if not selected_repair or selected_repair == "(همه)":
            messagebox.showinfo("دیباگ فیلتر", "هیچ نوع تعمیری انتخاب نشده است.")
            return

        mask = self.df_normalized[self.repair_col].astype(str) == selected_repair
        matching_records = self.df_normalized[mask]

        info_msg = f"""دیباگ فیلتر هوشمند:

نوع تعمیر انتخاب شده: '{selected_repair}'
تعداد رکوردهای مربوطه: {len(matching_records)}
قالب/قطعه/دستگاه‌های مربوطه: {list(matching_records[self.part_col].dropna().astype(str).unique())}
        """
        messagebox.showinfo("دیباگ فیلتر هوشمند", info_msg)

    def select_logo(self):
        path = filedialog.askopenfilename(
            title="انتخاب لوگو",
            filetypes=[("Image Files", "*.png;*.jpg;*.jpeg;*.gif;*.bmp"), ("All", "*.*")]
        )
        if not path:
            return

        self.settings["logo_path"] = path
        save_settings(self.settings)
        self.logo_path = self.fix_logo_path(path)
        self.update_logo_display()

        if self.logo_path and os.path.exists(self.logo_path):
            self.status_var.set("لوگو به روز شد")
            messagebox.showinfo("موفق", "لوگو با موفقیت بارگذاری شد")
        else:
            self.status_var.set("خطا در بارگذاری لوگو")
            messagebox.showerror("خطا", "لوگو یافت نشد یا قابل بارگذاری نیست")

    def remove_logo(self):
        self.settings["logo_path"] = ""
        self.logo_path = ""
        save_settings(self.settings)
        self.update_logo_display()
        self.status_var.set("لوگو حذف شد")

    def update_logo_display(self):
        if self.logo_label:
            self.logo_label.destroy()
            self.logo_label = None
            self.tk_img = None

        if self.logo_path and os.path.exists(self.logo_path):
            try:
                img = Image.open(self.logo_path)
                original_width, original_height = img.size
                max_size = 120
                ratio = min(max_size / original_width, max_size / original_height)
                new_size = (int(original_width * ratio), int(original_height * ratio))

                img = img.resize(new_size, Image.Resampling.LANCZOS)
                self.tk_img = ImageTk.PhotoImage(img)
                self.logo_label = tk.Label(
                    self.top_frame,
                    image=self.tk_img,
                    bg=self.colors.get("frame_bg", "#FFE5B4")
                )
                self.logo_label.pack(side="right", padx=10)
                self.status_var.set("لوگو بارگذاری شد")
            except Exception as e:
                logging.error(f"Error loading logo: {e}")
                self.status_var.set("خطا در بارگذاری لوگو")
                self.logo_label = tk.Label(
                    self.top_frame,
                    text="خطا در لوگو",
                    bg=self.colors.get("frame_bg", "#FFE5B4"),
                    fg="red",
                    font=("Arial", 10)
                )
                self.logo_label.pack(side="right", padx=10)
        else:
            self.logo_label = tk.Label(
                self.top_frame,
                text="بدون لوگو\n(برای افزودن از منوی فایل استفاده کنید)",
                bg=self.colors.get("frame_bg", "#FFE5B4"),
                fg="gray",
                font=("Arial", 9)
            )
            self.logo_label.pack(side="right", padx=10)

    def debug_logo_info(self):
        info_msg = f"""اطلاعات دیباگ لوگو:

مسیر لوگو در تنظیمات: {self.settings.get('logo_path', '')}
مسیر لوگو پس از اصلاح: {self.logo_path}
آیا فایل وجود دارد: {os.path.exists(self.logo_path) if self.logo_path else False}
مسیر جاری: {os.getcwd()}
        """
        messagebox.showinfo("اطلاعات دیباگ لوگو", info_msg)

    def show_help(self):
        help_text = """
راهنمای برنامه گزارش‌گیری قالب‌سازی

ویژگی‌ها:

1. فیلتر ساده:
- فیلتر بر اساس تاریخ (جلالی YYYY/MM/DD)
- فیلتر بر اساس نوع تعمیر
- فیلتر بر اساس قالب/قطعه/دستگاه

2. فیلتر ترکیبی پیشرفته:
- انتخاب چندین نوع تعمیر
- فیلتر بر اساس بازه ساعت کار شده

3. گروه‌بندی و جمع‌بندی:
- نمایش هر قالب فقط یک بار
- جمع‌بندی ساعت کاری
- خروجی 4 ستونی: قالب/کد/شماره/ساعت

4. داشبورد Power BI:
- نمودارهای متنوع از داده‌ها
- فیلترهای داخلی داشبورد
- امکان ذخیره‌ی گزارش داشبورد (در حد اطلاعات فیلترها)

توسعه‌دهنده: F.Alizadeh
        """
        messagebox.showinfo("راهنما", help_text)

    def debug_show_settings(self):
        messagebox.showinfo("settings.json", json.dumps(self.settings, ensure_ascii=False, indent=4))

    # -------------------- UI Setup --------------------
    def setup_ui(self):
        self.top_frame = ttk.Frame(self.root)
        self.top_frame.pack(fill="x", padx=10, pady=5)

        title_label = tk.Label(
            self.top_frame,
            text="برنامه‌ریز تعمیر و ساخت قالبسازی - F.Alizadeh",
            font=("Arial", 11, "bold"),
            bg=self.colors.get("frame_bg", "#FFE5B4")
        )
        title_label.pack(side="left", padx=10)

        self.update_logo_display()
        self.setup_filters_frame()
        self.setup_treeview()

        status_label = ttk.Label(self.root, textvariable=self.status_var, relief="sunken", anchor="w")
        status_label.pack(fill="x", padx=10, pady=5)

    def setup_filters_frame(self):
        self.frame_filters = ttk.LabelFrame(self.root, text="فیلتر ساده", padding=10)
        self.frame_filters.pack(fill="x", padx=10, pady=5)

        ttk.Label(self.frame_filters, text="مسیر فایل اکسل:").grid(row=0, column=0, sticky="w", padx=5, pady=2)
        self.file_entry = ttk.Entry(self.frame_filters, width=70)
        self.file_entry.grid(row=0, column=1, padx=5, pady=2)
        ttk.Button(self.frame_filters, text="انتخاب فایل", command=self.select_file).grid(row=0, column=2, padx=5, pady=2)

        ttk.Label(self.frame_filters, text="نام شیت:").grid(row=1, column=0, sticky="w", padx=5, pady=2)
        self.sheet_cb = ttk.Combobox(self.frame_filters, width=30, state="readonly")
        self.sheet_cb.grid(row=1, column=1, sticky="w", padx=5, pady=2)
        ttk.Button(self.frame_filters, text="بارگذاری شیت‌ها", command=self.load_sheets).grid(row=1, column=2, padx=5, pady=2)

        ttk.Label(self.frame_filters, text="تاریخ شروع (YYYY/MM/DD):").grid(row=2, column=0, sticky="w", padx=5, pady=2)
        self.start_entry = ttk.Entry(self.frame_filters, width=15)
        self.start_entry.grid(row=2, column=1, sticky="w", padx=5, pady=2)

        ttk.Label(self.frame_filters, text="تاریخ پایان (YYYY/MM/DD):").grid(row=3, column=0, sticky="w", padx=5, pady=2)
        self.end_entry = ttk.Entry(self.frame_filters, width=15)
        self.end_entry.grid(row=3, column=1, sticky="w", padx=5, pady=2)

        ttk.Label(self.frame_filters, text="نوع تعمیر:").grid(row=4, column=0, sticky="w", padx=5, pady=2)
        self.repair_cb = ttk.Combobox(self.frame_filters, width=30, state="readonly")
        self.repair_cb.grid(row=4, column=1, sticky="w", padx=5, pady=2)
        self.repair_cb.bind('<<ComboboxSelected>>', self.on_repair_type_changed)

        ttk.Label(self.frame_filters, text="قالب / قطعه / دستگاه:").grid(row=5, column=0, sticky="w", padx=5, pady=2)
        self.part_cb = ttk.Combobox(self.frame_filters, width=30, state="readonly")
        self.part_cb.grid(row=5, column=1, sticky="w", padx=5, pady=2)

        button_frame = ttk.Frame(self.frame_filters)
        button_frame.grid(row=6, column=0, columnspan=3, pady=10)

        ttk.Button(button_frame, text="📂 بارگذاری داده‌ها", command=self.load_values).pack(side="left", padx=5)
        ttk.Button(button_frame, text="🔍 اعمال فیلتر ساده", command=self.apply_simple_filter).pack(side="left", padx=5)
        ttk.Button(button_frame, text="💾 ذخیره", command=lambda: self.save_output(self.df_filtered)).pack(side="left", padx=5)
        ttk.Button(button_frame, text="پاک کردن فیلترها", command=self.clear_filters).pack(side="left", padx=5)

        # فیلتر ترکیبی
        frame_advanced = ttk.LabelFrame(self.root, text="فیلتر ترکیبی پیشرفته", padding=10)
        frame_advanced.pack(fill="x", padx=10, pady=5)

        ttk.Label(frame_advanced, text="انتخاب چندگانه نوع تعمیر:").grid(row=0, column=0, sticky="w", padx=5, pady=2)
        self.repair_listbox = tk.Listbox(frame_advanced, selectmode=tk.MULTIPLE, height=4, width=40)
        self.repair_listbox.grid(row=0, column=1, sticky="w", padx=5, pady=2)

        ttk.Label(frame_advanced, text="بازه ساعت کار شده:").grid(row=1, column=0, sticky="w", padx=5, pady=2)
        hour_frame = ttk.Frame(frame_advanced)
        hour_frame.grid(row=1, column=1, sticky="w", padx=5, pady=2)

        ttk.Label(hour_frame, text="از:").pack(side="left")
        self.hour_min_entry = ttk.Entry(hour_frame, width=8)
        self.hour_min_entry.pack(side="left", padx=2)

        ttk.Label(hour_frame, text="تا:").pack(side="left", padx=(10, 0))
        self.hour_max_entry = ttk.Entry(hour_frame, width=8)
        self.hour_max_entry.pack(side="left", padx=2)

        advanced_button_frame = ttk.Frame(frame_advanced)
        advanced_button_frame.grid(row=2, column=0, columnspan=2, pady=10)

        ttk.Button(advanced_button_frame, text="🔍 اعمال فیلتر ترکیبی", command=self.apply_advanced_filter).pack(side="left", padx=5)
        ttk.Button(advanced_button_frame, text="📊 گروه‌بندی و جمع‌بندی", command=self.apply_grouping_filter).pack(side="left", padx=5)
        ttk.Button(advanced_button_frame, text="💾 ذخیره", command=lambda: self.save_output(self.df_filtered)).pack(side="left", padx=5)

    def setup_treeview(self):
        tree_frame = ttk.Frame(self.root)
        tree_frame.pack(fill="both", expand=True, padx=10, pady=10)

        v_scrollbar = ttk.Scrollbar(tree_frame, orient="vertical")
        v_scrollbar.pack(side="right", fill="y")

        h_scrollbar = ttk.Scrollbar(tree_frame, orient="horizontal")
        h_scrollbar.pack(side="bottom", fill="x")

        self.tree = ttk.Treeview(
            tree_frame,
            columns=("نوع تعمیر", "قالب/قطعه/دستگاه", "شماره نامه درخواست", "کد قالب", "مقدار ساعت کار شده"),
            show="headings",
            height=18,
            yscrollcommand=v_scrollbar.set,
            xscrollcommand=h_scrollbar.set
        )

        v_scrollbar.config(command=self.tree.yview)
        h_scrollbar.config(command=self.tree.xview)

        for col in self.tree["columns"]:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=200, anchor="center")

        self.tree.pack(fill="both", expand=True)

    # -------------------- Filters Logic --------------------
    def on_repair_type_changed(self, event=None):
        if self.df_normalized is None or self.repair_col is None or self.part_col is None:
            return

        selected_repair = self.repair_cb.get()

        if not selected_repair or selected_repair == "(همه)":
            part_values = ["(همه)"] + sorted(self.df[self.part_col].dropna().astype(str).unique())
            self.part_cb["values"] = part_values
            self.status_var.set("همه قالب/قطعه/دستگاه‌ها نمایش داده می‌شوند")
        else:
            try:
                mask = self.df_normalized[self.repair_col].astype(str) == selected_repair
                filtered_parts = self.df_normalized.loc[mask, self.part_col].dropna().astype(str).unique()

                part_values = ["(همه)"] + sorted(filtered_parts)
                self.part_cb["values"] = part_values
                self.part_cb.set('')

                count = len(part_values) - 1
                self.status_var.set(f"{count} قالب/قطعه/دستگاه برای نوع تعمیر '{selected_repair}' یافت شد")
            except Exception as e:
                logging.error(f"Error in on_repair_type_changed: {e}")
                self.status_var.set("خطا در فیلتر کردن داده‌ها")

    def update_repair_listbox(self):
        if self.df_normalized is not None and self.repair_col is not None:
            self.repair_listbox.delete(0, tk.END)
            repair_types = sorted(self.df_normalized[self.repair_col].dropna().astype(str).unique())
            for repair_type in repair_types:
                self.repair_listbox.insert(tk.END, repair_type)

    def apply_advanced_filter(self):
        if self.df is None or self.df_normalized is None:
            messagebox.showwarning("هشدار", "ابتدا داده‌ها را بارگذاری کنید.")
            return

        df = self.df.copy()

        selected_repairs = [self.repair_listbox.get(i) for i in self.repair_listbox.curselection()]
        if selected_repairs and self.repair_col in self.df_normalized.columns:
            mask = self.df_normalized[self.repair_col].astype(str).isin(selected_repairs)
            df = df[mask]

        hour_min = self.hour_min_entry.get().strip()
        hour_max = self.hour_max_entry.get().strip()

        if (hour_min or hour_max) and self.perf_col in df.columns:
            try:
                df[self.perf_col] = pd.to_numeric(df[self.perf_col], errors="coerce")

                if hour_min and hour_max:
                    df = df[(df[self.perf_col] >= float(hour_min)) & (df[self.perf_col] <= float(hour_max))]
                elif hour_min:
                    df = df[df[self.perf_col] >= float(hour_min)]
                elif hour_max:
                    df = df[df[self.perf_col] <= float(hour_max)]
            except ValueError:
                messagebox.showerror("خطا", "مقادیر ساعت باید عددی باشند.")

        self.df_filtered = df
        self.update_treeview(df)

        filtered_count = len(df)
        self.status_var.set(f"فیلتر ترکیبی اعمال شد. {filtered_count} رکورد نمایش داده می‌شود")

    def apply_grouping_filter(self):
        if self.df is None:
            messagebox.showwarning("هشدار", "ابتدا داده‌ها را بارگذاری کنید.")
            return

        if self.df_filtered is None or self.df_filtered.empty:
            self.apply_advanced_filter()

        if self.df_filtered is None or self.df_filtered.empty:
            messagebox.showwarning("هشدار", "هیچ داده‌ای برای گروه‌بندی وجود ندارد.")
            return

        try:
            grouping_cols = []

            if self.part_col and self.part_col in self.df_filtered.columns:
                grouping_cols.append(self.part_col)

            if self.code_col and self.code_col in self.df_filtered.columns:
                grouping_cols.append(self.code_col)

            if self.req_col and self.req_col in self.df_filtered.columns:
                grouping_cols.append(self.req_col)

            if not grouping_cols:
                messagebox.showerror("خطا", "ستون‌های لازم برای گروه‌بندی یافت نشد.")
                return

            if self.perf_col in self.df_filtered.columns:
                self.df_filtered[self.perf_col] = pd.to_numeric(
                    self.df_filtered[self.perf_col],
                    errors="coerce"
                ).fillna(0)

            grouped_df = self.df_filtered.groupby(grouping_cols, as_index=False).agg({
                self.perf_col: 'sum'
            })

            grouped_df = grouped_df.sort_values(by=self.perf_col, ascending=False)

            self.df_grouped = grouped_df
            self.update_grouped_treeview(grouped_df)

            grouped_count = len(grouped_df)
            total_hours = grouped_df[self.perf_col].sum()

            self.status_var.set(f"گروه‌بندی انجام شد: {grouped_count} رکورد منحصر به فرد - مجموع ساعت: {total_hours:.2f}")

        except Exception as e:
            logging.error(f"Error in grouping: {e}")
            messagebox.showerror("خطا", f"خطا در گروه‌بندی داده‌ها: {str(e)}")

    def update_grouped_treeview(self, df):
        for item in self.tree.get_children():
            self.tree.delete(item)

        if df is None or df.empty:
            self.status_var.set("هیچ داده‌ای برای نمایش وجود ندارد")
            return

        self.tree["columns"] = ("قالب/قطعه/دستگاه", "کد قالب", "شماره نامه درخواست", "ساعت کار شده")

        for col in self.tree["columns"]:
            self.tree.heading(col, text="")

        self.tree.heading("قالب/قطعه/دستگاه", text="قالب/قطعه/دستگاه")
        self.tree.heading("کد قالب", text="کد قالب")
        self.tree.heading("شماره نامه درخواست", text="شماره نامه درخواست")
        self.tree.heading("ساعت کار شده", text="ساعت کار شده")

        self.tree.column("قالب/قطعه/دستگاه", width=200, anchor="center")
        self.tree.column("کد قالب", width=150, anchor="center")
        self.tree.column("شماره نامه درخواست", width=150, anchor="center")
        self.tree.column("ساعت کار شده", width=120, anchor="center")

        alt_bg = self.colors.get("tree_alt_bg", "#FFF5E0")
        tree_bg = self.colors.get("tree_bg", "#FFFFFF")

        for i, (_, row) in enumerate(df.iterrows()):
            bg_tag = "even" if i % 2 == 0 else "odd"

            part_value = row.get(self.part_col, "")
            code_value = row.get(self.code_col, "")
            req_value = row.get(self.req_col, "")

            try:
                perf_value = f"{float(row.get(self.perf_col, 0)):.2f}"
            except Exception:
                perf_value = str(row.get(self.perf_col, 0))

            self.tree.insert("", "end", values=(
                part_value,
                code_value,
                req_value,
                perf_value
            ), tags=(bg_tag,))

        self.tree.tag_configure("even", background=tree_bg)
        self.tree.tag_configure("odd", background=alt_bg)

        try:
            if self.perf_col in df.columns:
                total = df[self.perf_col].astype(float).sum()
                self.tree.insert("", "end", values=(
                    "جمع کل",
                    "",
                    "",
                    f"{total:.2f}"
                ), tags=("total",))

                self.tree.tag_configure(
                    "total",
                    background=self.colors.get("tree_total_bg", "#0000FF"),
                    foreground=self.colors.get("tree_total_fg", "#FFFFFF"),
                    font=("Arial", 10, "bold")
                )
        except Exception as e:
            logging.error(f"Error calculating total: {e}")

    def update_treeview(self, df):
        for item in self.tree.get_children():
            self.tree.delete(item)

        if df is None or df.empty:
            self.status_var.set("هیچ داده‌ای برای نمایش وجود ندارد")
            return

        self.tree["columns"] = ("نوع تعمیر", "قالب/قطعه/دستگاه", "شماره نامه درخواست", "کد قالب", "مقدار ساعت کار شده")

        for col in self.tree["columns"]:
            self.tree.heading(col, text="")

        self.tree.heading("نوع تعمیر", text="نوع تعمیر")
        self.tree.heading("قالب/قطعه/دستگاه", text="قالب/قطعه/دستگاه")
        self.tree.heading("شماره نامه درخواست", text="شماره نامه درخواست")
        self.tree.heading("کد قالب", text="کد قالب")
        self.tree.heading("مقدار ساعت کار شده", text="مقدار ساعت کار شده")

        for col in self.tree["columns"]:
            self.tree.column(col, width=180, anchor="center")

        alt_bg = self.colors.get("tree_alt_bg", "#FFF5E0")
        tree_bg = self.colors.get("tree_bg", "#FFFFFF")

        for i, (_, row) in enumerate(df.iterrows()):
            bg_tag = "even" if i % 2 == 0 else "odd"

            repair_value = row.get(self.repair_col, "")
            part_value = row.get(self.part_col, "")
            req_value = row.get(self.req_col, "")
            code_value = row.get(self.code_col, "")
            perf_value = row.get(self.perf_col, 0)

            try:
                perf_value = f"{float(perf_value):.2f}"
            except Exception:
                perf_value = str(perf_value)

            self.tree.insert("", "end", values=(
                repair_value,
                part_value,
                req_value,
                code_value,
                perf_value
            ), tags=(bg_tag,))

        self.tree.tag_configure("even", background=tree_bg)
        self.tree.tag_configure("odd", background=alt_bg)

        try:
            if self.perf_col in df.columns:
                total = df[self.perf_col].astype(float).sum()
                self.tree.insert("", "end", values=("جمع کل", "", "", "", f"{total:.2f}"), tags=("total",))
                self.tree.tag_configure(
                    "total",
                    background=self.colors.get("tree_total_bg", "#0000FF"),
                    foreground=self.colors.get("tree_total_fg", "#FFFFFF"),
                    font=("Arial", 10, "bold")
                )
        except Exception as e:
            logging.error(f"Error calculating total: {e}")

    def load_saved_fields(self):
        last_path = self.settings.get("last_excel_path", "")
        if last_path:
            self.file_entry.delete(0, tk.END)
            self.file_entry.insert(0, last_path)

        filt = self.settings.get("filters", {})
        self.start_entry.delete(0, tk.END)
        self.start_entry.insert(0, filt.get("start_date", ""))
        self.end_entry.delete(0, tk.END)
        self.end_entry.insert(0, filt.get("end_date", ""))

    def clear_filters(self):
        self.start_entry.delete(0, tk.END)
        self.end_entry.delete(0, tk.END)
        self.repair_cb.set('')
        self.part_cb.set('')
        self.hour_min_entry.delete(0, tk.END)
        self.hour_max_entry.delete(0, tk.END)
        self.repair_listbox.selection_clear(0, tk.END)
        self.status_var.set("فیلترها پاک شدند")

    # -------------------- File & Data Loading --------------------
    def select_file(self):
        path = filedialog.askopenfilename(
            title="انتخاب فایل اکسل",
            filetypes=[("Excel Files", "*.xlsx"), ("All", "*.*")]
        )
        if not path:
            return
        self.file_entry.delete(0, tk.END)
        self.file_entry.insert(0, path)
        self.settings["last_excel_path"] = path
        save_settings(self.settings)
        self.status_var.set(f"فایل انتخاب شد: {os.path.basename(path)}")

    def load_sheets(self):
        path = self.file_entry.get().strip()
        if not path or not os.path.exists(path):
            messagebox.showerror("خطا", "فایل یافت نشد.")
            return

        self.set_loading_cursor(True)
        try:
            wb = load_workbook(path, read_only=True)
            sheetnames = wb.sheetnames[:]
            wb.close()

            self.sheet_cb["values"] = sheetnames
            last_sheet = self.settings.get("last_sheet", "")
            if last_sheet and last_sheet in sheetnames:
                self.sheet_cb.set(last_sheet)

            self.status_var.set(f"{len(sheetnames)} شیت پیدا شد")

        except Exception as e:
            logging.error(f"Error loading sheets: {e}")
            messagebox.showerror("خطا", f"خطا در بارگذاری شیت‌ها: {str(e)}")
        finally:
            self.set_loading_cursor(False)

    def load_values(self):
        path = self.file_entry.get().strip()
        sheet = self.sheet_cb.get().strip()

        if not path or not sheet or not os.path.exists(path):
            messagebox.showerror("خطا", "فایل و شیت را انتخاب کنید.")
            return

        # پاک‌سازی قبلی
        for attr in ['df', 'df_filtered', 'df_normalized', 'df_grouped']:
            if hasattr(self, attr):
                setattr(self, attr, None)

        self.set_loading_cursor(True)
        try:
            wb = load_workbook(path, data_only=True, read_only=True)
            ws = wb[sheet]
            rows = list(ws.values)
            wb.close()

            if not rows:
                messagebox.showerror("خطا", "شیت انتخاب‌شده خالی است.")
                return

            headers = [str(x).strip() if x else "" for x in rows[0]]
            df = pd.DataFrame(rows[1:], columns=headers)
            self.df = df

            # تشخیص ستون‌ها
            self.detect_columns(df)

            # ساخت df_normalized پس از تشخیص ستون repair_col
            self.df_normalized = df.copy()
            if self.repair_col and self.repair_col in df.columns:
                self.df_normalized[self.repair_col] = self.df_normalized[self.repair_col].apply(normalize_repair_type)

            self.settings["last_sheet"] = sheet
            save_settings(self.settings)

            self.populate_comboboxes(self.df_normalized)
            self.update_repair_listbox()

            record_count = len(df)
            self.status_var.set(f"تعداد {record_count} رکورد بارگذاری شد")
            messagebox.showinfo("موفق", f"اطلاعات بارگذاری و نرمالایز شد. ({record_count} رکورد)")

        except Exception as e:
            logging.error(f"Error loading values: {e}")
            messagebox.showerror("خطا", f"خطا در بارگذاری داده‌ها: {str(e)}")
        finally:
            self.set_loading_cursor(False)

    def detect_columns(self, df):
        self.repair_col = find_column(df.columns, ["نوع تعمیر", "تعمیر", "repair"])
        self.part_col = find_column(df.columns, ["قالب / قطعه / دستگاه", "قالب", "قطعه", "دستگاه", "part", "device"])
        self.date_col = find_column(df.columns, ["تاریخ", "date"])
        self.perf_col = find_column(df.columns, ["مقدار ساعت کار شده", "ساعت", "hour", "time"])
        self.req_col = find_column(df.columns, ["شماره نامه درخواست", "شماره درخواست", "request"])
        self.code_col = find_column(df.columns, ["کد قالب", "کد", "code"])

    def populate_comboboxes(self, df):
        if self.repair_col in df.columns:
            repair_values = ["(همه)"] + sorted(df[self.repair_col].dropna().astype(str).unique())
            self.repair_cb["values"] = repair_values

        if self.part_col in df.columns:
            part_values = ["(همه)"] + sorted(df[self.part_col].dropna().astype(str).unique())
            self.part_cb["values"] = part_values

    def set_loading_cursor(self, loading):
        self.root.config(cursor="watch" if loading else "")
        self.root.update()

    # -------------------- Simple Filter --------------------
    def apply_simple_filter(self):
        if self.df is None or self.df_normalized is None:
            messagebox.showwarning("هشدار", "ابتدا داده‌ها را بارگذاری کنید.")
            return

        df = self.df.copy()

        s = self.start_entry.get().strip()
        e = self.end_entry.get().strip()

        if (s and not e) or (e and not s):
            messagebox.showwarning("هشدار", "هر دو فیلد تاریخ باید پر شوند یا خالی باشند.")
            return

        self.settings["filters"]["start_date"] = s
        self.settings["filters"]["end_date"] = e
        save_settings(self.settings)

        if s and e and self.date_col:
            try:
                s_g = JalaliDate.strptime(s, "%Y/%m/%d").to_gregorian()
                e_g = JalaliDate.strptime(e, "%Y/%m/%d").to_gregorian()
                df[self.date_col] = pd.to_datetime(df[self.date_col], errors="coerce")
                df = df[(df[self.date_col] >= s_g) & (df[self.date_col] <= e_g)]
                self.status_var.set(f"فیلتر تاریخ اعمال شد: {s} تا {e}")
            except Exception as exc:
                logging.error(f"Date filter error: {exc}")
                messagebox.showerror("خطا", "فرمت تاریخ اشتباه است. از فرمت YYYY/MM/DD استفاده کنید.")

        rep = self.repair_cb.get()
        if rep and rep != "(همه)" and self.repair_col:
            mask = self.df_normalized[self.repair_col].astype(str) == rep
            df = df[mask]
            self.settings["filters"]["repair_type"] = rep

        part = self.part_cb.get()
        if part and part != "(همه)" and self.part_col:
            df = df[df[self.part_col].astype(str) == part]
            self.settings["filters"]["part_type"] = part

        save_settings(self.settings)

        if self.perf_col in df.columns:
            df[self.perf_col] = pd.to_numeric(df[self.perf_col], errors="coerce").fillna(0)

        self.df_filtered = df
        self.update_treeview(df)

        filtered_count = len(df)
        self.status_var.set(f"فیلتر اعمال شد. {filtered_count} رکورد نمایش داده می‌شود")

    # -------------------- Save Output --------------------
    def save_output(self, df):
        if df is None or df.empty:
            messagebox.showerror("خطا", "هیچ داده‌ای برای ذخیره وجود ندارد.")
            return

        path = filedialog.asksaveasfilename(
            title="ذخیره گزارش",
            defaultextension=".xlsx",
            filetypes=[
                ("Excel", "*.xlsx"),
                ("CSV", "*.csv"),
                ("PDF", "*.pdf")
            ]
        )

        if not path:
            return

        self.set_loading_cursor(True)
        try:
            # اگر گروه‌بندی شده است، از df_grouped برای خروجی استفاده شود
            if self.df_grouped is not None and not self.df_grouped.empty:
                df_to_save = self.df_grouped.copy()
                columns_to_keep = []
                if self.part_col in df_to_save.columns:
                    columns_to_keep.append(self.part_col)
                if self.code_col in df_to_save.columns:
                    columns_to_keep.append(self.code_col)
                if self.req_col in df_to_save.columns:
                    columns_to_keep.append(self.req_col)
                if self.perf_col in df_to_save.columns:
                    columns_to_keep.append(self.perf_col)

                df_to_save = df_to_save[columns_to_keep]

                column_mapping = {
                    self.part_col: "قالب/قطعه/دستگاه",
                    self.code_col: "کد قالب",
                    self.req_col: "شماره نامه درخواست",
                    self.perf_col: "ساعت کار شده"
                }
                df_to_save = df_to_save.rename(columns=column_mapping)
            else:
                df_to_save = df

            if path.endswith(".xlsx"):
                self.save_excel(df_to_save, path)
            elif path.endswith(".csv"):
                self.save_csv(df_to_save, path)
            elif path.endswith(".pdf"):
                self.save_pdf(df_to_save, path)
            else:
                messagebox.showerror("خطا", "فرمت فایل پشتیبانی نمی‌شود.")
                return

            messagebox.showinfo("ذخیره شد", f"فایل با موفقیت ذخیره شد:\n{path}")
            self.status_var.set(f"فایل ذخیره شد: {os.path.basename(path)}")

        except Exception as e:
            logging.error(f"Error saving file: {e}")
            messagebox.showerror("خطا در ذخیره", f"خطا در ذخیره فایل: {str(e)}")
        finally:
            self.set_loading_cursor(False)

    def save_excel(self, df, path):
        wb = Workbook()
        ws = wb.active
        ws.title = "گزارش قالب‌سازی"

        if self.logo_path and os.path.exists(self.logo_path):
            try:
                img = XLImage(self.logo_path)
                img.width = 120
                img.height = 120
                ws.add_image(img, "H1")
            except Exception as e:
                logging.error(f"Error adding logo to Excel: {e}")

        ws.append(list(df.columns))

        for r in df.itertuples(index=False):
            ws.append(list(r))

        if "ساعت کار شده" in df.columns:
            total_row = len(df) + 2
            ws.cell(row=total_row, column=1, value="جمع کل")
            col_index = list(df.columns).index("ساعت کار شده") + 1
            ws.cell(row=total_row, column=col_index, value=df["ساعت کار شده"].sum())

            for col in range(1, len(df.columns) + 1):
                cell = ws.cell(row=total_row, column=col)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="0000FF")
                cell.alignment = Alignment(horizontal="center")

        for column_cells in ws.columns:
            values_len = [len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells]
            length = max(values_len) if values_len else 10
            ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)

        wb.save(path)

    def save_csv(self, df, path):
        df_out = df.copy()
        if "ساعت کار شده" in df_out.columns:
            total_row = {col: "" for col in df_out.columns}
            total_row["ساعت کار شده"] = df_out["ساعت کار شده"].sum()
            total_row[df_out.columns[0]] = "جمع کل"
            df_out = pd.concat([df_out, pd.DataFrame([total_row])], ignore_index=True)

        df_out.to_csv(path, index=False, encoding="utf-8-sig")

    def save_pdf(self, df, path):
        c = canvas.Canvas(path, pagesize=A4)
        width, height = A4

        font_name = self.persian_font

        c.setFont(font_name, 16)
        title = self.reshape_persian_text("گزارش قالب‌سازی")
        c.drawString(100, height - 50, title)

        if self.logo_path and os.path.exists(self.logo_path):
            try:
                c.drawImage(self.logo_path, width - 150, height - 120, width=100, height=100)
            except Exception as e:
                logging.error(f"Error adding logo to PDF: {e}")

        c.setFont(font_name, 12)
        y = height - 100

        # تشخیص نوع داده بر اساس نام ستون‌ها
        if set(["قالب/قطعه/دستگاه", "کد قالب", "شماره نامه درخواست", "ساعت کار شده"]).issubset(df.columns):
            headers = ["قالب/قطعه/دستگاه", "کد قالب", "شماره نامه درخواست", "ساعت کار شده"]
            col_widths = [120, 100, 120, 80]
        else:
            # حالت کلی: سعی می‌کنیم ستون‌های اصلی را پیدا کنیم
            headers = []
            col_widths = []
            for col in df.columns:
                headers.append(col)
                col_widths.append(100)

        x = 50
        for i, header in enumerate(headers):
            header_text = self.reshape_persian_text(header)
            c.drawString(x, y, header_text)
            x += col_widths[i]

        c.line(50, y - 5, width - 50, y - 5)

        c.setFont(font_name, 10)
        y -= 25

        for _, row in df.iterrows():
            if y < 100:
                c.showPage()
                c.setFont(font_name, 12)
                header_y = height - 50
                header_x = 50
                for i, header in enumerate(headers):
                    header_text = self.reshape_persian_text(header)
                    c.drawString(header_x, header_y, header_text)
                    header_x += col_widths[i]
                c.line(50, header_y - 5, width - 50, header_y - 5)
                c.setFont(font_name, 10)
                y = header_y - 30

            x = 50
            for i, header in enumerate(headers):
                value = str(row.get(header, ""))
                if len(value) > 20:
                    value = value[:20] + "..."
                value_text = self.reshape_persian_text(value)
                c.drawString(x, y, value_text)
                x += col_widths[i]

            y -= 20

        if "ساعت کار شده" in df.columns:
            if y < 100:
                c.showPage()
                c.setFont(font_name, 12)
                y = height - 50
            else:
                c.setFont(font_name, 12)

            total_text = f"جمع کل ساعت کار شده: {df['ساعت کار شده'].sum()}"
            total_text_reshaped = self.reshape_persian_text(total_text)
            c.drawString(50, y - 30, total_text_reshaped)

        c.save()

    # -------------------- Close --------------------
    def on_close(self):
        try:
            self.settings["window_size"] = self.root.geometry()
            save_settings(self.settings)
            self.status_var.set("برنامه بسته شد")
        except Exception as e:
            logging.error(f"Error on close: {e}")
        finally:
            self.root.destroy()


# -----------------------------
if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelReportApp(root)
    root.mainloop()