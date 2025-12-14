# -*- coding: utf-8 -*-
"""
Ù†Ø³Ø®Ù‡ Ù†Ù‡Ø§ÛŒÛŒ Ø¨Ø§ ÙÛŒÙ„ØªØ± Ø³Ø§Ø¯Ù‡ Ùˆ Ù‡ÙˆØ´Ù…Ù†Ø¯ Ùˆ Ø¬Ù…Ø¹ Ú©Ù„ Ø³Ø§Ø¹Øª Ú©Ø§Ø± Ø´Ø¯Ù‡ Ø¯Ø± Treeview
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import os
from persiantools.jdatetime import JalaliDate
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image as XLImage
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from PIL import Image, ImageTk
import sys



def find_column(columns, possible_names):
    for name in possible_names:
        for col in columns:
            if name.strip() in str(col).strip():
                return col
    return None

def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

class ExcelReportApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Ú¯Ø²Ø§Ø±Ø´ Ù‚Ø§Ù„Ø¨Ø³Ø§Ø²ÛŒ - Ù†Ø³Ø®Ù‡ Ù‡ÙˆØ´Ù…Ù†Ø¯")
        self.root.geometry("1200x800")
        self.root.configure(bg="#f5f5f5")

        self.df = None
        self.df_filtered = None
        self.file_path = None

        self.logo_path = resource_path("logo.png")
        self.create_menu()
        self.setup_ui()

    def create_menu(self):
        menubar = tk.Menu(self.root)
        file_menu = tk.Menu(menubar, tearoff=0)
        file_menu.add_command(label="âŒ Ø®Ø±ÙˆØ¬", command=self.root.quit)
        menubar.add_cascade(label="ÙØ§ÛŒÙ„", menu=file_menu)
        self.root.config(menu=menubar)
    

    def setup_ui(self):
        top_frame = ttk.Frame(self.root)
        top_frame.pack(fill="x", padx=10, pady=5)

        if os.path.exists(self.logo_path):
            img = Image.open(self.logo_path)
            img = img.resize((80, 50), Image.Resampling.LANCZOS)
            self.tk_img = ImageTk.PhotoImage(img)
            tk.Label(top_frame, image=self.tk_img).pack(side="left", padx=5)

        # ğŸ”¹ Ø§ØµÙ„Ø§Ø­ side="" â†’ Ø­Ø°Ù Ø´Ø¯
        tk.Label(top_frame, text="Ú©Ø§Ø±Ø´Ù†Ø§Ø³ Ø¨Ø±Ù†Ø§Ù…Ù‡ Ø±ÛŒØ²ÛŒ Ùˆ Ø³Ø§Ø®Øª Ù‚Ø§Ù„Ø¨Ø³Ø§Ø²ÛŒ: ÙÙˆØ§Ø¯ Ù…Ø·ÙˆØ± Ø¹Ù„ÛŒØ²Ø§Ø¯Ù‡",
                 font=("Arial", 11, "bold")).pack(padx=10)

        # ÙÛŒÙ„ØªØ± Ø³Ø§Ø¯Ù‡
        frame_simple = ttk.LabelFrame(self.root, text="ÙÛŒÙ„ØªØ± Ø³Ø§Ø¯Ù‡", padding=10)
        frame_simple.pack(padx=10, pady=5, fill="x")

        ttk.Label(frame_simple, text="Ù…Ø³ÛŒØ± ÙØ§ÛŒÙ„ Ø§Ú©Ø³Ù„:").grid(row=0, column=0, sticky="w")
        self.file_entry = ttk.Entry(frame_simple, width=70)
        self.file_entry.grid(row=0, column=1, padx=5)
        ttk.Button(frame_simple, text="Ø§Ù†ØªØ®Ø§Ø¨ ÙØ§ÛŒÙ„", command=self.select_file).grid(row=0, column=2)

        ttk.Label(frame_simple, text="Ù†Ø§Ù… Ø´ÛŒØª:").grid(row=1, column=0, sticky="w")
        self.sheet_cb = ttk.Combobox(frame_simple, width=30, state="readonly")
        self.sheet_cb.grid(row=1, column=1, sticky="w")
        ttk.Button(frame_simple, text="Ø¨Ø§Ø±Ú¯Ø°Ø§Ø±ÛŒ Ø´ÛŒØªâ€ŒÙ‡Ø§", command=self.load_sheets).grid(row=1, column=2)

        ttk.Label(frame_simple, text="ØªØ§Ø±ÛŒØ® Ø´Ø±ÙˆØ¹ (YYYY/MM/DD):").grid(row=2, column=0, sticky="w")
        self.start_entry = ttk.Entry(frame_simple, width=15)
        self.start_entry.grid(row=2, column=1, sticky="w")

        ttk.Label(frame_simple, text="ØªØ§Ø±ÛŒØ® Ù¾Ø§ÛŒØ§Ù† (YYYY/MM/DD):").grid(row=3, column=0, sticky="w")
        self.end_entry = ttk.Entry(frame_simple, width=15)
        self.end_entry.grid(row=3, column=1, sticky="w")

        ttk.Label(frame_simple, text="Ù†ÙˆØ¹ ØªØ¹Ù…ÛŒØ±:").grid(row=4, column=0, sticky="w")
        self.repair_cb = ttk.Combobox(frame_simple, width=30, state="readonly")
        self.repair_cb.grid(row=4, column=1, sticky="w")

        ttk.Label(frame_simple, text="Ù‚Ø§Ù„Ø¨ / Ù‚Ø·Ø¹Ù‡ / Ø¯Ø³ØªÚ¯Ø§Ù‡:").grid(row=5, column=0, sticky="w")
        self.part_cb = ttk.Combobox(frame_simple, width=30, state="readonly")
        self.part_cb.grid(row=5, column=1, sticky="w")

        ttk.Button(frame_simple, text="ğŸ“‚ Ø¨Ø§Ø±Ú¯Ø°Ø§Ø±ÛŒ Ø¯Ø§Ø¯Ù‡â€ŒÙ‡Ø§", command=self.load_values).grid(row=6, column=0, pady=5)
        style = ttk.Style()
        style.configure("Green.TButton", background="green", foreground="white")
        ttk.Button(frame_simple, text="ğŸ” Ø§Ø¹Ù…Ø§Ù„ ÙÛŒÙ„ØªØ± Ø³Ø§Ø¯Ù‡", command=self.apply_simple_filter,
                   style="Green.TButton").grid(row=6, column=1, pady=5)
        ttk.Button(frame_simple, text="ğŸ’¾ Ø°Ø®ÛŒØ±Ù‡", command=lambda: self.save_output(self.df_filtered),
                   style="Green.TButton").grid(row=6, column=2, pady=5)

        # ÙÛŒÙ„ØªØ± Ù‡ÙˆØ´Ù…Ù†Ø¯
        frame_smart = ttk.LabelFrame(self.root, text="ÙÛŒÙ„ØªØ± Ù‡ÙˆØ´Ù…Ù†Ø¯", padding=10)
        frame_smart.pack(padx=10, pady=5, fill="x")

        ttk.Button(frame_smart, text="ğŸ” Ø§Ø¹Ù…Ø§Ù„ ÙÛŒÙ„ØªØ± Ù‡ÙˆØ´Ù…Ù†Ø¯", command=self.apply_smart_filter,
                   style="Green.TButton").grid(row=0, column=0, pady=5)
        ttk.Button(frame_smart, text="ğŸ’¾ Ø°Ø®ÛŒØ±Ù‡", command=lambda: self.save_output(self.df_filtered),
                   style="Green.TButton").grid(row=0, column=1, pady=5)

        # Ù†Ù…Ø§ÛŒØ´ Ù†ØªØ§ÛŒØ¬
        self.tree = ttk.Treeview(self.root,
                                 columns=("Ù†ÙˆØ¹ ØªØ¹Ù…ÛŒØ±", "Ù‚Ø§Ù„Ø¨/Ù‚Ø·Ø¹Ù‡/Ø¯Ø³ØªÚ¯Ø§Ù‡", "Ø´Ù…Ø§Ø±Ù‡ Ù†Ø§Ù…Ù‡ Ø¯Ø±Ø®ÙˆØ§Ø³Øª", "Ú©Ø¯ Ù‚Ø§Ù„Ø¨", "Ù…Ù‚Ø¯Ø§Ø± Ø³Ø§Ø¹Øª Ú©Ø§Ø± Ø´Ø¯Ù‡"),
                                 show="headings", height=25)
        for col in self.tree["columns"]:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=200, anchor="center")
        self.tree.pack(padx=10, pady=10, fill="both", expand=True)

    # --- Ø³Ø§ÛŒØ± Ù…ØªØ¯Ù‡Ø§ Ù…Ø§Ù†Ù†Ø¯ Ù‚Ø¨Ù„ ---
    def select_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if path:
            self.file_entry.delete(0, tk.END)
            self.file_entry.insert(0, path)
            self.file_path = path

    def load_sheets(self):
        path = self.file_entry.get().strip()
        if not path or not os.path.exists(path):
            messagebox.showerror("Ø®Ø·Ø§", "ÙØ§ÛŒÙ„ Ù…Ø¹ØªØ¨Ø± Ø§Ù†ØªØ®Ø§Ø¨ Ù†Ø´Ø¯Ù‡ Ø§Ø³Øª.")
            return
        try:
            wb = load_workbook(path, read_only=True)
            self.sheet_cb["values"] = wb.sheetnames
            wb.close()
            messagebox.showinfo("Ø§Ù†Ø¬Ø§Ù… Ø´Ø¯", "Ø´ÛŒØªâ€ŒÙ‡Ø§ Ø¨Ø§ Ù…ÙˆÙÙ‚ÛŒØª Ø¨Ø§Ø±Ú¯Ø°Ø§Ø±ÛŒ Ø´Ø¯Ù†Ø¯.")
        except Exception as e:
            messagebox.showerror("Ø®Ø·Ø§", str(e))

    def load_values(self):
        path = self.file_entry.get().strip()
        sel_sheet = self.sheet_cb.get().strip()
        if not path or not sel_sheet:
            messagebox.showerror("Ø®Ø·Ø§", "ÙØ§ÛŒÙ„ ÛŒØ§ Ø´ÛŒØª Ø±Ø§ Ø§Ù†ØªØ®Ø§Ø¨ Ú©Ù†ÛŒØ¯.")
            return
        try:
            wb = load_workbook(path, data_only=True, read_only=True)
            ws = wb[sel_sheet]
            rows = list(ws.values)
            wb.close()
            header = [str(x).strip() if x else "" for x in rows[0]]
            df = pd.DataFrame(rows[1:], columns=header)
            df.columns = [str(c).strip() for c in df.columns]

            self.df = df

            self.repair_col = find_column(df.columns, ["Ù†ÙˆØ¹ ØªØ¹Ù…ÛŒØ±", "ØªØ¹Ù…ÛŒØ±"])
            self.part_col = find_column(df.columns, ["Ù‚Ø§Ù„Ø¨ / Ù‚Ø·Ø¹Ù‡ / Ø¯Ø³ØªÚ¯Ø§Ù‡"])
            self.date_col = find_column(df.columns, ["ØªØ§Ø±ÛŒØ®"])
            self.perf_col = find_column(df.columns, ["Ù…Ù‚Ø¯Ø§Ø± Ø³Ø§Ø¹Øª Ú©Ø§Ø± Ø´Ø¯Ù‡"])
            self.req_col = find_column(df.columns, ["Ø´Ù…Ø§Ø±Ù‡ Ù†Ø§Ù…Ù‡ Ø¯Ø±Ø®ÙˆØ§Ø³Øª", "Ø´Ù…Ø§Ø±Ù‡ Ø¯Ø±Ø®ÙˆØ§Ø³Øª"])
            self.code_col = find_column(df.columns, ["Ú©Ø¯ Ù‚Ø§Ù„Ø¨"])

            self.repair_cb["values"] = ["(Ù‡Ù…Ù‡)"] + sorted(df[self.repair_col].dropna().astype(str).unique())
            self.part_cb["values"] = ["(Ù‡Ù…Ù‡)"] + sorted(df[self.part_col].dropna().astype(str).unique())

            messagebox.showinfo("Ø§Ù†Ø¬Ø§Ù… Ø´Ø¯", "Ø§Ø·Ù„Ø§Ø¹Ø§Øª Ø¨Ø§ Ù…ÙˆÙÙ‚ÛŒØª Ø¨Ø§Ø±Ú¯Ø°Ø§Ø±ÛŒ Ø´Ø¯.")
        except Exception as e:
            messagebox.showerror("Ø®Ø·Ø§", str(e))

    def apply_simple_filter(self):
        if self.df is None:
            return
        df = self.df.copy()
        s = self.start_entry.get().strip()
        e = self.end_entry.get().strip()
        if s and e:
            try:
                s_g = JalaliDate.strptime(s, "%Y/%m/%d").to_gregorian()
                e_g = JalaliDate.strptime(e, "%Y/%m/%d").to_gregorian()
                df[self.date_col] = pd.to_datetime(df[self.date_col], errors="coerce")
                df = df[(df[self.date_col] >= s_g) & (df[self.date_col] <= e_g)]
            except:
                pass
        sel_repair = self.repair_cb.get()
        if sel_repair and sel_repair != "(Ù‡Ù…Ù‡)":
            df = df[df[self.repair_col].astype(str) == sel_repair]
        sel_part = self.part_cb.get()
        if sel_part and sel_part != "(Ù‡Ù…Ù‡)":
            df = df[df[self.part_col].astype(str) == sel_part]
        df[self.perf_col] = pd.to_numeric(df[self.perf_col], errors="coerce").fillna(0)
        self.df_filtered = df
        self.update_treeview(df)

    def apply_smart_filter(self):
        if self.df_filtered is None:
            self.apply_simple_filter()
        df = self.df_filtered.copy()
        for col in df.columns:
            if col not in [self.repair_col, self.part_col, self.date_col, self.perf_col, self.req_col, self.code_col]:
                df[col] = df[col].fillna("(Ø®Ø§Ù„ÛŒ)")
                values = df[col].dropna().unique()
                if len(values) > 0:
                    df = df[df[col].isin(values)]
        self.df_filtered = df
        self.update_treeview(df)

    def update_treeview(self, df):
        for i in self.tree.get_children():
            self.tree.delete(i)
        for _, row in df.iterrows():
            self.tree.insert("", "end", values=(
                row.get(self.repair_col, ""),
                row.get(self.part_col, ""),
                row.get(self.req_col, ""),
                row.get(self.code_col, ""),
                row.get(self.perf_col, 0)
            ))
        total = df[self.perf_col].sum()
        self.tree.insert("", "end", values=("Ø¬Ù…Ø¹ Ú©Ù„", "", "", "", total))

    def save_output(self, df):
        if df is None or df.empty:
            messagebox.showerror("Ø®Ø·Ø§", "Ø§Ø¨ØªØ¯Ø§ ÙÛŒÙ„ØªØ± Ø±Ø§ Ø§Ø¹Ù…Ø§Ù„ Ú©Ù†ÛŒØ¯.")
            return
        path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                            filetypes=[("Excel", "*.xlsx"), ("CSV", "*.csv"), ("PDF", "*.pdf")])
        if not path:
            return
        df_out = df.copy()
        df_out.loc["Ø¬Ù…Ø¹ Ú©Ù„"] = [""] * len(df_out.columns)
        df_out.at["Ø¬Ù…Ø¹ Ú©Ù„", self.perf_col] = df_out[self.perf_col].sum()
        try:
            if path.endswith(".xlsx"):
                wb = Workbook()
                ws = wb.active
                ws.title = "Ú¯Ø²Ø§Ø±Ø´"
                if os.path.exists(self.logo_path):
                    img = XLImage(self.logo_path)
                    img.width = 80
                    img.height = 50
                    ws.add_image(img, "A1")
                ws.append(list(df_out.columns))
                for r in df_out.itertuples(index=False):
                    ws.append(list(r))
                wb.save(path)
            elif path.endswith(".csv"):
                df_out.to_csv(path, index=False, encoding="utf-8-sig")
            elif path.endswith(".pdf"):
                c = canvas.Canvas(path, pagesize=A4)
                c.setFont("Helvetica", 10)
                y = 800
                if os.path.exists(self.logo_path):
                    c.drawImage(self.logo_path, 50, y - 50, width=80, height=50)
                for _, row in df_out.iterrows():
                    text = " | ".join([str(x) for x in row.values if x is not None])
                    c.drawString(40, y, text)
                    y -= 14
                    if y < 50:
                        c.showPage()
                        c.setFont("Helvetica", 10)
                        y = 800
                c.drawString(50, 30, "F.Alizadeh")
                c.save()
            messagebox.showinfo("Ø°Ø®ÛŒØ±Ù‡ Ø´Ø¯", f"ÙØ§ÛŒÙ„ Ø°Ø®ÛŒØ±Ù‡ Ø´Ø¯:\n{path}")
        except Exception as e:
            messagebox.showerror("Ø®Ø·Ø§ Ø¯Ø± Ø°Ø®ÛŒØ±Ù‡", str(e))


# ğŸ”¹ Ø§Ø¬Ø±Ø§ÛŒ Ø¨Ø±Ù†Ø§Ù…Ù‡
if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelReportApp(root)
    root.mainloop()