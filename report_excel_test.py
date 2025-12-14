# main.py
import tkinter as tk
from ui.main_window import ExcelReportApp
import logging

def main():
    """تابع اصلی برنامه"""
    try:
        root = tk.Tk()
        app = ExcelReportApp(root)
        root.mainloop()
    except Exception as e:
        logging.error(f"Application failed to start: {e}")
        raise

if __name__ == "__main__":
    main()
    # core/excel_processor.py
import pandas as pd
from openpyxl import load_workbook
import logging
from typing import Optional, List, Dict, Any
from utils.helpers import find_column, normalize_repair_type

class ExcelProcessor:
    """کلاس اصلی برای پردازش فایل‌های اکسل"""
    
    def __init__(self):
        self.df: Optional[pd.DataFrame] = None
        self.df_normalized: Optional[pd.DataFrame] = None
        self.column_mapping: Dict[str, str] = {}
        
    def load_excel(self, file_path: str, sheet_name: str) -> bool:
        """بارگذاری فایل اکسل"""
        try:
            # استفاده از context manager برای مدیریت حافظه
            with load_workbook(file_path, read_only=True, data_only=True) as wb:
                ws = wb[sheet_name]
                data = list(ws.values)
            
            if not data:
                return False
                
            headers = [str(cell).strip() if cell else "" for cell in data[0]]
            self.df = pd.DataFrame(data[1:], columns=headers)
            
            # تشخیص خودکار ستون‌ها
            self._auto_detect_columns()
            
            # ایجاد نسخه نرمالایز شده
            self._create_normalized_data()
            
            return True
            
        except Exception as e:
            logging.error(f"Error loading Excel file: {e}")
            return False
    
    def _auto_detect_columns(self) -> None:
        """تشخیص خودکار ستون‌های مهم"""
        if self.df is None:
            return
            
        column_patterns = {
            'repair_col': ['نوع تعمیر', 'تعمیر', 'repair', 'نوع', 'کار'],
            'part_col': ['قالب', 'قطعه', 'دستگاه', 'part', 'device', 'mold'],
            'date_col': ['تاریخ', 'date', 'زمان'],
            'perf_col': ['مقدار ساعت کار شده', 'ساعت', 'hour', 'time', 'مدت'],
            'req_col': ['شماره نامه درخواست', 'شماره درخواست', 'request'],
            'code_col': ['کد قالب', 'کد', 'code']
        }
        
        for col_type, patterns in column_patterns.items():
            self.column_mapping[col_type] = find_column(self.df.columns, patterns)
    
    def _create_normalized_data(self) -> None:
        """ایجاد نسخه نرمالایز شده از داده‌ها"""
        if self.df is None:
            return
            
        self.df_normalized = self.df.copy()
        
        # نرمال‌سازی ستون نوع تعمیر
        repair_col = self.column_mapping.get('repair_col')
        if repair_col and repair_col in self.df_normalized.columns:
            self.df_normalized[repair_col] = self.df_normalized[repair_col].apply(
                lambda x: normalize_repair_type(x) if pd.notna(x) else x
            )
    
    def get_column_values(self, column_type: str, normalized: bool = False) -> List[str]:
        """دریافت مقادیر منحصر به فرد یک ستون"""
        df_source = self.df_normalized if normalized else self.df
        column_name = self.column_mapping.get(column_type)
        
        if df_source is None or column_name not in df_source.columns:
            return []
            
        return sorted(df_source[column_name].dropna().astype(str).unique())
    
    def optimize_memory(self) -> None:
        """بهینه‌سازی مصرف حافظه"""
        if self.df is not None:
            # کاهش حجم انواع داده‌ها
            for col in self.df.select_dtypes(include=['object']):
                self.df[col] = self.df[col].astype('category')
            
            for col in self.df.select_dtypes(include=['int64']):
                self.df[col] = pd.to_numeric(self.df[col], downcast='integer')
            
            for col in self.df.select_dtypes(include=['float64']):
                self.df[col] = pd.to_numeric(self.df[col], downcast='float')
    
    def clear_data(self) -> None:
        """پاکسازی داده‌ها برای آزادسازی حافظه"""
        self.df = None
        self.df_normalized = None
        self.column_mapping.clear()
        
        import gc
        gc.collect()
        # core/data_filter.py
import pandas as pd
from persiantools.jdatetime import JalaliDate
from typing import Optional, List, Dict, Any
import logging

class DataFilter:
    """کلاس برای فیلتر کردن داده‌ها"""
    
    def __init__(self, excel_processor):
        self.excel_processor = excel_processor
        self.filtered_data: Optional[pd.DataFrame] = None
        
    def apply_date_filter(self, start_date: str, end_date: str) -> bool:
        """اعمال فیلتر تاریخ"""
        try:
            if self.excel_processor.df is None:
                return False
                
            date_col = self.excel_processor.column_mapping.get('date_col')
            if not date_col:
                return False
            
            # تبدیل تاریخ‌های شمسی به میلادی
            start_gregorian = JalaliDate.strptime(start_date, "%Y/%m/%d").to_gregorian()
            end_gregorian = JalaliDate.strptime(end_date, "%Y/%m/%d").to_gregorian()
            
            # تبدیل ستون تاریخ به datetime
            self.excel_processor.df[date_col] = pd.to_datetime(
                self.excel_processor.df[date_col], errors='coerce'
            )
            
            # اعمال فیلتر
            mask = (
                (self.excel_processor.df[date_col] >= start_gregorian) & 
                (self.excel_processor.df[date_col] <= end_gregorian)
            )
            
            self.filtered_data = self.excel_processor.df[mask].copy()
            return True
            
        except Exception as e:
            logging.error(f"Error applying date filter: {e}")
            return False
    
    def apply_repair_filter(self, repair_types: List[str]) -> bool:
        """اعمال فیلتر نوع تعمیر"""
        try:
            if self.excel_processor.df_normalized is None:
                return False
                
            repair_col = self.excel_processor.column_mapping.get('repair_col')
            if not repair_col:
                return False
            
            source_df = self.filtered_data if self.filtered_data is not None else self.excel_processor.df_normalized
            
            mask = source_df[repair_col].astype(str).isin(repair_types)
            self.filtered_data = source_df[mask].copy()
            return True
            
        except Exception as e:
            logging.error(f"Error applying repair filter: {e}")
            return False
    
    def apply_hour_filter(self, min_hours: Optional[float] = None, max_hours: Optional[float] = None) -> bool:
        """اعمال فیلتر بازه ساعتی"""
        try:
            perf_col = self.excel_processor.column_mapping.get('perf_col')
            if not perf_col:
                return False
            
            source_df = self.filtered_data if self.filtered_data is not None else self.excel_processor.df
            
            # تبدیل به عدد
            source_df[perf_col] = pd.to_numeric(source_df[perf_col], errors='coerce')
            
            mask = pd.Series(True, index=source_df.index)
            
            if min_hours is not None:
                mask &= (source_df[perf_col] >= min_hours)
            if max_hours is not None:
                mask &= (source_df[perf_col] <= max_hours)
            
            self.filtered_data = source_df[mask].copy()
            return True
            
        except Exception as e:
            logging.error(f"Error applying hour filter: {e}")
            return False
    
    def group_data(self) -> Optional[pd.DataFrame]:
        """گروه‌بندی داده‌ها"""
        try:
            if self.filtered_data is None:
                return None
            
            grouping_cols = []
            for col_type in ['part_col', 'code_col', 'req_col']:
                col_name = self.excel_processor.column_mapping.get(col_type)
                if col_name and col_name in self.filtered_data.columns:
                    grouping_cols.append(col_name)
            
            if not grouping_cols:
                return None
            
            perf_col = self.excel_processor.column_mapping.get('perf_col')
            if not perf_col:
                return None
            
            # گروه‌بندی و جمع‌بندی
            grouped_df = self.filtered_data.groupby(grouping_cols, as_index=False).agg({
                perf_col: 'sum'
            }).sort_values(by=perf_col, ascending=False)
            
            return grouped_df
            
        except Exception as e:
            logging.error(f"Error grouping data: {e}")
            return None
    
    def clear_filters(self) -> None:
        """پاک کردن فیلترها"""
        self.filtered_data = None
        # core/report_generator.py
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
import os
import logging
from typing import Optional
from utils.helpers import reshape_persian_text

class ReportGenerator:
    """کلاس برای تولید گزارش در فرمت‌های مختلف"""
    
    def __init__(self, excel_processor):
        self.excel_processor = excel_processor
        
    def generate_excel_report(self, df: pd.DataFrame, output_path: str, logo_path: Optional[str] = None) -> bool:
        """تولید گزارش اکسل"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "گزارش قالب‌سازی"
            
            # افزودن لوگو
            if logo_path and os.path.exists(logo_path):
                try:
                    img = XLImage(logo_path)
                    img.width = 120
                    img.height = 120
                    ws.add_image(img, "A1")
                except Exception as e:
                    logging.warning(f"Could not add logo to Excel: {e}")
            
            # افزودن هدرها
            headers = list(df.columns)
            ws.append(headers)
            
            # افزودن داده‌ها
            for _, row in df.iterrows():
                ws.append(row.tolist())
            
            # افزودن جمع کل
            perf_col = self.excel_processor.column_mapping.get('perf_col')
            if perf_col and perf_col in df.columns:
                total_row = len(df) + 3
                ws.cell(row=total_row, column=1, value="جمع کل")
                ws.cell(row=total_row, column=headers.index(perf_col) + 1, 
                       value=df[perf_col].sum())
                
                # استایل‌دهی
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=total_row, column=col)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill("solid", fgColor="0000FF")
                    cell.alignment = Alignment(horizontal="center")
            
            # تنظیم عرض ستون‌ها
            for column_cells in ws.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)
            
            wb.save(output_path)
            return True
            
        except Exception as e:
            logging.error(f"Error generating Excel report: {e}")
            return False
    
    def generate_pdf_report(self, df: pd.DataFrame, output_path: str, 
                          logo_path: Optional[str] = None, font_name: str = 'Helvetica') -> bool:
        """تولید گزارش PDF"""
        try:
            c = canvas.Canvas(output_path, pagesize=A4)
            width, height = A4
            
            # عنوان
            c.setFont(font_name, 16)
            title = reshape_persian_text("گزارش قالب‌سازی")
            c.drawString(100, height - 50, title)
            
            # لوگو
            if logo_path and os.path.exists(logo_path):
                try:
                    c.drawImage(logo_path, width - 150, height - 120, width=100, height=100)
                except Exception as e:
                    logging.warning(f"Could not add logo to PDF: {e}")
            
            # محتوای گزارش
            y_position = height - 100
            self._add_pdf_table(c, df, y_position, font_name)
            
            c.save()
            return True
            
        except Exception as e:
            logging.error(f"Error generating PDF report: {e}")
            return False
    
    def _add_pdf_table(self, c, df: pd.DataFrame, start_y: int, font_name: str) -> None:
        """افزودن جدول به PDF"""
        # پیاده‌سازی مشابه کد اصلی
        pass
    
    def generate_csv_report(self, df: pd.DataFrame, output_path: str) -> bool:
        """تولید گزارش CSV"""
        try:
            df.to_csv(output_path, index=False, encoding='utf-8-sig')
            return True
        except Exception as e:
            logging.error(f"Error generating CSV report: {e}")
            return False
        # utils/helpers.py
import os
import re
import pandas as pd
from typing import List, Optional
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

def find_column(columns: List[str], possible_names: List[str]) -> Optional[str]:
    """پیدا کردن ستون بر اساس نام‌های احتمالی"""
    for name in possible_names:
        for col in columns:
            if name.strip().lower() in str(col).strip().lower():
                return col
    return None

def normalize_repair_type(repair_type: str) -> str:
    """نرمال‌سازی نوع تعمیر"""
    if not isinstance(repair_type, str):
        return str(repair_type)
    
    repair_type = repair_type.strip()
    repair_type = re.sub(r'[:]', '', repair_type)
    repair_type = re.sub(r'\s+', ' ', repair_type)
    
    # منطق نرمال‌سازی
    patterns = {
        'قالب تعمیری': ['قالب.*تعمیر', 'تعمیر.*قالب'],
        'قطعه تعمیری': ['قطعه.*تعمیر', 'تعمیر.*قطعه'],
        'دستگاه تعمیری': ['دستگاه.*تعمیر', 'تعمیر.*دستگاه'],
        'قالب': ['قالب'],
        'قطعه': ['قطعه'],
        'دستگاه': ['دستگاه'],
        'تعمیری': ['تعمیر']
    }
    
    for normalized_name, pattern_list in patterns.items():
        for pattern in pattern_list:
            if re.search(pattern, repair_type, re.IGNORECASE):
                return normalized_name
    
    return repair_type

def register_persian_fonts() -> str:
    """ثبت فونت‌های فارسی"""
    font_paths = [
        "C:/Windows/Fonts/arial.ttf",
        "C:/Windows/Fonts/tahoma.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
    ]
    
    for font_path in font_paths:
        if os.path.exists(font_path):
            try:
                pdfmetrics.registerFont(TTFont('PersianFont', font_path))
                return 'PersianFont'
            except:
                continue
    
    return 'Helvetica'

def reshape_persian_text(text: str) -> str:
    """اصلاح متن فارسی برای نمایش"""
    try:
        import arabic_reshaper
        from bidi.algorithm import get_display
        
        text_str = str(text)
        if all(ord(c) < 128 for c in text_str):
            return text_str
        
        reshaped_text = arabic_reshaper.reshape(text_str)
        return get_display(reshaped_text)
    except ImportError:
        return str(text)
    # config/settings.py
import os
import json
import logging
from typing import Dict, Any

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SETTINGS_PATH = os.path.join(BASE_DIR, "settings.json")

class SettingsManager:
    """مدیریت تنظیمات برنامه"""
    
    _default_settings = {
        "logo_path": "",
        "last_excel_path": "",
        "last_sheet": "",
        "window_size": "1200x800",
        "filters": {
            "start_date": "",
            "end_date": "",
            "repair_type": "",
            "part_type": ""
        },
        "colors": {
            "bg_main": "#FFA500",
            "frame_bg": "#FFE5B4",
            "button_bg": "#FF8C00", 
            "button_fg": "#FFFFFF",
            "tree_bg": "#FFFFFF",
            "tree_alt_bg": "#FFF5E0",
            "tree_font_color": "#000000",
            "tree_total_bg": "#0000FF",
            "tree_total_fg": "#FFFFFF"
        }
    }
    
    @classmethod
    def load_settings(cls) -> Dict[str, Any]:
        """بارگذاری تنظیمات"""
        try:
            if not os.path.exists(SETTINGS_PATH):
                cls._create_default_settings()
            
            with open(SETTINGS_PATH, "r", encoding="utf-8") as f:
                settings = json.load(f)
            
            # ادغام با تنظیمات پیش‌فرض
            return cls._merge_with_defaults(settings)
            
        except Exception as e:
            logging.error(f"Error loading settings: {e}")
            return cls._default_settings.copy()
    
    @classmethod
    def save_settings(cls, settings: Dict[str, Any]) -> bool:
        """ذخیره تنظیمات"""
        try:
            with open(SETTINGS_PATH, "w", encoding="utf-8") as f:
                json.dump(settings, f, ensure_ascii=False, indent=4)
            return True
        except Exception as e:
            logging.error(f"Error saving settings: {e}")
            return False
    
    @classmethod
    def _create_default_settings(cls) -> None:
        """ایجاد فایل تنظیمات پیش‌فرض"""
        os.makedirs(os.path.dirname(SETTINGS_PATH), exist_ok=True)
        cls.save_settings(cls._default_settings)
    
    @classmethod
    def _merge_with_defaults(cls, settings: Dict[str, Any]) -> Dict[str, Any]:
        """ادغام با تنظیمات پیش‌فرض"""
        result = cls._default_settings.copy()
        
        for key, value in settings.items():
            if isinstance(value, dict) and key in result and isinstance(result[key], dict):
                result[key].update(value)
            else:
                result[key] = value
        
        return result
    # ui/main_window.py
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
import logging
from core.excel_processor import ExcelProcessor
from core.data_filter import DataFilter
from core.report_generator import ReportGenerator
from core.dashboard import PowerBIDashboard
from config.settings import SettingsManager
from utils.helpers import register_persian_fonts

class ExcelReportApp:
    """کلاس اصلی رابط کاربری"""
    
    def __init__(self, root):
        self.root = root
        self.settings_manager = SettingsManager()
        self.settings = self.settings_manager.load_settings()
        
        # Initialize core components
        self.excel_processor = ExcelProcessor()
        self.data_filter = DataFilter(self.excel_processor)
        self.report_generator = ReportGenerator(self.excel_processor)
        
        self._setup_ui()
        self._load_initial_settings()
    
    def _setup_ui(self):
        """راه‌اندازی رابط کاربری"""
        # پیاده‌سازی مشابه کد اصلی اما با ساختار ماژولار
        pass
    
    def _load_initial_settings(self):
        """بارگذاری تنظیمات اولیه"""
        # بارگذاری مسیرهای ذخیره شده
        pass
    
    def _create_menu(self):
        """ایجاد منوی برنامه"""
        pass
    
    def _create_filters_frame(self):
        """ایجاد فریم فیلترها"""
        pass
    
    def _create_treeview(self):
        """ایجاد Treeview برای نمایش داده‌ها"""
        pass
    
    # متدهای اصلی با استفاده از کامپوننت‌های جدید
    def load_excel_data(self):
        """بارگذاری داده‌های اکسل"""
        try:
            file_path = self.file_entry.get().strip()
            sheet_name = self.sheet_cb.get().strip()
            
            if not file_path or not sheet_name:
                messagebox.showwarning("هشدار", "لطفاً فایل و شیت را انتخاب کنید")
                return
            
            success = self.excel_processor.load_excel(file_path, sheet_name)
            if success:
                self._populate_filters()
                self.status_var.set("داده‌ها با موفقیت بارگذاری شدند")
            else:
                messagebox.showerror("خطا", "خطا در بارگذاری داده‌ها")
                
        except Exception as e:
            logging.error(f"Error loading Excel data: {e}")
            messagebox.showerror("خطا", f"خطا در بارگذاری داده‌ها: {e}")
    
    def apply_filters(self):
        """اعمال فیلترها"""
        try:
            # جمع‌آوری پارامترهای فیلتر
            filter_params = self._collect_filter_parameters()
            
            # اعمال فیلترها
            success = self._apply_all_filters(filter_params)
            
            if success:
                self._display_filtered_data()
            else:
                messagebox.showerror("خطا", "خطا در اعمال فیلترها")
                
        except Exception as e:
            logging.error(f"Error applying filters: {e}")
            messagebox.showerror("خطا", f"خطا در اعمال فیلترها: {e}")
    
    def _collect_filter_parameters(self):
        """جمع‌آوری پارامترهای فیلتر"""
        # پیاده‌سازی منطق جمع‌آوری پارامترها
        pass
    
    def _apply_all_filters(self, params):
        """اعمال تمام فیلترها"""
        # استفاده از DataFilter برای اعمال فیلترها
        pass
    
    def _display_filtered_data(self):
        """نمایش داده‌های فیلتر شده"""
        # استفاده از Treeview برای نمایش داده‌ها
        pass
    
    def on_closing(self):
        """مدیریت بسته شدن برنامه"""
        try:
            # ذخیره تنظیمات
            self.settings_manager.save_settings(self.settings)
            
            # پاکسازی حافظه
            self.excel_processor.clear_data()
            self.data_filter.clear_filters()
            
        except Exception as e:
            logging.error(f"Error during closing: {e}")
        finally:
            self.root.destroy()
            