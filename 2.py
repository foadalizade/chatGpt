# -*- coding: utf-8 -*-
"""
report_excel_with_inspector.py
نسخه‌ای از برنامه گزارش قالبسازی که دارای Inspector متنی است:
- سمت چپ: UI (فیلترها، Treeview، ...)،
- سمت راست بالا: Text نمایش کل سورس همین فایل،
- سمت راست پایین: Text نمایش توضیحات/راهنما،
- با حرکت ماوس روی ویجت‌ها یا روی ردیف‌های Treeview: هایلایت در سورس و نمایش توضیح مربوطه.

روش کار هایلایت سورس:
    در بخش‌های ساخت UI و متدهای مهم، بلاک‌های مخصوصی گذاشته‌ام:
    # --- BLOCK: <key> START
    ...
    # --- BLOCK: <key> END

    برنامه هنگام اجرا این فایل را می‌خواند، موقعیت خط START/END را پیدا می‌کند
    و هنگام رویداد Hover آن بازه را هایلایت می‌کند.

تذکر: اگر persiantools نصب نباشد، برنامه به جای تبدیل تاریخ جلالی به میلادی،
سعی می‌کند تاریخ‌های وارد شده را به صورت میلادی بخواند.
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import os
import traceback
import sys
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image as XLImage
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from PIL import Image, ImageTk
import datetime

# Try to import persiantools.JalaliDate; if missing, fallback gracefully.
try:
    from persiantools.jdatetime import JalaliDate
except Exception:
    JalaliDate = None

# -----------------------------
# Utility: find first column matching some possible names
# -----------------------------
def find_column(columns, possible_names):
    for name in possible_names:
        for col in columns:
            if name.strip() in str(col).strip():
                return col
    return None

# -----------------------------
# Main application
# -----------------------------
class ExcelReportApp:
    def __init__(self, root):
        self.root = root
        self.root.title("گزارش قالبسازی — با Inspector")
        self.root.geometry("1200x800")
        # ساده و خوانا: پس‌زمینه تیره برای زیبایی
        self.root.configure(bg="#1f1f2e")

        self.df = None
        self.df_filtered = None
        self.file_path = None

        # مسیر دقیق لوگو (مثل فایل خودت). اگر نیست، فقط نادیده گرفته می‌شود.
        self.logo_path = r"C:\Users\f.alizadeh\OneDrive\Desktop\1\logo.png"
        print(f"Logo path: {self.logo_path} Exists: {os.path.exists(self.logo_path)}")

        # نگاشت ویجت‌ها به کلید بلاک سورس؛ مقادیر START/END در زمان اجرا از فایل خوانده می‌شود.
        self.widget_key_map = {}

        # توضیحات متنی برای هر کلید (برای info_text)
        self.explanations = self._build_explanations()

        # نگهداری توضیحات مربوط به هر آیتم Treeview (item_id -> text)
        self.tree_item_info = {}

        # بلاک‌های سورس (start_line, end_line) پر می‌شود بعد از خواندن فایل.
        self.source_blocks = {}

        # رابط کاربری شامل دو ستون: چپ UI، راست inspector
        self._build_layout()
        # حالا UI را بساز
        self.setup_ui()
        # بعد از ساختن UI، خواندن سورس این فایل برای mapping
        self._load_own_source_and_find_blocks()

    # -------------------------
    def _build_explanations(self):
        return {
            "file_entry": "مسیر فایل اکسل را وارد یا با دکمه 'انتخاب فایل' انتخاب کنید.",
            "select_file_btn": "باز کردن دیالوگ برای انتخاب فایل اکسل (.xlsx).",
            "sheet_cb": "از اینجا نام شیت (Sheet) که داده‌ها در آن است را انتخاب کنید.",
            "load_sheets_btn": "خواندن نام شیت‌ها از فایل اکسل و قرار دادن در Combobox.",
            "start_entry": "شروع بازه تاریخ (YYYY/MM/DD). از تقویم جلالی استفاده می‌شود اگر نصب باشد.",
            "end_entry": "پایان بازه تاریخ (YYYY/MM/DD).",
            "repair_cb": "فیلتر بر اساس نوع تعمیر (مقادیر از شیت خوانده می‌شود).",
            "part_cb": "فیلتر بر اساس قالب/قطعه/دستگاه.",
            "load_values_btn": "خواندن داده‌ها از شیت انتخاب شده و آماده‌سازی فیلترها.",
            "apply_filter_btn": "اعمال فیلتر ساده با تاریخ، نوع تعمیر و قطعه.",
            "smart_filter_btn": "اعمال فیلتر هوشمند روی داده‌های فعلی.",
            "save_btn": "ذخیره خروجی فیلتر شده به فرمت Excel/CSV/PDF و درج لوگو.",
            "tree": "نمایش سطرهای فیلترشده. وقتی ماوس روی هر ردیف می‌رود، توضیح مربوط به آن ردیف در پایین نمایش داده می‌شود.",
            "update_treeview": "متدی که Treeview را پاک کرده و سطرهای جدید را وارد می‌کند و جمع کل را اضافه می‌کند.",
            "save_output": "متدی که خروجی را در فرمت‌های مختلف ذخیره می‌کند و لوگو را درج می‌کند."
        }

    # -------------------------
    def _build_layout(self):
        # کلیت: یک پنجره افقی — چپ: UI، راست: inspector (کد + توضیحات)
        self.main_pane = ttk.Panedwindow(self.root, orient=tk.HORIZONTAL)
        self.main_pane.pack(fill="both", expand=True, padx=6, pady=6)

        # چپ: frame_ui
        self.frame_ui = ttk.Frame(self.main_pane, width=720)
        self.main_pane.add(self.frame_ui, weight=3)

        # راست: frame_inspector
        self.frame_inspector = ttk.Frame(self.main_pane, width=480)
        self.main_pane.add(self.frame_inspector, weight=2)

        # داخل inspector: بالا -> code_text, پایین -> info_text
        self.code_text = tk.Text(self.frame_inspector, wrap="none", font=("Consolas", 11), height=28)
        self.code_vscroll = ttk.Scrollbar(self.frame_inspector, orient=tk.VERTICAL, command=self.code_text.yview)
        self.code_text.configure(yscrollcommand=self.code_vscroll.set)

        self.code_text.pack(side="top", fill="both", expand=True, padx=4, pady=(4,2))
        self.code_vscroll.pack(side="right", fill="y")

        # توضیحات زیرین
        ttk.Label(self.frame_inspector, text="توضیحات / راهنما:", font=("Arial", 10, "bold")).pack(anchor="w", padx=6)
        self.info_text = tk.Text(self.frame_inspector, wrap="word", height=8, font=("Arial", 10))
        self.info_text.pack(fill="x", padx=6, pady=(0,6))

        # تگ هایلایت برای کد و توضیحات
        self.code_text.tag_config("code_highlight", background="#fff59d")  # ملایم زرد
        self.info_text.tag_config("info_highlight", background="#fff59d")

    # -------------------------
    def setup_ui(self):
        # --- BLOCK: top_frame START
        top_frame = ttk.Frame(self.frame_ui)
        top_frame.pack(fill="x", padx=10, pady=4)
        tk.Label(top_frame, text="کارشناس برنامه‌ریزی و ساخت قالبسازی: فواد مطور علیزاده",
                 font=("Arial", 11, "bold")).pack(side="left", padx=10)
        # --- BLOCK: top_frame END
        # map widget key
        # (mapping to source blocks will be found later by searching markers in file)
        # --- BLOCK: filters START
        # فیلتر ساده
        frame_simple = ttk.LabelFrame(self.frame_ui, text="فیلتر ساده", padding=8)
        frame_simple.pack(padx=10, pady=6, fill="x")

        ttk.Label(frame_simple, text="مسیر فایل اکسل:").grid(row=0, column=0, sticky="w")
        self.file_entry = ttk.Entry(frame_simple, width=70)
        self.file_entry.grid(row=0, column=1, padx=5)
        self.file_entry._inspector_key = "file_entry"

        self.btn_select_file = ttk.Button(frame_simple, text="انتخاب فایل", command=self.select_file)
        self.btn_select_file.grid(row=0, column=2)
        self.btn_select_file._inspector_key = "select_file_btn"

        ttk.Label(frame_simple, text="نام شیت:").grid(row=1, column=0, sticky="w")
        self.sheet_cb = ttk.Combobox(frame_simple, width=30, state="readonly")
        self.sheet_cb.grid(row=1, column=1, sticky="w")
        self.sheet_cb._inspector_key = "sheet_cb"

        self.btn_load_sheets = ttk.Button(frame_simple, text="بارگذاری شیت‌ها", command=self.load_sheets)
        self.btn_load_sheets.grid(row=1, column=2)
        self.btn_load_sheets._inspector_key = "load_sheets_btn"

        ttk.Label(frame_simple, text="تاریخ شروع (YYYY/MM/DD):").grid(row=2, column=0, sticky="w")
        self.start_entry = ttk.Entry(frame_simple, width=15)
        self.start_entry.grid(row=2, column=1, sticky="w")
        self.start_entry._inspector_key = "start_entry"

        ttk.Label(frame_simple, text="تاریخ پایان (YYYY/MM/DD):").grid(row=3, column=0, sticky="w")
        self.end_entry = ttk.Entry(frame_simple, width=15)
        self.end_entry.grid(row=3, column=1, sticky="w")
        self.end_entry._inspector_key = "end_entry"

        ttk.Label(frame_simple, text="نوع تعمیر:").grid(row=4, column=0, sticky="w")
        self.repair_cb = ttk.Combobox(frame_simple, width=30, state="readonly")
        self.repair_cb.grid(row=4, column=1, sticky="w")
        self.repair_cb._inspector_key = "repair_cb"

        ttk.Label(frame_simple, text="قالب / قطعه / دستگاه:").grid(row=5, column=0, sticky="w")
        self.part_cb = ttk.Combobox(frame_simple, width=30, state="readonly")
        self.part_cb.grid(row=5, column=1, sticky="w")
        self.part_cb._inspector_key = "part_cb"

        self.btn_load_values = ttk.Button(frame_simple, text="📂 بارگذاری داده‌ها", command=self.load_values)
        self.btn_load_values.grid(row=6, column=0, pady=6)
        self.btn_load_values._inspector_key = "load_values_btn"

        style = ttk.Style()
        style.configure("Green.TButton", foreground="black")
        self.btn_apply_filter = ttk.Button(frame_simple, text="🔍 اعمال فیلتر ساده", command=self.apply_simple_filter, style="Green.TButton")
        self.btn_apply_filter.grid(row=6, column=1, pady=6)
        self.btn_apply_filter._inspector_key = "apply_filter_btn"

        self.btn_save = ttk.Button(frame_simple, text="💾 ذخیره", command=lambda: self.save_output(self.df_filtered), style="Green.TButton")
        self.btn_save.grid(row=6, column=2, pady=6)
        self.btn_save._inspector_key = "save_btn"
        # --- BLOCK: filters END

        # لوگو سمت راست بالای UI frame (اندازه 140x140)
        if os.path.exists(self.logo_path):
            try:
                img = Image.open(self.logo_path)
                img = img.resize((140, 140), Image.Resampling.LANCZOS)
                self.tk_img = ImageTk.PhotoImage(img)
                lbl_logo = tk.Label(top_frame, image=self.tk_img)
                lbl_logo.pack(side="right", padx=6)
                # give inspector key (logo is part of top_frame block anyway)
                lbl_logo._inspector_key = "logo"
            except Exception as e:
                print("Error loading logo:", e)

        # فیلتر هوشمند (مجزا)
        frame_smart = ttk.LabelFrame(self.frame_ui, text="فیلتر هوشمند", padding=8)
        frame_smart.pack(padx=10, pady=6, fill="x")
        self.btn_smart = ttk.Button(frame_smart, text="🔍 اعمال فیلتر هوشمند", command=self.apply_smart_filter, style="Green.TButton")
        self.btn_smart.grid(row=0, column=0, pady=6)
        self.btn_smart._inspector_key = "smart_filter_btn"

        self.btn_save2 = ttk.Button(frame_smart, text="💾 ذخیره", command=lambda: self.save_output(self.df_filtered), style="Green.TButton")
        self.btn_save2.grid(row=0, column=1, pady=6)
        self.btn_save2._inspector_key = "save_btn2"

        # نمایش نتایج - Treeview
        # --- BLOCK: tree START
        cols = ("نوع تعمیر", "قالب/قطعه/دستگاه", "شماره نامه درخواست", "کد قالب", "مقدار ساعت کار شده")
        self.tree = ttk.Treeview(self.frame_ui, columns=cols, show="headings", height=16)
        for col in cols:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=180, anchor="center")
        self.tree.pack(padx=10, pady=8, fill="both", expand=True)
        self.tree._inspector_key = "tree"
        # --- BLOCK: tree END

        # Bind events for inspector:
        # For main widgets, bind <Enter> / <Leave> to show explanation and highlight code.
        widgets = [
            self.file_entry, self.btn_select_file, self.sheet_cb, self.btn_load_sheets,
            self.start_entry, self.end_entry, self.repair_cb, self.part_cb,
            self.btn_load_values, self.btn_apply_filter, self.btn_save, self.btn_smart, self.tree
        ]
        for w in widgets:
            try:
                w.bind("<Enter>", self.on_widget_enter, add="+")
                w.bind("<Leave>", self.on_widget_leave, add="+")
            except Exception:
                pass

        # For Treeview rows, also bind Motion to give per-row hover
        self.tree.bind("<Motion>", self.on_tree_motion)

    # ------------------------- IO: خواندن سورس همین فایل و پیدا کردن بلاک‌ها
    def _load_own_source_and_find_blocks(self):
        try:
            src_path = os.path.abspath(__file__)
            with open(src_path, "r", encoding="utf-8") as f:
                src = f.read()
            # نمایش کامل سورس در code_text
            self.code_text.delete("1.0", "end")
            self.code_text.insert("1.0", src)
            # اکنون بلاک‌های مارکر را پیدا کن
            lines = src.splitlines()
            markers = {}
            for idx, line in enumerate(lines, start=1):
                line_stripped = line.strip()
                # markers of form: # --- BLOCK: <key> START / END
                if line_stripped.startswith("# --- BLOCK:"):
                    parts = line_stripped.split()
                    # expected: ['#', '---', 'BLOCK:', '<key>', 'START']
                    if len(parts) >= 5:
                        key = parts[3]
                        state = parts[4]
                        if key not in markers:
                            markers[key] = {}
                        markers[key][state] = idx
            # build source_blocks from markers
            for key, val in markers.items():
                start = val.get("START", None)
                end = val.get("END", None)
                if start and end:
                    self.source_blocks[key] = (start, end)
            # Additionally map some logical keys to function blocks if explicit markers not present
            # For important methods, search by def name
            func_names = ["update_treeview", "save_output", "load_values", "apply_simple_filter", "apply_smart_filter", "setup_ui"]
            for fn in func_names:
                if fn in self.source_blocks:
                    continue
                # find def line
                for idx, line in enumerate(lines, start=1):
                    if line.lstrip().startswith(f"def {fn}("):
                        # find end: next blank line followed by def or end of file; simple heuristic:
                        start = idx
                        end = start
                        for j in range(start+1, len(lines)+1):
                            # end on next "def " at column 1 or EOF
                            if lines[j-1].lstrip().startswith("def ") and j-1 != idx-1:
                                end = j-1
                                break
                            end = j
                        self.source_blocks[fn] = (start, end)
                        break
            # Map widget inspector keys to source block keys
            # Try mapping common keys to BLOCK names used in setup_ui
            # If not found, fall back to mapping to 'setup_ui' or specific function names
            default_map = {
                "file_entry": "filters",
                "select_file_btn": "filters",
                "sheet_cb": "filters",
                "load_sheets_btn": "filters",
                "start_entry": "filters",
                "end_entry": "filters",
                "repair_cb": "filters",
                "part_cb": "filters",
                "load_values_btn": "filters",
                "apply_filter_btn": "filters",
                "save_btn": "filters",
                "save_btn2": "filters",
                "smart_filter_btn": "filters",
                "tree": "tree",
                "update_treeview": "update_treeview",
                "save_output": "save_output"
            }
            for widget_key, block_key in default_map.items():
                if block_key in self.source_blocks:
                    self.widget_key_map[widget_key] = self.source_blocks[block_key]
            # also allow mapping by identical keys if exist
            for key in self.source_blocks:
                if key not in self.widget_key_map:
                    self.widget_key_map[key] = self.source_blocks[key]
        except Exception as e:
            print("[_load_own_source_and_find_blocks] error:", e)
            traceback.print_exc()

    # -------------------------
    # Event handlers for inspector
    def on_widget_enter(self, event):
        w = event.widget
        key = getattr(w, "_inspector_key", None)
        if key is None:
            # try to infer key from widget type
            if isinstance(w, ttk.Treeview):
                key = "tree"
        # Show explanation
        expl = self.explanations.get(key, "")
        self._show_info_text(expl)
        # Highlight code block if exists
        blk = self.widget_key_map.get(key)
        if blk:
            start, end = blk
            self._highlight_code_block(start, end)
        else:
            # fallback: clear highlight
            self._clear_code_highlight()

    def on_widget_leave(self, event):
        # clear info and code highlight
        # but keep last info for tree hover if still over row
        self._clear_info_highlight()
        self._clear_code_highlight()

    def on_tree_motion(self, event):
        # identify row
        rowid = self.tree.identify_row(event.y)
        if rowid:
            # get values and create a short description
            vals = self.tree.item(rowid, "values")
            desc = self._make_row_description(vals)
            # show in info_text and highlight generic update_treeview block
            self._show_info_text(desc)
            # highlight update_treeview block if exists
            blk = self.source_blocks.get("update_treeview") or self.source_blocks.get("tree")
            if blk:
                self._highlight_code_block(blk[0], blk[1])
            # also, if we stored per-item info, show it
            if rowid in self.tree_item_info:
                self._show_info_text(self.tree_item_info[rowid])
        else:
            # not on a row
            self._clear_info_highlight()
            self._clear_code_highlight()

    # -------------------------
    # helpers for info_text and code_text highlighting
    def _show_info_text(self, text):
        self.info_text.config(state="normal")
        self.info_text.delete("1.0", "end")
        if text:
            self.info_text.insert("1.0", text)
            # highlight entire info_text
            self.info_text.tag_add("info_highlight", "1.0", "end")
        self.info_text.config(state="disabled")

    def _clear_info_highlight(self):
        self.info_text.config(state="normal")
        self.info_text.delete("1.0", "end")
        self.info_text.config(state="disabled")

    def _highlight_code_block(self, start_line, end_line):
        try:
            # remove old
            self._clear_code_highlight()
            # create tag across lines
            start_idx = f"{start_line}.0"
            end_idx = f"{end_line}.0"
            self.code_text.tag_add("code_highlight", start_idx, end_idx)
            # scroll to start line
            self.code_text.see(start_idx)
        except Exception as e:
            print("[_highlight_code_block] error:", e)

    def _clear_code_highlight(self):
        self.code_text.tag_remove("code_highlight", "1.0", "end")

    # -------------------------
    # UI functionality (مشابه برنامه پیشین)
    def select_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx"), ("All", "*.*")])
        if not path:
            return
        self.file_entry.delete(0, tk.END)
        self.file_entry.insert(0, path)
        self.file_path = path

    def load_sheets(self):
        path = self.file_entry.get().strip()
        if not path or not os.path.exists(path):
            messagebox.showerror("خطا", "فایل معتبر انتخاب نشده است.")
            return
        try:
            wb = load_workbook(path, read_only=True)
            self.sheet_cb["values"] = wb.sheetnames
            wb.close()
            messagebox.showinfo("انجام شد", "شیت‌ها با موفقیت بارگذاری شدند.")
        except Exception as e:
            messagebox.showerror("خطا", str(e))

    def load_values(self):
        path = self.file_entry.get().strip()
        sel_sheet = self.sheet_cb.get().strip()
        if not path or not sel_sheet:
            messagebox.showerror("خطا", "فایل یا شیت را انتخاب کنید.")
            return
        try:
            wb = load_workbook(path, data_only=True, read_only=True)
            ws = wb[sel_sheet]
            rows = list(ws.values)
            wb.close()
            header = [str(x).strip() if x else "" for x in rows[0]]
            df = pd.DataFrame(rows[1:], columns=header)
            df.columns = [str(c).strip() for c in df.columns]

            self.df = df

            self.repair_col = find_column(df.columns, ["نوع تعمیر", "تعمیر"])
            self.part_col = find_column(df.columns, ["قالب / قطعه / دستگاه", "قالب"])
            self.date_col = find_column(df.columns, ["تاریخ"])
            self.perf_col = find_column(df.columns, ["مقدار ساعت کار شده", "ساعت"])
            self.req_col = find_column(df.columns, ["شماره نامه درخواست", "شماره درخواست"])
            self.code_col = find_column(df.columns, ["کد قالب", "کد"])

            # fill comboboxes safely
            try:
                if self.repair_col in df.columns:
                    self.repair_cb["values"] = ["(همه)"] + sorted(df[self.repair_col].dropna().astype(str).unique())
                if self.part_col in df.columns:
                    self.part_cb["values"] = ["(همه)"] + sorted(df[self.part_col].dropna().astype(str).unique())
            except Exception:
                pass

            messagebox.showinfo("انجام شد", "اطلاعات با موفقیت بارگذاری شد.")
        except Exception as e:
            messagebox.showerror("خطا", str(e))

    def apply_simple_filter(self):
        if self.df is None:
            messagebox.showwarning("هشدار", "ابتدا داده‌ها را بارگذاری کنید.")
            return
        df = self.df.copy()
        s = self.start_entry.get().strip()
        e = self.end_entry.get().strip()
        if s and e and self.date_col:
            # اگر JalaliDate در دسترس باشد تبدیل انجام شود
            if JalaliDate is not None:
                try:
                    s_g = JalaliDate.strptime(s, "%Y/%m/%d").to_gregorian()
                    e_g = JalaliDate.strptime(e, "%Y/%m/%d").to_gregorian()
                    df[self.date_col] = pd.to_datetime(df[self.date_col], errors="coerce")
                    df = df[(df[self.date_col] >= s_g) & (df[self.date_col] <= e_g)]
                except Exception:
                    pass
            else:
                # fallback: تلاش برای parse تاریخ به میلادی
                try:
                    s_g = pd.to_datetime(s, errors="coerce")
                    e_g = pd.to_datetime(e, errors="coerce")
                    df[self.date_col] = pd.to_datetime(df[self.date_col], errors="coerce")
                    if pd.notna(s_g) and pd.notna(e_g):
                        df = df[(df[self.date_col] >= s_g) & (df[self.date_col] <= e_g)]
                except Exception:
                    pass

        sel_repair = self.repair_cb.get()
        if sel_repair and sel_repair != "(همه)":
            try:
                df = df[df[self.repair_col].astype(str) == sel_repair]
            except Exception:
                pass

        sel_part = self.part_cb.get()
        if sel_part and sel_part != "(همه)":
            try:
                df = df[df[self.part_col].astype(str) == sel_part]
            except Exception:
                pass

        # ensure performance column numeric
        try:
            df[self.perf_col] = pd.to_numeric(df[self.perf_col], errors="coerce").fillna(0)
        except Exception:
            pass

        self.df_filtered = df
        self.update_treeview(df)

    def apply_smart_filter(self):
        if self.df_filtered is None:
            self.apply_simple_filter()
        df = self.df_filtered.copy()
        for col in df.columns:
            df[col] = df[col].fillna("(خالی)")
        # simple heuristic that keeps only values that occur (this is a placeholder for smarter logic)
        self.df_filtered = df
        self.update_treeview(df)

    def _make_row_description(self, values_tuple):
        try:
            typ, part, req, code, hours = values_tuple
            return f"نوع تعمیر: {typ}\nقالب/قطعه: {part}\nشماره نامه: {req}\nکد قالب: {code}\nمقدار ساعت کار شده: {hours}"
        except Exception:
            return str(values_tuple)

    def update_treeview(self, df):
        # پاک‌سازی
        for i in self.tree.get_children():
            self.tree.delete(i)
        # درج سطرها و ذخیره توضیحات برای هر آیتم
        for _, row in df.iterrows():
            vals = (
                row.get(self.repair_col, ""),
                row.get(self.part_col, ""),
                row.get(self.req_col, ""),
                row.get(self.code_col, ""),
                row.get(self.perf_col, 0)
            )
            item = self.tree.insert("", "end", values=vals)
            # store info text for hover
            self.tree_item_info[item] = self._make_row_description(vals)
        # جمع کل
        try:
            total = df[self.perf_col].sum()
        except Exception:
            total = 0
        total_item = self.tree.insert("", "end", values=("جمع کل", "", "", "", total))
        self.tree_item_info[total_item] = "این ردیف نشان‌دهنده مجموع کل مقدار ساعت کار شده است."

    def save_output(self, df):
        if df is None or df.empty:
            messagebox.showerror("خطا", "ابتدا فیلتر را اعمال کنید.")
            return
        path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                            filetypes=[("Excel", "*.xlsx"), ("CSV", "*.csv"), ("PDF", "*.pdf")])
        if not path:
            return
        df_out = df.copy()
        # اضافه کردن ردیف جمع کل به انتها با index مشخص
        try:
            df_out.loc["جمع کل"] = [""] * len(df_out.columns)
            if self.perf_col in df_out.columns:
                df_out.at["جمع کل", self.perf_col] = df_out[self.perf_col].sum()
        except Exception:
            pass

        try:
            if path.endswith(".xlsx"):
                wb = Workbook()
                ws = wb.active
                ws.title = "گزارش"
                if os.path.exists(self.logo_path):
                    try:
                        img = XLImage(self.logo_path)
                        img.width = 140
                        img.height = 140
                        ws.add_image(img, "A1")
                    except Exception:
                        pass
                ws.append(list(df_out.columns))
                for r in df_out.itertuples(index=False):
                    ws.append(list(r))
                wb.save(path)
            elif path.endswith(".csv"):
                df_out.to_csv(path, index=False, encoding="utf-8-sig")
            elif path.endswith(".pdf"):
                c = canvas.Canvas(path, pagesize=A4)
                c.setFont("Helvetica", 10)
                y = 800
                if os.path.exists(self.logo_path):
                    try:
                        c.drawImage(self.logo_path, 50, y - 50, width=140, height=140)
                    except Exception:
                        pass
                for _, row in df_out.iterrows():
                    text = " | ".join([str(x) for x in row.values if x is not None])
                    c.drawString(40, y, text)
                    y -= 14
                    if y < 50:
                        c.showPage()
                        c.setFont("Helvetica", 10)
                        y = 800
                c.drawString(50, 30, "F.Alizadeh")
                c.save()
            messagebox.showinfo("ذخیره شد", f"فایل ذخیره شد:\n{path}")
        except Exception as e:
            messagebox.showerror("خطا در ذخیره", str(e))

# -----------------------------
# اجرا
# -----------------------------
if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelReportApp(root)
    root.mainloop()
