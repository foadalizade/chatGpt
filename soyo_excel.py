# report_excel.py
# -*- coding: utf-8 -*-

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import os
import json
from persiantools.jdatetime import JalaliDate
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from PIL import Image, ImageTk
import traceback

# -----------------------------
# مسیر فایل settings را کنار خود اسکریپت تعریف می‌کنیم
# -----------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SETTINGS_PATH = os.path.join(BASE_DIR, "settings.json")
DEFAULT_LOGO = os.path.join(BASE_DIR, "Logo.png")  # مسیر پیش‌فرض لوگو

# -----------------------------
def load_settings():
    """لود کردن تنظیمات از فایل JSON"""
    default = {
        "logo_path": DEFAULT_LOGO,
        "last_excel_path": "",
        "last_sheet": "",
        "window_size": "1200x800",
        "filters": {
            "start_date": "",
            "end_date": "",
            "repair_type": "",
            "part_type": ""
        }
    }
    try:
        if not os.path.exists(SETTINGS_PATH):
            with open(SETTINGS_PATH, "w", encoding="utf-8") as f:
                json.dump(default, f, ensure_ascii=False, indent=4)
            print(f"[settings] created default settings at: {SETTINGS_PATH}")
            return default
        with open(SETTINGS_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
        # اطمینان از وجود کلیدها
        for k, v in default.items():
            if k not in data:
                data[k] = v
        if "filters" not in data or not isinstance(data["filters"], dict):
            data["filters"] = default["filters"]
        return data
    except Exception as e:
        print("[settings] error loading settings:", e)
        traceback.print_exc()
        return default

def save_settings(data):
    """ذخیره تنظیمات در فایل JSON"""
    try:
        with open(SETTINGS_PATH, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
        print(f"[settings] saved to: {SETTINGS_PATH}")
    except Exception as e:
        print("[settings] error saving settings:", e)
        traceback.print_exc()

# -----------------------------
def find_column(columns, possible_names):
    """پیدا کردن ستون با نام‌های احتمالی"""
    for name in possible_names:
        for col in columns:
            if name.strip() in str(col).strip():
                return col
    return None

# -----------------------------
class ExcelReportApp:
    def __init__(self, root):
        self.root = root
        self.settings = load_settings()

        self.root.title("گزارش قالبسازی")
        geom = self.settings.get("window_size", "1200x800")
        try:
            self.root.geometry(geom)
        except:
            self.root.geometry("1200x800")

        self.root.configure(bg="#f5f5f5")

        self.df = None
        self.df_filtered = None
        self.logo_path = self.settings.get("logo_path", DEFAULT_LOGO)

        self.create_menu()
        self.setup_ui()
        self.load_saved_fields()

        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

    # -----------------------------
    def create_menu(self):
        menubar = tk.Menu(self.root)
        file_menu = tk.Menu(menubar, tearoff=0)
        file_menu.add_command(label="ذخیره تنظیمات", command=lambda: save_settings(self.settings))
        file_menu.add_command(label="نمایش تنظیمات", command=self.debug_show_settings)
        file_menu.add_separator()
        file_menu.add_command(label="❌ خروج", command=self.root.quit)
        menubar.add_cascade(label="فایل", menu=file_menu)
        self.root.config(menu=menubar)

    def debug_show_settings(self):
        messagebox.showinfo("settings.json", json.dumps(self.settings, ensure_ascii=False, indent=4))

    # -----------------------------
    def setup_ui(self):
        top_frame = ttk.Frame(self.root)
        top_frame.pack(fill="x", padx=10, pady=5)

        tk.Label(top_frame,
                 text="کارشناس برنامه‌ریزی و ساخت قالبسازی: فواد مطور علیزاده",
                 font=("Arial", 11, "bold")).pack(side="left", padx=10)

        # نمایش لوگو
        if self.logo_path and os.path.exists(self.logo_path):
            try:
                img = Image.open(self.logo_path)
                img = img.resize((120, 120), Image.Resampling.LANCZOS)
                self.tk_img = ImageTk.PhotoImage(img)
                tk.Label(top_frame, image=self.tk_img).pack(side="right", padx=10)
            except Exception as e:
                print("[logo] error loading:", e)
                traceback.print_exc()

        # --- فیلتر ساده ---
        frame_simple = ttk.LabelFrame(self.root, text="فیلتر ساده", padding=10)
        frame_simple.pack(fill="x", padx=10, pady=5)

        ttk.Label(frame_simple, text="مسیر فایل اکسل:").grid(row=0, column=0, sticky="w")
        self.file_entry = ttk.Entry(frame_simple, width=70)
        self.file_entry.grid(row=0, column=1, padx=5)
        ttk.Button(frame_simple, text="انتخاب فایل", command=self.select_file).grid(row=0, column=2)

        ttk.Label(frame_simple, text="نام شیت:").grid(row=1, column=0, sticky="w")
        self.sheet_cb = ttk.Combobox(frame_simple, width=30, state="readonly")
        self.sheet_cb.grid(row=1, column=1, sticky="w")
        ttk.Button(frame_simple, text="بارگذاری شیت‌ها", command=self.load_sheets).grid(row=1, column=2)

        ttk.Label(frame_simple, text="تاریخ شروع (YYYY/MM/DD):").grid(row=2, column=0, sticky="w")
        self.start_entry = ttk.Entry(frame_simple, width=15)
        self.start_entry.grid(row=2, column=1, sticky="w")

        ttk.Label(frame_simple, text="تاریخ پایان (YYYY/MM/DD):").grid(row=3, column=0, sticky="w")
        self.end_entry = ttk.Entry(frame_simple, width=15)
        self.end_entry.grid(row=3, column=1, sticky="w")

        ttk.Label(frame_simple, text="نوع تعمیر:").grid(row=4, column=0, sticky="w")
        self.repair_cb = ttk.Combobox(frame_simple, width=30, state="readonly")
        self.repair_cb.grid(row=4, column=1, sticky="w")

        ttk.Label(frame_simple, text="قالب / قطعه / دستگاه:").grid(row=5, column=0, sticky="w")
        self.part_cb = ttk.Combobox(frame_simple, width=30, state="readonly")
        self.part_cb.grid(row=5, column=1, sticky="w")

        ttk.Button(frame_simple, text="📂 بارگذاری داده‌ها", command=self.load_values).grid(row=6, column=0, pady=5)
        ttk.Button(frame_simple, text="🔍 اعمال فیلتر ساده", command=self.apply_simple_filter).grid(row=6, column=1, pady=5)
        ttk.Button(frame_simple, text="💾 ذخیره", command=lambda: self.save_output(self.df_filtered)).grid(row=6, column=2, pady=5)

        # --- فیلتر هوشمند ---
        frame_smart = ttk.LabelFrame(self.root, text="فیلتر هوشمند", padding=10)
        frame_smart.pack(fill="x", padx=10, pady=5)

        ttk.Button(frame_smart, text="🔍 اعمال فیلتر هوشمند", command=self.apply_smart_filter).grid(row=0, column=0, pady=5)
        ttk.Button(frame_smart, text="💾 ذخیره", command=lambda: self.save_output(self.df_filtered)).grid(row=0, column=1, pady=5)

        # --- نمایش نتایج ---
        self.tree = ttk.Treeview(self.root,
                                 columns=("نوع تعمیر", "قالب/قطعه/دستگاه", "شماره نامه درخواست", "کد قالب", "مقدار ساعت کار شده"),
                                 show="headings", height=18)
        for col in self.tree["columns"]:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=200, anchor="center")
        self.tree.pack(padx=10, pady=10, fill="both", expand=True)

    # -----------------------------
    def load_saved_fields(self):
        last_path = self.settings.get("last_excel_path", "")
        if last_path:
            self.file_entry.delete(0, tk.END)
            self.file_entry.insert(0, last_path)

        filt = self.settings.get("filters", {})
        self.start_entry.delete(0, tk.END)
        self.start_entry.insert(0, filt.get("start_date", ""))
        self.end_entry.delete(0, tk.END)
        self.end_entry.insert(0, filt.get("end_date", ""))

    # -----------------------------
    def select_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx"), ("All", "*.*")])
        if not path:
            return
        self.file_entry.delete(0, tk.END)
        self.file_entry.insert(0, path)
        self.settings["last_excel_path"] = path
        save_settings(self.settings)

    def load_sheets(self):
        path = self.file_entry.get().strip()
        if not path or not os.path.exists(path):
            messagebox.showerror("خطا", "فایل یافت نشد")
            return
        try:
            wb = load_workbook(path, read_only=True)
            sheetnames = wb.sheetnames[:]
            wb.close()
            self.sheet_cb["values"] = sheetnames
            last_sheet = self.settings.get("last_sheet", "")
            if last_sheet in sheetnames:
                self.sheet_cb.set(last_sheet)
            messagebox.showinfo("انجام شد", "شیت‌ها بارگذاری شدند.")
        except Exception as e:
            print("[load_sheets] error:", e)
            traceback.print_exc()
            messagebox.showerror("خطا", str(e))

    # -----------------------------
    def load_values(self):
        path = self.file_entry.get().strip()
        sheet = self.sheet_cb.get().strip()
        if not path or not sheet or not os.path.exists(path):
            messagebox.showerror("خطا", "لطفاً فایل و شیت را انتخاب کنید.")
            return
        try:
            wb = load_workbook(path, data_only=True, read_only=True)
            ws = wb[sheet]
            rows = list(ws.values)
            wb.close()
            if not rows:
                messagebox.showerror("خطا", "شیت انتخاب‌شده خالی است.")
                return
            headers = [str(x).strip() if x else "" for x in rows[0]]
            df = pd.DataFrame(rows[1:], columns=headers)
            self.df = df

            # ذخیره شیت انتخابی
            self.settings["last_sheet"] = sheet
            save_settings(self.settings)

            # پیدا کردن ستون‌ها
            self.repair_col = find_column(df.columns, ["نوع تعمیر", "تعمیر"])
            self.part_col = find_column(df.columns, ["قالب / قطعه / دستگاه", "قالب"])
            self.date_col = find_column(df.columns, ["تاریخ"])
            self.perf_col = find_column(df.columns, ["مقدار ساعت کار شده", "ساعت"])
            self.req_col = find_column(df.columns, ["شماره نامه درخواست", "شماره درخواست"])
            self.code_col = find_column(df.columns, ["کد قالب", "کد"])

            # پر کردن کمبوباکس‌ها
            if self.repair_col in df.columns:
                self.repair_cb["values"] = ["(همه)"] + sorted(df[self.repair_col].dropna().astype(str).unique())
            if self.part_col in df.columns:
                self.part_cb["values"] = ["(همه)"] + sorted(df[self.part_col].dropna().astype(str).unique())

            messagebox.showinfo("موفق", "اطلاعات بارگذاری شد.")
        except Exception as e:
            print("[load_values] error:", e)
            traceback.print_exc()
            messagebox.showerror("خطا", str(e))

    # -----------------------------
    def apply_simple_filter(self):
        if self.df is None:
            messagebox.showwarning("هشدار", "ابتدا داده‌ها را بارگذاری کنید.")
            return
        df = self.df.copy()
        s = self.start_entry.get().strip()
        e = self.end_entry.get().strip()
        self.settings["filters"]["start_date"] = s
        self.settings["filters"]["end_date"] = e

        if s and e and self.date_col in df.columns:
            try:
                s_g = JalaliDate.strptime(s, "%Y/%m/%d").to_gregorian()
                e_g = JalaliDate.strptime(e, "%Y/%m/%d").to_gregorian()
                df[self.date_col] = pd.to_datetime(df[self.date_col], errors="coerce")
                df = df[(df[self.date_col] >= s_g) & (df[self.date_col] <= e_g)]
            except Exception as exc:
                print("[apply_simple_filter] date filter error:", exc)

        rep = self.repair_cb.get()
        if rep and rep != "(همه)" and self.repair_col in df.columns:
            df = df[df[self.repair_col].astype(str) == rep]
            self.settings["filters"]["repair_type"] = rep

        part = self.part_cb.get()
        if part and part != "(همه)" and self.part_col in df.columns:
            df = df[df[self.part_col].astype(str) == part]
            self.settings["filters"]["part_type"] = part

        save_settings(self.settings)

        if self.perf_col in df.columns:
            df[self.perf_col] = pd.to_numeric(df[self.perf_col], errors="coerce").fillna(0)

        self.df_filtered = df
        self.update_treeview(df)

    # -----------------------------
    def apply_smart_filter(self):
        if self.df_filtered is None:
            self.apply_simple_filter()
        df = self.df_filtered.copy()
        for col in df.columns:
            df[col] = df[col].fillna("(خالی)")
        self.df_filtered = df
        self.update_treeview(df)

    # -----------------------------
    def update_treeview(self, df):
        self.tree.delete(*self.tree.get_children())
        for _, row in df.iterrows():
            self.tree.insert("", "end", values=(
                row.get(self.repair_col, ""),
                row.get(self.part_col, ""),
                row.get(self.req_col, ""),
                row.get(self.code_col, ""),
                row.get(self.perf_col, 0)
            ))
        try:
            if self.perf_col in df.columns:
                total = df[self.perf_col].sum()
                self.tree.insert("", "end", values=("جمع کل", "", "", "", total))
        except Exception as e:
            print("[update_treeview] error calculating total:", e)

    # -----------------------------
    def save_output(self, df):
        if df is None or df.empty:
            messagebox.showerror("خطا", "ابتدا فیلتر را اعمال کنید.")
            return
        path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                            filetypes=[("Excel", "*.xlsx"), ("CSV", "*.csv"), ("PDF", "*.pdf")])
        if not path:
            return
        df_out = df.copy()
        try:
            df_out.loc["جمع کل"] = [""] * len(df_out.columns)
            if self.perf_col in df_out.columns:
                df_out.at["جمع کل", self.perf_col] = df_out[self.perf_col].sum()
        except Exception as e:
            print("[save_output] preparing df_out error:", e)

        try:
            if path.endswith(".xlsx"):
                wb = Workbook()
                ws = wb.active
                if self.logo_path and os.path.exists(self.logo_path):
                    try:
                        img = XLImage(self.logo_path)
                        img.width = 120
                        img.height = 120
                        ws.add_image(img, "H1")
                    except Exception as e:
                        print("[save_output] logo add error:", e)
                ws.append(list(df_out.columns))
                for r in df_out.itertuples(index=False):
                    ws.append(list(r))
                last_row = ws.max_row
                for col in range(1, ws.max_column + 1):
                    c = ws.cell(row=last_row, column=col)
                    c.font = Font(bold=True, color="FFFFFF")
                    c.fill = PatternFill("solid", fgColor="0000FF")
                    c.alignment = Alignment(horizontal="center")
                wb.save(path)

            elif path.endswith(".csv"):
                df_out.to_csv(path, index=False, encoding="utf-8-sig")

            elif path.endswith(".pdf"):
                c = canvas.Canvas(path, pagesize=A4)
                c.setFont("Helvetica", 10)
                y = 800
                if self.logo_path and os.path.exists(self.logo_path):
                    c.drawImage(self.logo_path, 450, y - 120, width=120, height=120)
                for _, row in df_out.iterrows():
                    text = " | ".join([str(x) for x in row.values])
                    c.drawString(40, y, text)
                    y -= 14
                    if y < 50:
                        c.showPage()
                        c.setFont("Helvetica", 10)
                        y = 800
                c.setFont("Helvetica-Bold", 12)
                if self.perf_col in df_out.columns:
                    c.drawString(40, 20, f"جمع کل: {df_out[self.perf_col].sum()}")
                c.save()

            messagebox.showinfo("ذخیره شد", "فایل با موفقیت ذخیره شد.")
        except Exception as e:
            print("[save_output] error:", e)
            traceback.print_exc()
            messagebox.showerror("خطا در ذخیره", str(e))

    # -----------------------------
    def on_close(self):
        try:
            self.settings["window_size"] = self.root.geometry()
            save_settings(self.settings)
        except Exception as e:
            print("[on_close] save error:", e)
        self.root.destroy()

# -----------------------------
if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelReportApp(root)
    root.mainloop()


