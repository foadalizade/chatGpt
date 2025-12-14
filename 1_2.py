# -*- coding: utf-8 -*-
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import os
from persiantools.jdatetime import JalaliDate
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image as XLImage
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from PIL import Image, ImageTk
import sys

def find_column(columns, possible_names):
    for name in possible_names:
        for col in columns:
            if name.strip() in str(col).strip():
                return col
    return None

def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

class ExcelReportApp:
    def __init__(self, root):
        self.root = root
        self.root.title("گزارش قالبسازی - نسخه هوشمند")
        self.root.geometry("1200x800")
        self.root.configure(bg="#31DB22")

        self.df = None
        self.df_filtered = None
        self.file_path = None

        self.logo_path = resource_path("logo.png")

        self.create_menu()
        self.setup_ui()

    def create_menu(self):
        menubar = tk.Menu(self.root)
        file_menu = tk.Menu(menubar, tearoff=0)
        file_menu.add_command(label="❌ خروج", command=self.root.quit)
        menubar.add_cascade(label="فایل", menu=file_menu)
        self.root.config(menu=menubar)

    # --- متد انتخاب فایل ---
    def select_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if path:
            self.file_entry.delete(0, tk.END)
            self.file_entry.insert(0, path)
            self.file_path = path

    # --- متد بارگذاری شیت ---
    def load_sheets(self):
        path = self.file_entry.get().strip()
        if not path or not os.path.exists(path):
            messagebox.showerror("خطا", "فایل معتبر انتخاب نشده است.")
            return
        try:
            wb = load_workbook(path, read_only=True)
            self.sheet_cb["values"] = wb.sheetnames
            wb.close()
            messagebox.showinfo("انجام شد", "شیت‌ها با موفقیت بارگذاری شدند.")
        except Exception as e:
            messagebox.showerror("خطا", str(e))

    # --- متد بارگذاری داده‌ها ---
    def load_values(self):
        path = self.file_entry.get().strip()
        sel_sheet = self.sheet_cb.get().strip()
        if not path or not sel_sheet:
            messagebox.showerror("خطا", "فایل یا شیت را انتخاب کنید.")
            return
        try:
            wb = load_workbook(path, data_only=True, read_only=True)
            ws = wb[sel_sheet]
            rows = list(ws.values)
            wb.close()
            header = [str(x).strip() if x else "" for x in rows[0]]
            df = pd.DataFrame(rows[1:], columns=header)
            df.columns = [str(c).strip() for c in df.columns]
            self.df = df

            # پیدا کردن ستون‌ها
            self.repair_col = find_column(df.columns, ["نوع تعمیر", "تعمیر"])
            self.part_col = find_column(df.columns, ["قالب / قطعه / دستگاه"])
            self.date_col = find_column(df.columns, ["تاریخ"])
            self.perf_col = find_column(df.columns, ["مقدار ساعت کار شده"])
            self.req_col = find_column(df.columns, ["شماره نامه درخواست", "شماره درخواست"])
            self.code_col = find_column(df.columns, ["کد قالب"])

            self.repair_cb["values"] = ["(همه)"] + sorted(df[self.repair_col].dropna().astype(str).unique())
            self.part_cb["values"] = ["(همه)"] + sorted(df[self.part_col].dropna().astype(str).unique())

            messagebox.showinfo("انجام شد", "اطلاعات با موفقیت بارگذاری شد.")
        except Exception as e:
            messagebox.showerror("خطا", str(e))

    # --- فیلتر ساده ---
    def apply_simple_filter(self):
        if self.df is None:
            return
        df = self.df.copy()
        s = self.start_entry.get().strip()
        e = self.end_entry.get().strip()
        if s and e:
            try:
                s_g = JalaliDate.strptime(s, "%Y/%m/%d").to_gregorian()
                e_g = JalaliDate.strptime(e, "%Y/%m/%d").to_gregorian()
                df[self.date_col] = pd.to_datetime(df[self.date_col], errors="coerce")
                df = df[(df[self.date_col] >= s_g) & (df[self.date_col] <= e_g)]
            except:
                pass
        sel_repair = self.repair_cb.get()
        if sel_repair and sel_repair != "(همه)":
            df = df[df[self.repair_col].astype(str) == sel_repair]
        sel_part = self.part_cb.get()
        if sel_part and sel_part != "(همه)":
            df = df[df[self.part_col].astype(str) == sel_part]
        df[self.perf_col] = pd.to_numeric(df[self.perf_col], errors="coerce").fillna(0)
        self.df_filtered = df
        self.update_treeview(df)

    # --- فیلتر هوشمند ---
    def apply_smart_filter(self):
        if self.df_filtered is None:
            self.apply_simple_filter()
        df = self.df_filtered.copy()
        for col in df.columns:
            if col not in [self.repair_col, self.part_col, self.date_col, self.perf_col, self.req_col, self.code_col]:
                df[col] = df[col].fillna("(خالی)")
                values = df[col].dropna().unique()
                if len(values) > 0:
                    df = df[df[col].isin(values)]
        self.df_filtered = df
        self.update_treeview(df)

    # --- بروزرسانی Treeview ---
    def update_treeview(self, df):
        for i in self.tree.get_children():
            self.tree.delete(i)
        for _, row in df.iterrows():
            self.tree.insert("", "end", values=(
                row.get(self.repair_col, ""),
                row.get(self.part_col, ""),
                row.get(self.req_col, ""),
                row.get(self.code_col, ""),
                row.get(self.perf_col, 0)
            ))
        total = df[self.perf_col].sum()
        self.tree.insert("", "end", values=("جمع کل", "", "", "", total))

    # --- ذخیره خروجی ---
    def save_output(self, df):
        if df is None or df.empty:
            messagebox.showerror("خطا", "ابتدا فیلتر را اعمال کنید.")
            return
        path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                            filetypes=[("Excel", "*.xlsx"), ("CSV", "*.csv"), ("PDF", "*.pdf")])
        if not path:
            return
        df_out = df.copy()
        df_out.loc["جمع کل"] = [""] * len(df_out.columns)
        df_out.at["جمع کل", self.perf_col] = df_out[self.perf_col].sum()
        try:
            if path.endswith(".xlsx"):
                wb = Workbook()
                ws = wb.active
                ws.title = "گزارش"
                if os.path.exists(self.logo_path):
                    img = XLImage(self.logo_path)
                    img.width = 80
                    img.height = 50
                    ws.add_image(img, "A1")
                ws.append(list(df_out.columns))
                for r in df_out.itertuples(index=False):
                    ws.append(list(r))
                wb.save(path)
            elif path.endswith(".csv"):
                df_out.to_csv(path, index=False, encoding="utf-8-sig")
            elif path.endswith(".pdf"):
                c = canvas.Canvas(path, pagesize=A4)
                c.setFont("Helvetica", 10)
                y = 800
                if os.path.exists(self.logo_path):
                    c.drawImage(self.logo_path, 50, y - 50, width=80, height=50)
                for _, row in df_out.iterrows():
                    text = " | ".join([str(x) for x in row.values if x is not None])
                    c.drawString(40, y, text)
                    y -= 14
                    if y < 50:
                        c.showPage()
                        c.setFont("Helvetica", 10)
                        y = 800
                c.drawString(50, 30, "F.Alizadeh")
                c.save()
            messagebox.showinfo("ذخیره شد", f"فایل ذخیره شد:\n{path}")
        except Exception as e:
            messagebox.showerror("خطا در ذخیره", str(e))

    # --- ساخت رابط کاربری ---
    def setup_ui(self):
        bg_color = "#F77171"
        button_color = "#FF8C42"
        button_hover = "#3A1E07"
        tree_header_bg = button_color
        tree_header_fg = "white"

        # فریم بالا و لوگو
        top_frame = tk.Frame(self.root, bg=bg_color)
        top_frame.pack(fill="x", padx=10, pady=5)

        if os.path.exists(self.logo_path):
            img = Image.open(self.logo_path)
            img = img.resize((80, 50), Image.Resampling.LANCZOS)
            self.tk_img = ImageTk.PhotoImage(img)
            tk.Label(top_frame, image=self.tk_img, bg=bg_color).pack(side="left", padx=5)

        tk.Label(top_frame, text="کارشناس برنامه‌ریزی و ساخت قالبسازی: فواد مطور علیزاده",
                 font=("Arial", 11, "bold"), bg=bg_color).pack(padx=10)

        # فیلتر ساده
        frame_simple = tk.LabelFrame(self.root, text="فیلتر ساده", bg=bg_color, font=("Arial", 10, "bold"), padx=10, pady=10)
        frame_simple.pack(fill="x", padx=10, pady=5)

        tk.Label(frame_simple, text="مسیر فایل اکسل:", bg=bg_color).grid(row=0, column=0, sticky="w")
        self.file_entry = tk.Entry(frame_simple, width=70)
        self.file_entry.grid(row=0, column=1, padx=5)
        tk.Button(frame_simple, text="انتخاب فایل", command=self.select_file).grid(row=0, column=2)

        tk.Label(frame_simple, text="نام شیت:", bg=bg_color).grid(row=1, column=0, sticky="w")
        self.sheet_cb = ttk.Combobox(frame_simple, width=30, state="readonly")
        self.sheet_cb.grid(row=1, column=1, sticky="w")
        tk.Button(frame_simple, text="بارگذاری شیت‌ها", command=self.load_sheets).grid(row=1, column=2)

        tk.Label(frame_simple, text="تاریخ شروع (YYYY/MM/DD):", bg=bg_color).grid(row=2, column=0, sticky="w")
        self.start_entry = tk.Entry(frame_simple, width=15)
        self.start_entry.grid(row=2, column=1, sticky="w")

        tk.Label(frame_simple, text="تاریخ پایان (YYYY/MM/DD):", bg=bg_color).grid(row=3, column=0, sticky="w")
        self.end_entry = tk.Entry(frame_simple, width=15)
        self.end_entry.grid(row=3, column=1, sticky="w")

        tk.Label(frame_simple, text="نوع تعمیر:", bg=bg_color).grid(row=4, column=0, sticky="w")
        self.repair_cb = ttk.Combobox(frame_simple, width=30, state="readonly")
        self.repair_cb.grid(row=4, column=1, sticky="w")

        tk.Label(frame_simple, text="قالب / قطعه / دستگاه:", bg=bg_color).grid(row=5, column=0, sticky="w")
        self.part_cb = ttk.Combobox(frame_simple, width=30, state="readonly")
        self.part_cb.grid(row=5, column=1, sticky="w")

        tk.Button(frame_simple, text="📂 بارگذاری داده‌ها", command=self.load_values).grid(row=6, column=0, pady=5)
        tk.Button(frame_simple, text="🔍 اعمال فیلتر ساده", command=self.apply_simple_filter).grid(row=6, column=1, pady=5)
        tk.Button(frame_simple, text="💾 ذخیره", command=lambda: self.save_output(self.df_filtered)).grid(row=6, column=2, pady=5)

        # فیلتر هوشمند
        frame_smart = tk.LabelFrame(self.root, text="فیلتر هوشمند", bg=bg_color, font=("Arial", 10, "bold"), padx=10, pady=10)
        frame_smart.pack(fill="x", padx=10, pady=5)

        tk.Button(frame_smart, text="🔍 اعمال فیلتر هوشمند", command=self.apply_smart_filter).grid(row=0, column=0, pady=5)
        tk.Button(frame_smart, text="💾 ذخیره", command=lambda: self.save_output(self.df_filtered)).grid(row=0, column=1, pady=5)

        # Treeview
        self.tree = ttk.Treeview(self.root,
                                 columns=("نوع تعمیر", "قالب/قطعه/دستگاه", "شماره نامه درخواست", "کد قالب", "مقدار ساعت کار شده"),
                                 show="headings", height=25)
        for col in self.tree["columns"]:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=200, anchor="center")
        self.tree.pack(padx=10, pady=10, fill="both", expand=True)

# 🔹 اجرای برنامه
if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelReportApp(root)
    root.mainloop()
