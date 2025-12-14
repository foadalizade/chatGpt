# simple_app.py - نسخه بسیار ساده و مطمئن
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
import sys

try:
    import pandas as pd
    from openpyxl import load_workbook
    print("✅ همه کتابخانه‌ها وارد شدند")
except ImportError as e:
    print(f"❌ خطا در وارد کردن کتابخانه: {e}")
    messagebox.showerror("خطا", f"کتابخانه لازم نصب نیست: {e}")
    sys.exit(1)

class SimpleExcelApp:
    def __init__(self, root):
        self.root = root
        self.df = None
        
        self.setup_ui()
    
    def setup_ui(self):
        """راه‌اندازی رابط کاربری بسیار ساده"""
        self.root.title("برنامه ساده گزارش‌گیری")
        self.root.geometry("800x600")
        
        # عنوان
        title = tk.Label(self.root, text="برنامه ساده کار با اکسل", 
                        font=("Arial", 16, "bold"))
        title.pack(pady=20)
        
        # دکمه انتخاب فایل
        btn_frame = tk.Frame(self.root)
        btn_frame.pack(pady=20)
        
        tk.Button(btn_frame, text="انتخاب فایل اکسل", 
                 command=self.select_file, font=("Arial", 12),
                 bg="lightblue", padx=20, pady=10).pack(pady=10)
        
        # نمایش اطلاعات
        self.info_text = tk.Text(self.root, height=20, width=80, font=("Arial", 10))
        self.info_text.pack(pady=20, padx=20, fill="both", expand=True)
        
        # نوار وضعیت
        self.status_var = tk.StringVar()
        self.status_var.set("آماده - فایل اکسل را انتخاب کنید")
        status_bar = tk.Label(self.root, textvariable=self.status_var, 
                             relief="sunken", bd=1)
        status_bar.pack(side="bottom", fill="x")
    
    def select_file(self):
        """انتخاب فایل ساده"""
        try:
            file_path = filedialog.askopenfilename(
                title="انتخاب فایل اکسل",
                filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
            )
            
            if file_path:
                self.status_var.set(f"فایل انتخاب شد: {os.path.basename(file_path)}")
                self.process_excel_file(file_path)
                
        except Exception as e:
            error_msg = f"خطا در انتخاب فایل: {str(e)}"
            self.status_var.set(error_msg)
            messagebox.showerror("خطا", error_msg)
            print(f"❌ {error_msg}")
    
    def process_excel_file(self, file_path):
        """پردازش فایل اکسل"""
        try:
            self.status_var.set("در حال پردازش فایل...")
            self.root.update()
            
            # ابتدا فقط شیت‌ها را بگیریم
            wb = load_workbook(file_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            
            info = "=" * 50 + "\n"
            info += "گزارش پردازش فایل\n"
            info += "=" * 50 + "\n\n"
            
            info += f"📁 فایل: {os.path.basename(file_path)}\n"
            info += f"📋 تعداد شیت‌ها: {len(sheet_names)}\n"
            info += f"📜 نام شیت‌ها: {', '.join(sheet_names)}\n\n"
            
            # حالا داده‌های اولین شیت را بخوانیم
            if sheet_names:
                first_sheet = sheet_names[0]
                info += f"📊 در حال خواندن شیت: {first_sheet}\n\n"
                
                wb2 = load_workbook(file_path, read_only=True, data_only=True)
                ws = wb2[first_sheet]
                data = list(ws.values)
                wb2.close()
                
                if data:
                    headers = [str(cell) if cell is not None else "خالی" for cell in data[0]]
                    info += f"🏷️ ستون‌ها ({len(headers)} عدد):\n"
                    for i, header in enumerate(headers):
                        info += f"  {i+1}. {header}\n"
                    
                    info += f"\n📈 تعداد ردیف‌های داده: {len(data)-1}\n"
                    
                    # نمونه داده
                    if len(data) > 1:
                        info += "\n📄 نمونه داده (ردیف اول):\n"
                        first_row = data[1]
                        for i, cell in enumerate(first_row):
                            cell_value = str(cell) if cell is not None else "خالی"
                            info += f"  • {headers[i]}: {cell_value}\n"
                else:
                    info += "❌ شیت انتخاب شده خالی است\n"
            
            self.info_text.delete(1.0, tk.END)
            self.info_text.insert(1.0, info)
            self.status_var.set("پردازش کامل شد")
            
            messagebox.showinfo("موفق", "فایل با موفقیت پردازش شد")
            
        except Exception as e:
            error_msg = f"خطا در پردازش فایل: {str(e)}"
            self.status_var.set(error_msg)
            self.info_text.delete(1.0, tk.END)
            self.info_text.insert(1.0, f"خطا:\n{error_msg}")
            messagebox.showerror("خطا", error_msg)
            print(f"❌ {error_msg}")

def main():
    """تابع اصلی"""
    try:
        print("🚀 شروع برنامه ساده")
        root = tk.Tk()
        app = SimpleExcelApp(root)
        root.mainloop()
    except Exception as e:
        print(f"❌ خطای شدید: {e}")
        messagebox.showerror("خطا", f"برنامه نمی‌تواند اجرا شود: {e}")

if __name__ == "__main__":
    main()
