# main.py - برنامه گزارش‌گیری قالب‌سازی - نسخه اصلاح شده نهایی
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import os
import logging
from openpyxl import load_workbook
import traceback

# تنظیمات لاگینگ
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

class ExcelProcessor:
    """کلاس برای پردازش فایل‌های اکسل"""
    
    def __init__(self):
        self.df = None
        self.column_mapping = {}
        
    def load_excel(self, file_path, sheet_name):
        """بارگذاری فایل اکسل"""
        try:
            print(f"📂 در حال بارگذاری فایل: {file_path}")
            print(f"📋 شیت انتخاب شده: {sheet_name}")
            
            # بارگذاری فایل اکسل
            wb = load_workbook(file_path, read_only=False, data_only=True)
            ws = wb[sheet_name]
            data = list(ws.values)
            wb.close()
            
            print(f"📊 تعداد ردیف‌های خوانده شده: {len(data)}")
            
            if not data or len(data) <= 1:  # فقط هدرها یا هیچ داده‌ای نیست
                print("❌ فایل اکسل خالی است یا فقط هدر دارد")
                return False
                
            # استخراج هدرها و داده‌ها
            headers = [str(cell).strip() if cell is not None else f"Column_{i}" for i, cell in enumerate(data[0])]
            print(f"🏷️ هدرهای شناسایی شده: {headers}")
            
            # ایجاد DataFrame
            self.df = pd.DataFrame(data[1:], columns=headers)
            
            # حذف ستون‌های کاملاً خالی
            self.df = self.df.dropna(axis=1, how='all')
            
            # تشخیص خودکار ستون‌ها
            self._auto_detect_columns()
            
            print("✅ فایل اکسل با موفقیت بارگذاری شد")
            return True
            
        except Exception as e:
            print(f"❌ خطا در بارگذاری فایل اکسل: {e}")
            logging.error(f"Error loading Excel file: {e}")
            logging.error(traceback.format_exc())
            return False
    
    def _auto_detect_columns(self):
        """تشخیص خودکار ستون‌های مهم"""
        if self.df is None or self.df.empty:
            return
            
        column_patterns = {
            'repair_col': ['نوع تعمیر', 'تعمیر', 'repair', 'نوع', 'کار'],
            'part_col': ['قالب', 'قطعه', 'دستگاه', 'part', 'device', 'مورد'],
            'date_col': ['تاریخ', 'date', 'زمان'],
            'perf_col': ['مقدار ساعت کار شده', 'ساعت', 'hour', 'time', 'مدت', 'ساعت کار'],
            'req_col': ['شماره نامه درخواست', 'شماره درخواست', 'request', 'شماره'],
            'code_col': ['کد قالب', 'کد', 'code', 'شناسه']
        }
        
        print("🔍 در حال تشخیص ستون‌ها...")
        for col_type, patterns in column_patterns.items():
            found_col = None
            for col in self.df.columns:
                col_str = str(col).strip().lower()
                for pattern in patterns:
                    if pattern.lower() in col_str:
                        found_col = col
                        print(f"   ✅ {col_type} تشخیص داده شد: {col}")
                        break
                if found_col:
                    break
            self.column_mapping[col_type] = found_col
            if not found_col:
                print(f"   ❌ {col_type} تشخیص داده نشد")

class ExcelReportApp:
    """کلاس اصلی رابط کاربری"""
    
    def __init__(self, root):
        self.root = root
        self.excel_processor = ExcelProcessor()
        
        self._setup_ui()
    
    def _setup_ui(self):
        """راه‌اندازی رابط کاربری"""
        self.root.title("گزارش قالبسازی - نسخه اصلاح شده")
        self.root.geometry("1000x800")
        
        # Create main frames
        self._create_main_frame()
        
        # Status bar
        self.status_var = tk.StringVar()
        self.status_var.set("آماده - لطفاً فایل اکسل را انتخاب کنید")
        status_bar = ttk.Label(self.root, textvariable=self.status_var, relief="sunken")
        status_bar.pack(side="bottom", fill="x")
    
    def _create_main_frame(self):
        """ایجاد فریم اصلی"""
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill="both", expand=True)
        
        # Title
        title_label = ttk.Label(main_frame, text="برنامه گزارش‌گیری قالب‌سازی - نسخه اصلاح شده", 
                               font=("Arial", 16, "bold"))
        title_label.pack(pady=10)
        
        # File selection
        file_frame = ttk.LabelFrame(main_frame, text="انتخاب فایل", padding="10")
        file_frame.pack(fill="x", pady=5)
        
        # Configure grid weights for responsive layout
        file_frame.columnconfigure(1, weight=1)
        
        ttk.Label(file_frame, text="مسیر فایل اکسل:").grid(row=0, column=0, sticky="w", padx=5, pady=2)
        self.file_entry = ttk.Entry(file_frame)
        self.file_entry.grid(row=0, column=1, padx=5, pady=2, sticky="ew")
        
        ttk.Button(file_frame, text="انتخاب فایل", 
                  command=self.select_excel_file).grid(row=0, column=2, padx=5, pady=2)
        
        ttk.Label(file_frame, text="نام شیت:").grid(row=1, column=0, sticky="w", padx=5, pady=2)
        self.sheet_combo = ttk.Combobox(file_frame, state="readonly")
        self.sheet_combo.grid(row=1, column=1, sticky="w", padx=5, pady=2)
        
        ttk.Button(file_frame, text="بارگذاری شیت‌ها", 
                  command=self.load_sheets).grid(row=1, column=2, padx=5, pady=2)
        
        # Buttons frame
        button_frame = ttk.Frame(file_frame)
        button_frame.grid(row=2, column=0, columnspan=3, pady=10)
        
        ttk.Button(button_frame, text="بارگذاری داده‌ها", 
                  command=self.load_data).pack(side="left", padx=5)
        
        ttk.Button(button_frame, text="نمایش اطلاعات کامل", 
                  command=self.show_info).pack(side="left", padx=5)
        
        ttk.Button(button_frame, text="نمایش نمونه داده‌ها", 
                  command=self.show_sample_data).pack(side="left", padx=5)
        
        # Info display
        info_frame = ttk.LabelFrame(main_frame, text="اطلاعات و نتایج", padding="10")
        info_frame.pack(fill="both", expand=True, pady=5)
        
        # Create text widget with scrollbar
        text_frame = ttk.Frame(info_frame)
        text_frame.pack(fill="both", expand=True)
        
        self.info_text = tk.Text(text_frame, height=20, width=80, font=("Arial", 10))
        scrollbar_v = ttk.Scrollbar(text_frame, orient="vertical", command=self.info_text.yview)
        scrollbar_h = ttk.Scrollbar(text_frame, orient="horizontal", command=self.info_text.xview)
        
        self.info_text.configure(yscrollcommand=scrollbar_v.set, xscrollcommand=scrollbar_h.set)
        
        self.info_text.grid(row=0, column=0, sticky="nsew")
        scrollbar_v.grid(row=0, column=1, sticky="ns")
        scrollbar_h.grid(row=1, column=0, sticky="ew")
        
        text_frame.columnconfigure(0, weight=1)
        text_frame.rowconfigure(0, weight=1)
    
    def select_excel_file(self):
        """انتخاب فایل اکسل"""
        try:
            file_path = filedialog.askopenfilename(
                title="انتخاب فایل اکسل",
                filetypes=[("Excel Files", "*.xlsx *.xls"), ("All files", "*.*")]
            )
            if file_path:
                self.file_entry.delete(0, tk.END)
                self.file_entry.insert(0, file_path)
                self.status_var.set(f"فایل انتخاب شد: {os.path.basename(file_path)}")
                print(f"📁 فایل انتخاب شده: {file_path}")
        except Exception as e:
            messagebox.showerror("خطا", f"خطا در انتخاب فایل: {e}")
    
    def load_sheets(self):
        """بارگذاری لیست شیت‌ها"""
        file_path = self.file_entry.get().strip()
        if not file_path:
            messagebox.showerror("خطا", "لطفاً ابتدا یک فایل انتخاب کنید")
            return
        
        if not os.path.exists(file_path):
            messagebox.showerror("خطا", "فایل انتخاب شده وجود ندارد")
            return
        
        try:
            print(f"🔍 در حال بارگذاری شیت‌های فایل: {file_path}")
            
            # بارگذاری فایل اکسل برای خواندن نام شیت‌ها
            wb = load_workbook(file_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            
            print(f"📋 شیت‌های پیدا شده: {sheet_names}")
            
            if not sheet_names:
                messagebox.showwarning("هشدار", "هیچ شیتی در فایل اکسل پیدا نشد")
                return
            
            self.sheet_combo['values'] = sheet_names
            self.sheet_combo.set(sheet_names[0])
            
            self.status_var.set(f"{len(sheet_names)} شیت پیدا شد - اولین شیت انتخاب شد")
            messagebox.showinfo("موفق", f"{len(sheet_names)} شیت در فایل پیدا شد")
            
        except Exception as e:
            error_msg = f"خطا در بارگذاری شیت‌ها: {str(e)}"
            print(f"❌ {error_msg}")
            logging.error(traceback.format_exc())
            messagebox.showerror("خطا", error_msg)
    
    def load_data(self):
        """بارگذاری داده‌ها"""
        file_path = self.file_entry.get().strip()
        sheet_name = self.sheet_combo.get().strip()
        
        if not file_path or not sheet_name:
            messagebox.showwarning("هشدار", "لطفاً فایل و شیت را انتخاب کنید")
            return
        
        try:
            print(f"🔄 شروع بارگذاری داده‌ها از شیت: {sheet_name}")
            self.status_var.set("در حال بارگذاری داده‌ها...")
            self.root.update()
            
            success = self.excel_processor.load_excel(file_path, sheet_name)
            if success:
                # روش صحیح بررسی وجود DataFrame
                if self.excel_processor.df is not None and not self.excel_processor.df.empty:
                    record_count = len(self.excel_processor.df)
                    success_msg = f"داده‌ها با موفقیت بارگذاری شدند ({record_count} رکورد)"
                    self.status_var.set(success_msg)
                    print(f"✅ {success_msg}")
                    messagebox.showinfo("موفق", success_msg)
                else:
                    error_msg = "داده‌ها بارگذاری شدند اما فایل خالی است"
                    self.status_var.set(error_msg)
                    print(f"⚠️ {error_msg}")
                    messagebox.showwarning("هشدار", error_msg)
            else:
                error_msg = "خطا در بارگذاری داده‌ها - لطفاً فایل را بررسی کنید"
                self.status_var.set(error_msg)
                print(f"❌ {error_msg}")
                messagebox.showerror("خطا", error_msg)
                
        except Exception as e:
            error_msg = f"خطا در بارگذاری داده‌ها: {str(e)}"
            print(f"❌ {error_msg}")
            logging.error(traceback.format_exc())
            self.status_var.set("خطا در بارگذاری داده‌ها")
            messagebox.showerror("خطا", error_msg)
    
    def show_info(self):
        """نمایش اطلاعات کامل داده‌ها"""
        # روش صحیح بررسی وجود DataFrame
        if self.excel_processor.df is None or self.excel_processor.df.empty:
            messagebox.showwarning("هشدار", "لطفاً ابتدا داده‌ها را بارگذاری کنید")
            return
        
        try:
            info = "="*60 + "\n"
            info += "گزارش کامل داده‌های بارگذاری شده\n"
            info += "="*60 + "\n\n"
            
            # اطلاعات کلی
            info += "📊 اطلاعات کلی:\n"
            info += f"• تعداد رکوردها: {len(self.excel_processor.df):,}\n"
            info += f"• تعداد ستون‌ها: {len(self.excel_processor.df.columns)}\n"
            
            # ستون‌های تشخیص داده شده
            info += "\n🔍 ستون‌های تشخیص داده شده:\n"
            detected_count = 0
            for col_type, col_name in self.excel_processor.column_mapping.items():
                status = "✅" if col_name else "❌"
                info += f"  {status} {col_type}: {col_name if col_name else 'یافت نشد'}\n"
                if col_name:
                    detected_count += 1
            
            info += f"\n• مجموع ستون‌های تشخیص داده شده: {detected_count} از {len(self.excel_processor.column_mapping)}\n\n"
            
            # اطلاعات ستون‌ها
            info += "📋 اطلاعات ستون‌ها:\n"
            for i, col in enumerate(self.excel_processor.df.columns):
                non_null_count = self.excel_processor.df[col].count()
                null_count = len(self.excel_processor.df) - non_null_count
                # روش صحیح دریافت نوع داده ستون
                dtype = str(self.excel_processor.df[col].dtype)
                info += f"  {i+1:2d}. {col}: {dtype} - {non_null_count} مقدار ({null_count} خالی)\n"
            
            self.info_text.delete(1.0, tk.END)
            self.info_text.insert(1.0, info)
            self.status_var.set("اطلاعات کامل نمایش داده شد")
            print("📄 اطلاعات کامل نمایش داده شد")
            
        except Exception as e:
            error_msg = f"خطا در نمایش اطلاعات: {str(e)}"
            print(f"❌ {error_msg}")
            messagebox.showerror("خطا", error_msg)
    
    def show_sample_data(self):
        """نمایش نمونه داده‌ها"""
        # روش صحیح بررسی وجود DataFrame
        if self.excel_processor.df is None or self.excel_processor.df.empty:
            messagebox.showwarning("هشدار", "لطفاً ابتدا داده‌ها را بارگذاری کنید")
            return
        
        try:
            info = "="*60 + "\n"
            info += "نمونه داده‌های بارگذاری شده (5 رکورد اول)\n"
            info += "="*60 + "\n\n"
            
            # نمایش 5 رکورد اول
            sample_df = self.excel_processor.df.head(5)
            
            # فرمت‌بندی زیبا برای نمایش
            info += "📄 داده‌های نمونه:\n\n"
            for idx, (_, row) in enumerate(sample_df.iterrows()):
                info += f"رکورد {idx + 1}:\n"
                for col in sample_df.columns:
                    value = row[col]
                    # روش صحیح بررسی NaN
                    if pd.isna(value):
                        value = "---"
                    elif isinstance(value, (int, float)):
                        value = str(value)
                    else:
                        value = str(value)
                    
                    # کوتاه کردن متن‌های طولانی
                    if len(value) > 50:
                        value = value[:47] + "..."
                    
                    info += f"  • {col}: {value}\n"
                info += "-" * 40 + "\n"
            
            self.info_text.delete(1.0, tk.END)
            self.info_text.insert(1.0, info)
            self.status_var.set("نمونه داده‌ها نمایش داده شد")
            print("📊 نمونه داده‌ها نمایش داده شد")
            
        except Exception as e:
            error_msg = f"خطا در نمایش نمونه داده‌ها: {str(e)}"
            print(f"❌ {error_msg}")
            messagebox.showerror("خطا", error_msg)

def main():
    """تابع اصلی برنامه"""
    try:
        print("🚀 شروع برنامه گزارش‌گیری قالب‌سازی")
        root = tk.Tk()
        app = ExcelReportApp(root)
        root.mainloop()
    except Exception as e:
        error_msg = f"برنامه نمی‌تواند اجرا شود: {e}"
        print(f"❌ {error_msg}")
        logging.error(traceback.format_exc())
        messagebox.showerror("خطای شدید", error_msg)

if __name__ == "__main__":
    main()