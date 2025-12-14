# main_advanced.py - برنامه کامل گزارش‌گیری با فیلترها
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import os
import logging
from openpyxl import load_workbook
import traceback

# تنظیمات لاگینگ
logging.basicConfig(level=logging.INFO)

class ExcelProcessor:
    """کلاس برای پردازش فایل‌های اکسل"""
    
    def __init__(self):
        self.df = None
        self.df_normalized = None
        self.column_mapping = {}
        
    def load_excel(self, file_path, sheet_name):
        """بارگذاری فایل اکسل"""
        try:
            print(f"📂 در حال بارگذاری فایل: {file_path}")
            print(f"📋 شیت انتخاب شده: {sheet_name}")
            
            wb = load_workbook(file_path, read_only=False, data_only=True)
            ws = wb[sheet_name]
            data = list(ws.values)
            wb.close()
            
            print(f"📊 تعداد ردیف‌های خوانده شده: {len(data)}")
            
            if not data or len(data) <= 1:
                return False
                
            headers = [str(cell).strip() if cell is not None else f"Column_{i}" for i, cell in enumerate(data[0])]
            self.df = pd.DataFrame(data[1:], columns=headers)
            self.df = self.df.dropna(axis=1, how='all')
            
            self._auto_detect_columns()
            self._create_normalized_data()
            
            return True
            
        except Exception as e:
            print(f"❌ خطا در بارگذاری فایل اکسل: {e}")
            return False
    
    def _auto_detect_columns(self):
        """تشخیص خودکار ستون‌های مهم"""
        if self.df is None or self.df.empty:
            return
            
        column_patterns = {
            'repair_col': ['نوع تعمیر', 'تعمیر', 'repair'],
            'part_col': ['قالب', 'قطعه', 'دستگاه', 'part', 'device'],
            'date_col': ['تاریخ', 'date'],
            'perf_col': ['مقدار ساعت کار شده', 'ساعت', 'hour', 'time'],
            'req_col': ['شماره نامه درخواست', 'شماره درخواست', 'request'],
            'code_col': ['کد قالب', 'کد', 'code']
        }
        
        for col_type, patterns in column_patterns.items():
            found_col = None
            for col in self.df.columns:
                col_str = str(col).strip().lower()
                for pattern in patterns:
                    if pattern.lower() in col_str:
                        found_col = col
                        break
                if found_col:
                    break
            self.column_mapping[col_type] = found_col
    
    def _create_normalized_data(self):
        """ایجاد نسخه نرمالایز شده"""
        if self.df is None or self.df.empty:
            return
            
        self.df_normalized = self.df.copy()
        repair_col = self.column_mapping.get('repair_col')
        if repair_col and repair_col in self.df_normalized.columns:
            self.df_normalized[repair_col] = self.df_normalized[repair_col].apply(self._normalize_repair_type)
    
    def _normalize_repair_type(self, repair_type):
        """نرمال‌سازی نوع تعمیر"""
        if not isinstance(repair_type, str):
            return str(repair_type)
        
        repair_type = str(repair_type).strip().lower()
        
        if 'قالب' in repair_type and 'تعمیر' in repair_type:
            return 'قالب تعمیری'
        elif 'قطعه' in repair_type and 'تعمیر' in repair_type:
            return 'قطعه تعمیری'
        elif 'دستگاه' in repair_type and 'تعمیر' in repair_type:
            return 'دستگاه تعمیری'
        elif 'قالب' in repair_type:
            return 'قالب'
        elif 'قطعه' in repair_type:
            return 'قطعه'
        elif 'دستگاه' in repair_type:
            return 'دستگاه'
        elif 'تعمیر' in repair_type:
            return 'تعمیری'
        else:
            return repair_type

class DataFilter:
    """کلاس برای فیلتر کردن داده‌ها"""
    
    def __init__(self, excel_processor):
        self.excel_processor = excel_processor
        self.filtered_data = None
        
    def apply_simple_filter(self, start_date="", end_date="", repair_type="", part_type=""):
        """اعمال فیلتر ساده"""
        if self.excel_processor.df is None or self.excel_processor.df.empty:
            return None
            
        df = self.excel_processor.df.copy()
        
        # فیلتر نوع تعمیر
        if repair_type and repair_type != "(همه)":
            repair_col = self.excel_processor.column_mapping.get('repair_col')
            if repair_col and repair_col in df.columns:
                # استفاده از داده‌های نرمالایز شده برای فیلتر
                normalized_mask = self.excel_processor.df_normalized[repair_col] == repair_type
                df = df[normalized_mask]
        
        # فیلتر قالب/قطعه/دستگاه
        if part_type and part_type != "(همه)":
            part_col = self.excel_processor.column_mapping.get('part_col')
            if part_col and part_col in df.columns:
                df = df[df[part_col].astype(str) == part_type]
        
        self.filtered_data = df
        return df

class AdvancedReportApp:
    """برنامه پیشرفته با قابلیت فیلتر"""
    
    def __init__(self, root):
        self.root = root
        self.excel_processor = ExcelProcessor()
        self.data_filter = DataFilter(self.excel_processor)
        
        self._setup_ui()
    
    def _setup_ui(self):
        """راه‌اندازی رابط کاربری"""
        self.root.title("گزارش قالبسازی - نسخه پیشرفته")
        self.root.geometry("1200x800")
        
        self._create_main_frame()
        
        # Status bar
        self.status_var = tk.StringVar()
        self.status_var.set("آماده - لطفاً فایل اکسل را انتخاب کنید")
        status_bar = ttk.Label(self.root, textvariable=self.status_var, relief="sunken")
        status_bar.pack(side="bottom", fill="x")
    
    def _create_main_frame(self):
        """ایجاد فریم اصلی"""
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill="both", expand=True)
        
        # Title
        title_label = ttk.Label(main_frame, text="برنامه پیشرفته گزارش‌گیری قالب‌سازی", 
                               font=("Arial", 16, "bold"))
        title_label.pack(pady=10)
        
        # File selection
        self._create_file_section(main_frame)
        
        # Filters
        self._create_filter_section(main_frame)
        
        # Results
        self._create_results_section(main_frame)
    
    def _create_file_section(self, parent):
        """بخش انتخاب فایل"""
        file_frame = ttk.LabelFrame(parent, text="انتخاب فایل و داده‌ها", padding="10")
        file_frame.pack(fill="x", pady=5)
        
        file_frame.columnconfigure(1, weight=1)
        
        # File path
        ttk.Label(file_frame, text="مسیر فایل اکسل:").grid(row=0, column=0, sticky="w", padx=5, pady=2)
        self.file_entry = ttk.Entry(file_frame)
        self.file_entry.grid(row=0, column=1, padx=5, pady=2, sticky="ew")
        ttk.Button(file_frame, text="انتخاب فایل", command=self.select_excel_file).grid(row=0, column=2, padx=5, pady=2)
        
        # Sheet selection
        ttk.Label(file_frame, text="نام شیت:").grid(row=1, column=0, sticky="w", padx=5, pady=2)
        self.sheet_combo = ttk.Combobox(file_frame, state="readonly")
        self.sheet_combo.grid(row=1, column=1, sticky="w", padx=5, pady=2)
        ttk.Button(file_frame, text="بارگذاری شیت‌ها", command=self.load_sheets).grid(row=1, column=2, padx=5, pady=2)
        
        # Load data button
        ttk.Button(file_frame, text="بارگذاری داده‌ها", command=self.load_data).grid(row=2, column=1, pady=10)
    
    def _create_filter_section(self, parent):
        """بخش فیلترها"""
        filter_frame = ttk.LabelFrame(parent, text="فیلترهای داده", padding="10")
        filter_frame.pack(fill="x", pady=5)
        
        filter_frame.columnconfigure(1, weight=1)
        
        # Repair type filter
        ttk.Label(filter_frame, text="نوع تعمیر:").grid(row=0, column=0, sticky="w", padx=5, pady=2)
        self.repair_combo = ttk.Combobox(filter_frame, state="readonly")
        self.repair_combo.grid(row=0, column=1, sticky="w", padx=5, pady=2)
        self.repair_combo.bind('<<ComboboxSelected>>', self.on_repair_changed)
        
        # Part filter
        ttk.Label(filter_frame, text="قالب/قطعه/دستگاه:").grid(row=1, column=0, sticky="w", padx=5, pady=2)
        self.part_combo = ttk.Combobox(filter_frame, state="readonly")
        self.part_combo.grid(row=1, column=1, sticky="w", padx=5, pady=2)
        
        # Filter buttons
        button_frame = ttk.Frame(filter_frame)
        button_frame.grid(row=2, column=0, columnspan=3, pady=10)
        
        ttk.Button(button_frame, text="اعمال فیلتر", command=self.apply_filter).pack(side="left", padx=5)
        ttk.Button(button_frame, text="نمایش همه داده‌ها", command=self.show_all_data).pack(side="left", padx=5)
        ttk.Button(button_frame, text="گروه‌بندی داده‌ها", command=self.group_data).pack(side="left", padx=5)
    
    def _create_results_section(self, parent):
        """بخش نتایج"""
        results_frame = ttk.LabelFrame(parent, text="نتایج و گزارش‌ها", padding="10")
        results_frame.pack(fill="both", expand=True, pady=5)
        
        # Create notebook for tabs
        self.notebook = ttk.Notebook(results_frame)
        self.notebook.pack(fill="both", expand=True)
        
        # Info tab
        info_tab = ttk.Frame(self.notebook)
        self.notebook.add(info_tab, text="اطلاعات کلی")
        
        self.info_text = tk.Text(info_tab, height=15, font=("Arial", 10))
        scrollbar_v = ttk.Scrollbar(info_tab, orient="vertical", command=self.info_text.yview)
        scrollbar_h = ttk.Scrollbar(info_tab, orient="horizontal", command=self.info_text.xview)
        self.info_text.configure(yscrollcommand=scrollbar_v.set, xscrollcommand=scrollbar_h.set)
        
        self.info_text.grid(row=0, column=0, sticky="nsew")
        scrollbar_v.grid(row=0, column=1, sticky="ns")
        scrollbar_h.grid(row=1, column=0, sticky="ew")
        
        info_tab.columnconfigure(0, weight=1)
        info_tab.rowconfigure(0, weight=1)
        
        # Data tab
        data_tab = ttk.Frame(self.notebook)
        self.notebook.add(data_tab, text="نمایش داده‌ها")
        
        # Create treeview for data display
        columns = ("نوع تعمیر", "قالب/قطعه/دستگاه", "شماره درخواست", "کد قالب", "ساعت کاری")
        self.tree = ttk.Treeview(data_tab, columns=columns, show="headings", height=15)
        
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=150)
        
        tree_scrollbar = ttk.Scrollbar(data_tab, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=tree_scrollbar.set)
        
        self.tree.grid(row=0, column=0, sticky="nsew")
        tree_scrollbar.grid(row=0, column=1, sticky="ns")
        
        data_tab.columnconfigure(0, weight=1)
        data_tab.rowconfigure(0, weight=1)
    
    def select_excel_file(self):
        """انتخاب فایل اکسل"""
        file_path = filedialog.askopenfilename(
            title="انتخاب فایل اکسل",
            filetypes=[("Excel Files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if file_path:
            self.file_entry.delete(0, tk.END)
            self.file_entry.insert(0, file_path)
            self.status_var.set(f"فایل انتخاب شد: {os.path.basename(file_path)}")
    
    def load_sheets(self):
        """بارگذاری لیست شیت‌ها"""
        file_path = self.file_entry.get().strip()
        if not file_path or not os.path.exists(file_path):
            messagebox.showerror("خطا", "لطفاً ابتدا یک فایل معتبر انتخاب کنید")
            return
        
        try:
            with load_workbook(file_path, read_only=True) as wb:
                sheet_names = wb.sheetnames
            
            self.sheet_combo['values'] = sheet_names
            if sheet_names:
                self.sheet_combo.set(sheet_names[0])
            
            self.status_var.set(f"{len(sheet_names)} شیت پیدا شد")
            
        except Exception as e:
            messagebox.showerror("خطا", f"خطا در بارگذاری شیت‌ها: {e}")
    
    def load_data(self):
        """بارگذاری داده‌ها"""
        file_path = self.file_entry.get().strip()
        sheet_name = self.sheet_combo.get().strip()
        
        if not file_path or not sheet_name:
            messagebox.showwarning("هشدار", "لطفاً فایل و شیت را انتخاب کنید")
            return
        
        try:
            success = self.excel_processor.load_excel(file_path, sheet_name)
            if success and self.excel_processor.df is not None and not self.excel_processor.df.empty:
                record_count = len(self.excel_processor.df)
                self.status_var.set(f"{record_count} رکورد با موفقیت بارگذاری شد")
                
                # پر کردن فیلترها
                self._populate_filters()
                self.show_all_data()
                
                messagebox.showinfo("موفق", f"داده‌ها با موفقیت بارگذاری شدند ({record_count} رکورد)")
            else:
                messagebox.showerror("خطا", "خطا در بارگذاری داده‌ها")
                
        except Exception as e:
            messagebox.showerror("خطا", f"خطا در بارگذاری داده‌ها: {e}")
    
    def _populate_filters(self):
        """پر کردن فیلترها با داده‌ها"""
        if self.excel_processor.df_normalized is None:
            return
        
        # پر کردن فیلتر نوع تعمیر
        repair_col = self.excel_processor.column_mapping.get('repair_col')
        if repair_col and repair_col in self.excel_processor.df_normalized.columns:
            repair_types = ["(همه)"] + sorted(self.excel_processor.df_normalized[repair_col].dropna().unique().tolist())
            self.repair_combo['values'] = repair_types
            self.repair_combo.set("(همه)")
        
        # پر کردن فیلتر قالب/قطعه/دستگاه
        part_col = self.excel_processor.column_mapping.get('part_col')
        if part_col and part_col in self.excel_processor.df.columns:
            part_types = ["(همه)"] + sorted(self.excel_processor.df[part_col].dropna().astype(str).unique().tolist())
            self.part_combo['values'] = part_types
            self.part_combo.set("(همه)")
    
    def on_repair_changed(self, event=None):
        """وقتی نوع تعمیر تغییر کرد"""
        selected_repair = self.repair_combo.get()
        if not selected_repair or selected_repair == "(همه)" or self.excel_processor.df_normalized is None:
            return
        
        # اینجا می‌توانید منطق فیلتر هوشمند را اضافه کنید
    
    def apply_filter(self):
        """اعمال فیلترها"""
        if self.excel_processor.df is None:
            messagebox.showwarning("هشدار", "لطفاً ابتدا داده‌ها را بارگذاری کنید")
            return
        
        repair_type = self.repair_combo.get()
        part_type = self.part_combo.get()
        
        filtered_df = self.data_filter.apply_simple_filter(
            repair_type=repair_type,
            part_type=part_type
        )
        
        if filtered_df is not None and not filtered_df.empty:
            self._display_data_in_treeview(filtered_df)
            self.status_var.set(f"فیلتر اعمال شد - {len(filtered_df)} رکورد")
        else:
            messagebox.showwarning("هشدار", "هیچ داده‌ای با فیلترهای انتخاب شده یافت نشد")
    
    def show_all_data(self):
        """نمایش همه داده‌ها"""
        if self.excel_processor.df is not None and not self.excel_processor.df.empty:
            self._display_data_in_treeview(self.excel_processor.df)
            self._show_basic_info()
            self.status_var.set(f"همه داده‌ها نمایش داده شد - {len(self.excel_processor.df)} رکورد")
    
    def group_data(self):
        """گروه‌بندی داده‌ها"""
        # اینجا می‌توانید منطق گروه‌بندی را اضافه کنید
        messagebox.showinfo("گروه‌بندی", "قابلیت گروه‌بندی به زودی اضافه خواهد شد")
    
    def _display_data_in_treeview(self, df):
        """نمایش داده‌ها در Treeview"""
        # پاک کردن داده‌های قبلی
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # نمایش داده‌ها
        repair_col = self.excel_processor.column_mapping.get('repair_col')
        part_col = self.excel_processor.column_mapping.get('part_col')
        req_col = self.excel_processor.column_mapping.get('req_col')
        code_col = self.excel_processor.column_mapping.get('code_col')
        perf_col = self.excel_processor.column_mapping.get('perf_col')
        
        for _, row in df.head(100).iterrows():  # فقط 100 رکورد اول نمایش داده می‌شود
            repair_value = row[repair_col] if repair_col and repair_col in row else ""
            part_value = row[part_col] if part_col and part_col in row else ""
            req_value = row[req_col] if req_col and req_col in row else ""
            code_value = row[code_col] if code_col and code_col in row else ""
            perf_value = row[perf_col] if perf_col and perf_col in row else ""
            
            self.tree.insert("", "end", values=(
                repair_value,
                part_value,
                req_value,
                code_value,
                perf_value
            ))
    
    def _show_basic_info(self):
        """نمایش اطلاعات پایه"""
        if self.excel_processor.df is None:
            return
        
        info = "="*50 + "\n"
        info += "اطلاعات کلی داده‌ها\n"
        info += "="*50 + "\n\n"
        
        info += f"تعداد رکوردها: {len(self.excel_processor.df):,}\n"
        info += f"تعداد ستون‌ها: {len(self.excel_processor.df.columns)}\n\n"
        
        info += "ستون‌های تشخیص داده شده:\n"
        for col_type, col_name in self.excel_processor.column_mapping.items():
            status = "✅" if col_name else "❌"
            info += f"  {status} {col_type}: {col_name if col_name else 'یافت نشد'}\n"
        
        self.info_text.delete(1.0, tk.END)
        self.info_text.insert(1.0, info)

def main():
    """تابع اصلی برنامه"""
    try:
        root = tk.Tk()
        app = AdvancedReportApp(root)
        root.mainloop()
    except Exception as e:
        messagebox.showerror("خطای شدید", f"برنامه نمی‌تواند اجرا شود: {e}")

if __name__ == "__main__":
    main()
