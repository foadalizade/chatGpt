# complete_app.py - برنامه کامل با قابلیت‌های پیشرفته
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import os
import logging
from openpyxl import load_workbook
from persiantools.jdatetime import JalaliDate
import traceback
from datetime import datetime

# تنظیمات لاگینگ
logging.basicConfig(level=logging.INFO)

class ExcelProcessor:
    def __init__(self):
        self.df = None
        self.df_normalized = None
        self.column_mapping = {}
        
    def load_excel(self, file_path, sheet_name):
        """بارگذاری فایل اکسل"""
        try:
            print(f"📂 در حال بارگذاری فایل: {file_path}")
            print(f"📋 شیت انتخاب شده: {sheet_name}")
            
            wb = load_workbook(file_path, read_only=False, data_only=True)
            ws = wb[sheet_name]
            data = list(ws.values)
            wb.close()
            
            print(f"📊 تعداد ردیف‌های خوانده شده: {len(data)}")
            
            if not data or len(data) <= 1:
                return False
                
            headers = [str(cell).strip() if cell is not None else f"Column_{i}" for i, cell in enumerate(data[0])]
            self.df = pd.DataFrame(data[1:], columns=headers)
            
            # حذف ستون‌های کاملاً خالی
            self.df = self.df.dropna(axis=1, how='all')
            
            self._auto_detect_columns()
            self._create_normalized_data()
            
            return True
            
        except Exception as e:
            print(f"❌ خطا در بارگذاری فایل اکسل: {e}")
            return False
    
    def _auto_detect_columns(self):
        """تشخیص خودکار ستون‌های مهم"""
        if self.df is None or self.df.empty:
            return
            
        column_patterns = {
            'repair_col': ['نوع تعمیر', 'تعمیر', 'repair'],
            'part_col': ['قالب', 'قطعه', 'دستگاه', 'part', 'device'],
            'date_col': ['تاریخ', 'date'],
            'perf_col': ['مقدار ساعت کار شده', 'ساعت', 'hour', 'time'],
            'req_col': ['شماره نامه درخواست', 'شماره درخواست', 'request'],
            'code_col': ['کد قالب', 'کد', 'code']
        }
        
        for col_type, patterns in column_patterns.items():
            found_col = None
            for col in self.df.columns:
                col_str = str(col).strip().lower()
                for pattern in patterns:
                    if pattern.lower() in col_str:
                        found_col = col
                        break
                if found_col:
                    break
            self.column_mapping[col_type] = found_col
    
    def _create_normalized_data(self):
        """ایجاد نسخه نرمالایز شده"""
        if self.df is None or self.df.empty:
            return
            
        self.df_normalized = self.df.copy()
        repair_col = self.column_mapping.get('repair_col')
        if repair_col and repair_col in self.df_normalized.columns:
            self.df_normalized[repair_col] = self.df_normalized[repair_col].apply(self._normalize_repair_type)
    
    def _normalize_repair_type(self, repair_type):
        """نرمال‌سازی نوع تعمیر"""
        if not isinstance(repair_type, str):
            return str(repair_type)
        
        repair_type = str(repair_type).strip().lower()
        
        if 'قالب' in repair_type and 'تعمیر' in repair_type:
            return 'قالب تعمیری'
        elif 'قطعه' in repair_type and 'تعمیر' in repair_type:
            return 'قطعه تعمیری'
        elif 'دستگاه' in repair_type and 'تعمیر' in repair_type:
            return 'دستگاه تعمیری'
        elif 'قالب' in repair_type:
            return 'قالب'
        elif 'قطعه' in repair_type:
            return 'قطعه'
        elif 'دستگاه' in repair_type:
            return 'دستگاه'
        elif 'تعمیر' in repair_type:
            return 'تعمیری'
        else:
            return repair_type

class DataFilter:
    """کلاس برای فیلتر کردن داده‌ها"""
    
    def __init__(self, excel_processor):
        self.excel_processor = excel_processor
        self.filtered_data = None
        
    def apply_simple_filter(self, repair_type="", part_type=""):
        """اعمال فیلتر ساده"""
        if self.excel_processor.df is None or self.excel_processor.df.empty:
            return None
            
        df = self.excel_processor.df.copy()
        
        # فیلتر نوع تعمیر
        if repair_type and repair_type != "(همه)":
            repair_col = self.excel_processor.column_mapping.get('repair_col')
            if repair_col and repair_col in df.columns:
                # استفاده از داده‌های نرمالایز شده برای فیلتر
                if self.excel_processor.df_normalized is not None:
                    normalized_mask = self.excel_processor.df_normalized[repair_col] == repair_type
                    df = df[normalized_mask]
        
        # فیلتر قالب/قطعه/دستگاه
        if part_type and part_type != "(همه)":
            part_col = self.excel_processor.column_mapping.get('part_col')
            if part_col and part_col in df.columns:
                df = df[df[part_col].astype(str) == part_type]
        
        self.filtered_data = df
        return df
    
    def apply_advanced_filter(self, selected_repairs, hour_min=None, hour_max=None):
        """اعمال فیلتر پیشرفته"""
        if self.excel_processor.df is None or self.excel_processor.df.empty:
            return None
            
        df = self.excel_processor.df.copy()
        
        # فیلتر انواع تعمیر
        if selected_repairs:
            repair_col = self.excel_processor.column_mapping.get('repair_col')
            if repair_col and repair_col in df.columns:
                mask = df[repair_col].astype(str).isin(selected_repairs)
                df = df[mask]
        
        # فیلتر بازه ساعتی
        if hour_min is not None or hour_max is not None:
            perf_col = self.excel_processor.column_mapping.get('perf_col')
            if perf_col and perf_col in df.columns:
                df[perf_col] = pd.to_numeric(df[perf_col], errors='coerce')
                
                if hour_min is not None:
                    df = df[df[perf_col] >= hour_min]
                if hour_max is not None:
                    df = df[df[perf_col] <= hour_max]
        
        self.filtered_data = df
        return df
    
    def group_data(self):
        """گروه‌بندی داده‌ها"""
        if self.filtered_data is None or self.filtered_data.empty:
            return self.excel_processor.df
            
        grouping_cols = []
        for col_type in ['part_col', 'code_col', 'req_col']:
            col_name = self.excel_processor.column_mapping.get(col_type)
            if col_name and col_name in self.filtered_data.columns:
                grouping_cols.append(col_name)
        
        if not grouping_cols:
            return self.filtered_data
        
        perf_col = self.excel_processor.column_mapping.get('perf_col')
        if not perf_col:
            return self.filtered_data
        
        # گروه‌بندی و جمع‌بندی
        grouped_df = self.filtered_data.groupby(grouping_cols, as_index=False).agg({
            perf_col: 'sum'
        }).sort_values(by=perf_col, ascending=False)
        
        return grouped_df

class ReportGenerator:
    """کلاس برای تولید گزارش"""
    
    @staticmethod
    def save_to_excel(df, file_path):
        """ذخیره در فایل اکسل"""
        try:
            df.to_excel(file_path, index=False, engine='openpyxl')
            return True
        except Exception as e:
            print(f"❌ خطا در ذخیره فایل اکسل: {e}")
            return False
    
    @staticmethod
    def save_to_csv(df, file_path):
        """ذخیره در فایل CSV"""
        try:
            df.to_csv(file_path, index=False, encoding='utf-8-sig')
            return True
        except Exception as e:
            print(f"❌ خطا در ذخیره فایل CSV: {e}")
            return False

class CompleteReportApp:
    """برنامه کامل با تمام قابلیت‌ها"""
    
    def __init__(self, root):
        self.root = root
        self.excel_processor = ExcelProcessor()
        self.data_filter = DataFilter(self.excel_processor)
        
        self._setup_ui()
    
    def _setup_ui(self):
        """راه‌اندازی رابط کاربری"""
        self.root.title("برنامه کامل گزارش‌گیری قالب‌سازی")
        self.root.geometry("1200x900")
        
        self._create_main_frame()
        
        # Status bar
        self.status_var = tk.StringVar()
        self.status_var.set("آماده - لطفاً فایل اکسل را انتخاب کنید")
        status_bar = ttk.Label(self.root, textvariable=self.status_var, relief="sunken")
        status_bar.pack(side="bottom", fill="x")
    
    def _create_main_frame(self):
        """ایجاد فریم اصلی"""
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill="both", expand=True)
        
        # Title
        title_label = ttk.Label(main_frame, text="برنامه کامل گزارش‌گیری قالب‌سازی", 
                               font=("Arial", 16, "bold"))
        title_label.pack(pady=10)
        
        # Create notebook for tabs
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.pack(fill="both", expand=True, pady=10)
        
        # Tab 1: Data Loading
        self._create_data_tab()
        
        # Tab 2: Simple Filters
        self._create_simple_filter_tab()
        
        # Tab 3: Advanced Filters
        self._create_advanced_filter_tab()
        
        # Tab 4: Results
        self._create_results_tab()
    
    def _create_data_tab(self):
        """تب بارگذاری داده‌ها"""
        data_tab = ttk.Frame(self.notebook)
        self.notebook.add(data_tab, text="بارگذاری داده‌ها")
        
        # File selection
        file_frame = ttk.LabelFrame(data_tab, text="انتخاب فایل", padding="10")
        file_frame.pack(fill="x", pady=5, padx=10)
        
        file_frame.columnconfigure(1, weight=1)
        
        ttk.Label(file_frame, text="مسیر فایل اکسل:").grid(row=0, column=0, sticky="w", padx=5, pady=2)
        self.file_entry = ttk.Entry(file_frame)
        self.file_entry.grid(row=0, column=1, padx=5, pady=2, sticky="ew")
        
        ttk.Button(file_frame, text="انتخاب فایل", 
                  command=self.select_excel_file).grid(row=0, column=2, padx=5, pady=2)
        
        ttk.Label(file_frame, text="نام شیت:").grid(row=1, column=0, sticky="w", padx=5, pady=2)
        self.sheet_combo = ttk.Combobox(file_frame, state="readonly")
        self.sheet_combo.grid(row=1, column=1, sticky="w", padx=5, pady=2)
        
        ttk.Button(file_frame, text="بارگذاری شیت‌ها", 
                  command=self.load_sheets).grid(row=1, column=2, padx=5, pady=2)
        
        # Load data button
        ttk.Button(file_frame, text="بارگذاری داده‌ها", 
                  command=self.load_data, style="Accent.TButton").grid(row=2, column=1, pady=10)
        
        # Info display
        info_frame = ttk.LabelFrame(data_tab, text="اطلاعات فایل", padding="10")
        info_frame.pack(fill="both", expand=True, pady=5, padx=10)
        
        self.info_text = tk.Text(info_frame, height=15, font=("Arial", 10))
        scrollbar_v = ttk.Scrollbar(info_frame, orient="vertical", command=self.info_text.yview)
        scrollbar_h = ttk.Scrollbar(info_frame, orient="horizontal", command=self.info_text.xview)
        
        self.info_text.configure(yscrollcommand=scrollbar_v.set, xscrollcommand=scrollbar_h.set)
        
        self.info_text.grid(row=0, column=0, sticky="nsew")
        scrollbar_v.grid(row=0, column=1, sticky="ns")
        scrollbar_h.grid(row=1, column=0, sticky="ew")
        
        info_frame.columnconfigure(0, weight=1)
        info_frame.rowconfigure(0, weight=1)
    
    def _create_simple_filter_tab(self):
        """تب فیلترهای ساده"""
        filter_tab = ttk.Frame(self.notebook)
        self.notebook.add(filter_tab, text="فیلتر ساده")
        
        filter_frame = ttk.LabelFrame(filter_tab, text="فیلترهای ساده", padding="10")
        filter_frame.pack(fill="x", pady=5, padx=10)
        
        filter_frame.columnconfigure(1, weight=1)
        
        # Repair type filter
        ttk.Label(filter_frame, text="نوع تعمیر:").grid(row=0, column=0, sticky="w", padx=5, pady=5)
        self.repair_combo = ttk.Combobox(filter_frame, state="readonly")
        self.repair_combo.grid(row=0, column=1, sticky="ew", padx=5, pady=5)
        
        # Part filter
        ttk.Label(filter_frame, text="قالب/قطعه/دستگاه:").grid(row=1, column=0, sticky="w", padx=5, pady=5)
        self.part_combo = ttk.Combobox(filter_frame, state="readonly")
        self.part_combo.grid(row=1, column=1, sticky="ew", padx=5, pady=5)
        
        # Filter buttons
        button_frame = ttk.Frame(filter_frame)
        button_frame.grid(row=2, column=0, columnspan=2, pady=10)
        
        ttk.Button(button_frame, text="اعمال فیلتر", 
                  command=self.apply_simple_filter).pack(side="left", padx=5)
        
        ttk.Button(button_frame, text="نمایش همه داده‌ها", 
                  command=self.show_all_data).pack(side="left", padx=5)
    
    def _create_advanced_filter_tab(self):
        """تب فیلترهای پیشرفته"""
        advanced_tab = ttk.Frame(self.notebook)
        self.notebook.add(advanced_tab, text="فیلتر پیشرفته")
        
        advanced_frame = ttk.LabelFrame(advanced_tab, text="فیلترهای پیشرفته", padding="10")
        advanced_frame.pack(fill="x", pady=5, padx=10)
        
        # Multiple repair types
        ttk.Label(advanced_frame, text="انواع تعمیر (چندگانه):").grid(row=0, column=0, sticky="w", padx=5, pady=5)
        self.repair_listbox = tk.Listbox(advanced_frame, selectmode=tk.MULTIPLE, height=4)
        self.repair_listbox.grid(row=0, column=1, sticky="ew", padx=5, pady=5)
        
        # Hour range
        ttk.Label(advanced_frame, text="بازه ساعت کاری:").grid(row=1, column=0, sticky="w", padx=5, pady=5)
        
        hour_frame = ttk.Frame(advanced_frame)
        hour_frame.grid(row=1, column=1, sticky="ew", padx=5, pady=5)
        
        ttk.Label(hour_frame, text="از:").pack(side="left")
        self.hour_min_entry = ttk.Entry(hour_frame, width=8)
        self.hour_min_entry.pack(side="left", padx=2)
        
        ttk.Label(hour_frame, text="تا:").pack(side="left", padx=(10, 0))
        self.hour_max_entry = ttk.Entry(hour_frame, width=8)
        self.hour_max_entry.pack(side="left", padx=2)
        
        # Advanced filter buttons
        adv_button_frame = ttk.Frame(advanced_frame)
        adv_button_frame.grid(row=2, column=0, columnspan=2, pady=10)
        
        ttk.Button(adv_button_frame, text="اعمال فیلتر پیشرفته", 
                  command=self.apply_advanced_filter).pack(side="left", padx=5)
        
        ttk.Button(adv_button_frame, text="گروه‌بندی داده‌ها", 
                  command=self.group_data).pack(side="left", padx=5)
    
    def _create_results_tab(self):
        """تب نتایج"""
        results_tab = ttk.Frame(self.notebook)
        self.notebook.add(results_tab, text="نتایج و خروجی")
        
        # Results display
        results_frame = ttk.LabelFrame(results_tab, text="داده‌های فیلتر شده", padding="10")
        results_frame.pack(fill="both", expand=True, pady=5, padx=10)
        
        # Create treeview
        self.tree = ttk.Treeview(results_frame, show="headings", height=15)
        
        tree_scrollbar_v = ttk.Scrollbar(results_frame, orient="vertical", command=self.tree.yview)
        tree_scrollbar_h = ttk.Scrollbar(results_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=tree_scrollbar_v.set, xscrollcommand=tree_scrollbar_h.set)
        
        self.tree.grid(row=0, column=0, sticky="nsew")
        tree_scrollbar_v.grid(row=0, column=1, sticky="ns")
        tree_scrollbar_h.grid(row=1, column=0, sticky="ew")
        
        results_frame.columnconfigure(0, weight=1)
        results_frame.rowconfigure(0, weight=1)
        
        # Export buttons
        export_frame = ttk.Frame(results_tab)
        export_frame.pack(fill="x", pady=10, padx=10)
        
        ttk.Button(export_frame, text="ذخیره در Excel", 
                  command=self.export_to_excel).pack(side="left", padx=5)
        
        ttk.Button(export_frame, text="ذخیره در CSV", 
                  command=self.export_to_csv).pack(side="left", padx=5)
        
        ttk.Button(export_frame, text="نمایش آمار", 
                  command=self.show_statistics).pack(side="left", padx=5)
    
    def select_excel_file(self):
        """انتخاب فایل اکسل"""
        file_path = filedialog.askopenfilename(
            title="انتخاب فایل اکسل",
            filetypes=[("Excel Files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if file_path:
            self.file_entry.delete(0, tk.END)
            self.file_entry.insert(0, file_path)
            self.status_var.set(f"فایل انتخاب شد: {os.path.basename(file_path)}")
    
    def load_sheets(self):
        """بارگذاری لیست شیت‌ها"""
        file_path = self.file_entry.get().strip()
        if not file_path or not os.path.exists(file_path):
            messagebox.showerror("خطا", "لطفاً ابتدا یک فایل معتبر انتخاب کنید")
            return
        
        try:
            with load_workbook(file_path, read_only=True) as wb:
                sheet_names = wb.sheetnames
            
            self.sheet_combo['values'] = sheet_names
            if sheet_names:
                self.sheet_combo.set(sheet_names[0])
            
            self.status_var.set(f"{len(sheet_names)} شیت پیدا شد")
            
        except Exception as e:
            messagebox.showerror("خطا", f"خطا در بارگذاری شیت‌ها: {e}")
    
    def load_data(self):
        """بارگذاری داده‌ها"""
        file_path = self.file_entry.get().strip()
        sheet_name = self.sheet_combo.get().strip()
        
        if not file_path or not sheet_name:
            messagebox.showwarning("هشدار", "لطفاً فایل و شیت را انتخاب کنید")
            return
        
        try:
            success = self.excel_processor.load_excel(file_path, sheet_name)
            if success and self.excel_processor.df is not None and not self.excel_processor.df.empty:
                record_count = len(self.excel_processor.df)
                self.status_var.set(f"{record_count} رکورد با موفقیت بارگذاری شد")
                
                # پر کردن فیلترها
                self._populate_filters()
                self.show_all_data()
                self._show_basic_info()
                
                messagebox.showinfo("موفق", f"داده‌ها با موفقیت بارگذاری شدند ({record_count} رکورد)")
            else:
                messagebox.showerror("خطا", "خطا در بارگذاری داده‌ها")
                
        except Exception as e:
            messagebox.showerror("خطا", f"خطا در بارگذاری داده‌ها: {e}")
    
    def _populate_filters(self):
        """پر کردن فیلترها با داده‌ها"""
        if self.excel_processor.df_normalized is None:
            return
        
        # پر کردن فیلتر نوع تعمیر
        repair_col = self.excel_processor.column_mapping.get('repair_col')
        if repair_col and repair_col in self.excel_processor.df_normalized.columns:
            repair_types = ["(همه)"] + sorted(self.excel_processor.df_normalized[repair_col].dropna().unique().tolist())
            self.repair_combo['values'] = repair_types
            self.repair_combo.set("(همه)")
            
            # پر کردن لیست‌باکس برای فیلتر پیشرفته
            self.repair_listbox.delete(0, tk.END)
            for repair_type in repair_types:
                if repair_type != "(همه)":
                    self.repair_listbox.insert(tk.END, repair_type)
        
        # پر کردن فیلتر قالب/قطعه/دستگاه
        part_col = self.excel_processor.column_mapping.get('part_col')
        if part_col and part_col in self.excel_processor.df.columns:
            part_types = ["(همه)"] + sorted(self.excel_processor.df[part_col].dropna().astype(str).unique().tolist())
            self.part_combo['values'] = part_types
            self.part_combo.set("(همه)")
    
    def _show_basic_info(self):
        """نمایش اطلاعات پایه"""
        if self.excel_processor.df is None:
            return
        
        info = "="*60 + "\n"
        info += "اطلاعات کلی داده‌ها\n"
        info += "="*60 + "\n\n"
        
        info += f"تعداد رکوردها: {len(self.excel_processor.df):,}\n"
        info += f"تعداد ستون‌ها: {len(self.excel_processor.df.columns)}\n\n"
        
        info += "ستون‌های تشخیص داده شده:\n"
        for col_type, col_name in self.excel_processor.column_mapping.items():
            status = "✅" if col_name else "❌"
            info += f"  {status} {col_type}: {col_name if col_name else 'یافت نشد'}\n"
        
        self.info_text.delete(1.0, tk.END)
        self.info_text.insert(1.0, info)
    
    def apply_simple_filter(self):
        """اعمال فیلتر ساده"""
        if self.excel_processor.df is None:
            messagebox.showwarning("هشدار", "لطفاً ابتدا داده‌ها را بارگذاری کنید")
            return
        
        repair_type = self.repair_combo.get()
        part_type = self.part_combo.get()
        
        filtered_df = self.data_filter.apply_simple_filter(
            repair_type=repair_type,
            part_type=part_type
        )
        
        if filtered_df is not None and not filtered_df.empty:
            self._display_data_in_treeview(filtered_df)
            self.status_var.set(f"فیلتر اعمال شد - {len(filtered_df)} رکورد")
        else:
            messagebox.showwarning("هشدار", "هیچ داده‌ای با فیلترهای انتخاب شده یافت نشد")
    
    def apply_advanced_filter(self):
        """اعمال فیلتر پیشرفته"""
        if self.excel_processor.df is None:
            messagebox.showwarning("هشدار", "لطفاً ابتدا داده‌ها را بارگذاری کنید")
            return
        
        selected_repairs = [self.repair_listbox.get(i) for i in self.repair_listbox.curselection()]
        
        hour_min = self.hour_min_entry.get().strip()
        hour_max = self.hour_max_entry.get().strip()
        
        hour_min_float = float(hour_min) if hour_min else None
        hour_max_float = float(hour_max) if hour_max else None
        
        filtered_df = self.data_filter.apply_advanced_filter(
            selected_repairs=selected_repairs,
            hour_min=hour_min_float,
            hour_max=hour_max_float
        )
        
        if filtered_df is not None and not filtered_df.empty:
            self._display_data_in_treeview(filtered_df)
            self.status_var.set(f"فیلتر پیشرفته اعمال شد - {len(filtered_df)} رکورد")
        else:
            messagebox.showwarning("هشدار", "هیچ داده‌ای با فیلترهای انتخاب شده یافت نشد")
    
    def group_data(self):
        """گروه‌بندی داده‌ها"""
        if self.excel_processor.df is None:
            messagebox.showwarning("هشدار", "لطفاً ابتدا داده‌ها را بارگذاری کنید")
            return
        
        grouped_df = self.data_filter.group_data()
        
        if grouped_df is not None and not grouped_df.empty:
            self._display_data_in_treeview(grouped_df)
            self.status_var.set(f"داده‌ها گروه‌بندی شد - {len(grouped_df)} گروه")
        else:
            messagebox.showwarning("هشدار", "خطا در گروه‌بندی داده‌ها")
    
    def show_all_data(self):
        """نمایش همه داده‌ها"""
        if self.excel_processor.df is not None and not self.excel_processor.df.empty:
            self._display_data_in_treeview(self.excel_processor.df)
            self.status_var.set(f"همه داده‌ها نمایش داده شد - {len(self.excel_processor.df)} رکورد")
    
    def _display_data_in_treeview(self, df):
        """نمایش داده‌ها در Treeview"""
        # پاک کردن داده‌های قبلی
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # تنظیم ستون‌ها
        columns = list(df.columns)
        self.tree["columns"] = columns
        
        # تنظیم هدرها
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=100)
        
        # نمایش داده‌ها (حداکثر 1000 رکورد)
        for _, row in df.head(1000).iterrows():
            values = [row[col] if pd.notna(row[col]) else "" for col in columns]
            self.tree.insert("", "end", values=values)
    
    def export_to_excel(self):
        """ذخیره در Excel"""
        if self.data_filter.filtered_data is None:
            messagebox.showwarning("هشدار", "هیچ داده‌ای برای ذخیره وجود ندارد")
            return
        
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        
        if file_path:
            success = ReportGenerator.save_to_excel(self.data_filter.filtered_data, file_path)
            if success:
                messagebox.showinfo("موفق", f"فایل با موفقیت ذخیره شد:\n{file_path}")
            else:
                messagebox.showerror("خطا", "خطا در ذخیره فایل")
    
    def export_to_csv(self):
        """ذخیره در CSV"""
        if self.data_filter.filtered_data is None:
            messagebox.showwarning("هشدار", "هیچ داده‌ای برای ذخیره وجود ندارد")
            return
        
        file_path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv")]
        )
        
        if file_path:
            success = ReportGenerator.save_to_csv(self.data_filter.filtered_data, file_path)
            if success:
                messagebox.showinfo("موفق", f"فایل با موفقیت ذخیره شد:\n{file_path}")
            else:
                messagebox.showerror("خطا", "خطا در ذخیره فایل")
    
    def show_statistics(self):
        """نمایش آمار"""
        if self.data_filter.filtered_data is None:
            messagebox.showwarning("هشدار", "هیچ داده‌ای برای نمایش آمار وجود ندارد")
            return
        
        df = self.data_filter.filtered_data
        
        stats = "="*60 + "\n"
        stats += "آمار داده‌ها\n"
        stats += "="*60 + "\n\n"
        
        stats += f"تعداد رکوردها: {len(df):,}\n"
        
        perf_col = self.excel_processor.column_mapping.get('perf_col')
        if perf_col and perf_col in df.columns:
            df[perf_col] = pd.to_numeric(df[perf_col], errors='coerce')
            total_hours = df[perf_col].sum()
            avg_hours = df[perf_col].mean()
            max_hours = df[perf_col].max()
            min_hours = df[perf_col].min()
            
            stats += f"\n📊 آمار ساعت کاری:\n"
            stats += f"  • مجموع ساعت: {total_hours:.2f}\n"
            stats += f"  • میانگین ساعت: {avg_hours:.2f}\n"
            stats += f"  • بیشترین ساعت: {max_hours:.2f}\n"
            stats += f"  • کمترین ساعت: {min_hours:.2f}\n"
        
        repair_col = self.excel_processor.column_mapping.get('repair_col')
        if repair_col and repair_col in df.columns:
            repair_stats = df[repair_col].value_counts()
            stats += f"\n🔧 توزیع انواع تعمیر:\n"
            for repair_type, count in repair_stats.head(10).items():
                stats += f"  • {repair_type}: {count} رکورد\n"
        
        messagebox.showinfo("آمار داده‌ها", stats)

def main():
    """تابع اصلی"""
    try:
        root = tk.Tk()
        app = CompleteReportApp(root)
        root.mainloop()
    except Exception as e:
        messagebox.showerror("خطای شدید", f"برنامه نمی‌تواند اجرا شود: {e}")

if __name__ == "__main__":
    main()
